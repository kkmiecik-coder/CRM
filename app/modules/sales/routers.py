# app/modules/sales/routers.py
"""
Sales Routes
======================

API endpoints i routes dla modułu Sales.

Element 1: Recruitment
- GET /sales/application - strona rekrutacyjna
- POST /sales/api/application/validate - walidacja pola
- POST /sales/api/application/submit - wysłanie aplikacji
- POST /sales/api/application/generate-nda - generowanie PDF
- POST /sales/api/application/check-email - sprawdzenie email

Admin Panel:
- GET /sales/ - panel administracyjny
- GET /sales/admin/api/stats - statystyki
- GET /sales/admin/api/applications - lista aplikacji z filtrowaniem
- GET /sales/admin/api/application/<id> - szczegóły aplikacji
- POST /sales/admin/api/application/<id>/status - zmiana statusu
- POST /sales/admin/api/application/<id>/note - dodanie notatki
- GET /sales/admin/api/application/<id>/nda - pobieranie pliku NDA
- GET /sales/admin/api/export - eksport do XLSX
- GET /sales/admin/api/export-applications - eksport aplikacji do XLSX

Autor: Development Team
Data: 2025-10-24
"""

from flask import render_template, request, jsonify, current_app, send_file, session, redirect, url_for, flash, make_response
from modules.sales import sales_bp
from modules.sales.services import ApplicationService, EmailService
from modules.sales.validators import validate_application_data, validate_file
from modules.sales.utils import rate_limit, generate_nda_pdf
from modules.users.decorators import require_module_access
from extensions import db
from modules.sales.models import SalesApplication
import io
import os
import sys
from datetime import datetime, timedelta

from sqlalchemy import func, or_, desc
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from functools import wraps

def json_response(data, status=200):
    """Helper do tworzenia JSON response z właściwym Content-Type"""
    response = make_response(jsonify(data), status)
    response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response


# ============================================================================
# ELEMENT 1: RECRUITMENT - VIEWS
# ============================================================================

@sales_bp.route('/application')
def recruitment():
    """Strona rekrutacyjna (Element 1)"""
    return render_template('sales_recruitment.html')


# ============================================================================
# API ENDPOINTS - ELEMENT 1: RECRUITMENT
# ============================================================================

@sales_bp.route('/api/application/validate', methods=['POST'])
def validate_application_field():
    """
    Walidacja pojedynczego pola formularza (AJAX)
    
    Request JSON:
    {
        "field_name": "email",
        "field_value": "test@example.com"
    }
    
    Response:
    {
        "success": true,
        "valid": true,
        "error": null
    }
    """
    try:
        data = request.get_json()
        field_name = data.get('field_name')
        field_value = data.get('field_value')
        
        if not field_name:
            return jsonify({
                'success': False,
                'error': 'Brak nazwy pola'
            }), 400
        
        # Waliduj pojedyncze pole
        temp_form_data = {field_name: field_value}
        is_valid, errors = validate_application_data(temp_form_data)
        
        field_error = errors.get(field_name)
        
        return jsonify({
            'success': True,
            'valid': field_error is None,
            'error': field_error
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Validation error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Błąd walidacji'
        }), 500


@sales_bp.route('/api/application/submit', methods=['POST'])
def submit_application():
    """
    Wysłanie kompletnego formularza aplikacyjnego z plikiem NDA
    
    Request (multipart/form-data):
    - Wszystkie pola formularza
    - nda_file: plik NDA
    
    Response:
    {
        "success": true,
        "message": "Aplikacja została wysłana",
        "application_id": 123
    }
    """
    try:
        # Pobierz dane z formularza
        form_data = request.form.to_dict()
        
        # Pobierz plik NDA
        nda_file = request.files.get('nda_file')
        
        current_app.logger.info(f"Received application from: {form_data.get('email')}")
        
        # ========================================================================
        # WALIDACJA DANYCH FORMULARZA
        # ========================================================================
        
        is_valid, errors = validate_application_data(form_data)

        if not is_valid:
            current_app.logger.warning(f"Validation errors: {errors}")
            return json_response({
                'success': False,
                'errors': errors
            }, 400)
        
        # ========================================================================
        # WALIDACJA PLIKU NDA
        # ========================================================================
        
        file_valid, file_error = validate_file(nda_file)
        
        if not file_valid:
            current_app.logger.warning(f"File validation error: {file_error}")
            return json_response({
                'success': False,
                'error': file_error
            }, 400)
        
        # ========================================================================
        # SPRAWDŹ CZY EMAIL JUŻ ISTNIEJE
        # ========================================================================
        
        existing_application = ApplicationService.get_application_by_email(
            form_data['email']
        )
        
        if existing_application:
            current_app.logger.warning(
                f"Duplicate application attempt: {form_data['email']}"
            )
            return json_response({
                'success': False,
                'error': 'Aplikacja z tym adresem email już istnieje'
            }, 400)
        
        # ========================================================================
        # UTWÓRZ APLIKACJĘ
        # ========================================================================
        
        # Pobierz IP i user agent
        ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ip_address and ',' in ip_address:
            ip_address = ip_address.split(',')[0].strip()
        
        user_agent = request.headers.get('User-Agent', '')
        
        # Utwórz aplikację w bazie
        application = ApplicationService.create_application(
            form_data=form_data,
            file=nda_file,
            ip_address=ip_address,
            user_agent=user_agent
        )
        
        current_app.logger.info(
            f"Application created successfully: ID={application.id}, Email={application.email}"
        )
        
        # ========================================================================
        # WYSYŁKA EMAILI
        # ========================================================================
        
        try:
            # Email do kandydata
            EmailService.send_application_confirmation(application)
            
            # Email do admina
            EmailService.send_admin_notification(application)
            
        except Exception as email_error:
            current_app.logger.error(f"Email sending failed: {str(email_error)}")
            # Nie przerywamy procesu jeśli email się nie wyśle
        
        # ========================================================================
        # RESPONSE
        # ========================================================================
        
        return json_response({
            'success': True,
            'message': 'Aplikacja została wysłana pomyślnie',
            'application_id': application.id
        }, 201)
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Application submission error: {str(e)}", exc_info=True)
        
        return json_response({
            'success': False,
            'error': 'Wystąpił błąd podczas przetwarzania aplikacji. Spróbuj ponownie.'
        }, 500)


@sales_bp.route('/api/application/generate-nda', methods=['POST'])
def generate_nda():
    """
    Generowanie PDF z umową NDA na podstawie danych z formularza
    
    Request JSON:
    {
        "first_name": "Jan",
        "last_name": "Kowalski",
        "email": "jan@example.com",
        "city": "Warszawa",
        ... (wszystkie dane formularza)
    }
    
    Response:
    - PDF file (application/pdf)
    """
    try:
        data = request.get_json()
        
        # Waliduj podstawowe dane
        required_fields = ['first_name', 'last_name', 'email', 'city']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                'success': False,
                'error': f'Brak wymaganych pól: {", ".join(missing_fields)}'
            }), 400
        
        current_app.logger.info(f"Generating NDA for: {data.get('email')}")
        
        # Generuj PDF - zwróci surowe bajty
        pdf_bytes = generate_nda_pdf(data)
        
        if not pdf_bytes:
            return jsonify({
                'success': False,
                'error': 'Nie udało się wygenerować PDF'
            }), 500
        
        # Przygotuj nazwę pliku
        filename = f"NDA_{data['last_name']}_{data['first_name']}.pdf"
        
        current_app.logger.info(f"NDA generated successfully: {filename}")
        
        # Zwróć PDF używając Response z surowymi bajtami
        from flask import Response
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        current_app.logger.error(f"NDA generation error: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Wystąpił błąd podczas generowania PDF'
        }), 500


@sales_bp.route('/api/application/check-email', methods=['POST'])
def check_email_exists():
    """
    Sprawdź czy email już istnieje w bazie (pomocnicze API)
    
    Request JSON:
    {
        "email": "test@example.com"
    }
    
    Response:
    {
        "success": true,
        "exists": false
    }
    """
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({
                'success': False,
                'error': 'Brak adresu email'
            }), 400
        
        existing = ApplicationService.get_application_by_email(email)
        
        return jsonify({
            'success': True,
            'exists': existing is not None
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Email check error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Błąd sprawdzania email'
        }), 500


# ============================================================================
# ADMIN PANEL - VIEWS
# ============================================================================

@sales_bp.route('/admin')
@require_module_access('sales')
def admin_dashboard():
    """Strona główna panelu admina"""
    return render_template('sales_admin.html')


# ============================================================================
# ADMIN PANEL - API ENDPOINTS
# ============================================================================

@sales_bp.route('/admin/api/stats')
@require_module_access('sales')
def get_admin_stats():
    """Statystyki aplikacji dla dashboardu"""

    try:
        # ============================================================================
        # STATYSTYKI APLIKACJI REKRUTACYJNYCH
        # ============================================================================
        total_applications = SalesApplication.query.count()
        pending_count = SalesApplication.query.filter_by(status='pending').count()
        contacted_count = SalesApplication.query.filter_by(status='contacted').count()
        accepted_count = SalesApplication.query.filter_by(status='accepted').count()
        rejected_count = SalesApplication.query.filter_by(status='rejected').count()
        
        return jsonify({
            'success': True,
            'data': {
                # Statystyki aplikacji
                'total_applications': total_applications,
                'pending_count': pending_count,
                'contacted_count': contacted_count,
                'accepted_count': accepted_count,
                'rejected_count': rejected_count
            }
        }), 200
    except Exception as e:
        current_app.logger.error(f"Admin stats error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania statystyk'}), 500


@sales_bp.route('/admin/api/applications')
@require_module_access('sales')
def get_admin_applications():
    """Lista aplikacji z filtrowaniem i paginacją"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '')
        is_b2b = request.args.get('is_b2b', '')
        
        query = SalesApplication.query
        
        # Filtrowanie po statusie
        if status_filter:
            query = query.filter_by(status=status_filter)
        
        # Filtrowanie B2B/B2C
        if is_b2b == 'true':
            query = query.filter_by(is_b2b=True)
        elif is_b2b == 'false':
            query = query.filter_by(is_b2b=False)
        
        # Wyszukiwanie
        if search:
            pattern = f'%{search}%'
            query = query.filter(
                or_(
                    SalesApplication.first_name.ilike(pattern),
                    SalesApplication.last_name.ilike(pattern),
                    SalesApplication.email.ilike(pattern),
                    SalesApplication.phone.ilike(pattern),
                    SalesApplication.company_name.ilike(pattern),
                    SalesApplication.nip.ilike(pattern)
                )
            )
        
        # Paginacja
        pagination = query.order_by(desc(SalesApplication.created_at)).paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        # Formatowanie wyników
        applications = []
        for app in pagination.items:
            app_data = {
                'id': app.id,
                'first_name': app.first_name,
                'last_name': app.last_name,
                'email': app.email,
                'phone': app.phone,
                'city': app.city,
                'address': app.address,
                'postal_code': app.postal_code,
                'voivodeship': app.voivodeship,
                'business_location': app.business_location,
                'status': app.status,
                'is_b2b': app.is_b2b,
                'created_at': app.created_at.strftime('%Y-%m-%d %H:%M') if app.created_at else None,
                'has_nda_file': bool(app.nda_filepath)
            }
            
            # Dodaj dane B2B jeśli istnieją
            if app.is_b2b:
                app_data['company_name'] = app.company_name
                app_data['nip'] = app.nip
            
            applications.append(app_data)
        
        return jsonify({
            'success': True,
            'data': {
                'applications': applications,
                'pagination': {
                    'page': pagination.page,
                    'per_page': pagination.per_page,
                    'total': pagination.total,
                    'pages': pagination.pages
                }
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Admin applications list error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania listy aplikacji'}), 500


@sales_bp.route('/admin/api/application/<int:application_id>')
@require_module_access('sales')
def get_admin_application_detail(application_id):
    """Szczegóły pojedynczej aplikacji"""
    try:
        app = SalesApplication.query.get_or_404(application_id)
        
        # Podstawowe dane
        detail = {
            'id': app.id,
            'first_name': app.first_name,
            'last_name': app.last_name,
            'email': app.email,
            'phone': app.phone,
            'city': app.city,
            'address': app.address,
            'postal_code': app.postal_code,
            'voivodeship': app.voivodeship,
            'business_location': app.business_location,
            'about_text': app.about_text,
            'status': app.status,
            'is_b2b': app.is_b2b,
            'data_processing_consent': app.data_processing_consent,
            'created_at': app.created_at.strftime('%Y-%m-%d %H:%M:%S') if app.created_at else None,
            'updated_at': app.updated_at.strftime('%Y-%m-%d %H:%M:%S') if app.updated_at else None,
            'ip_address': app.ip_address,
            'user_agent': app.user_agent,
            'notes': json.loads(app.notes) if app.notes else [],
            'has_nda_file': bool(app.nda_filepath)
        }
        
        # Dane NDA
        if app.nda_filepath:
            detail['nda_filename'] = app.nda_filename
            detail['nda_filesize'] = app.nda_filesize
            detail['nda_mime_type'] = app.nda_mime_type
        
        # Dane B2B
        if app.is_b2b:
            detail['company_name'] = app.company_name
            detail['nip'] = app.nip
            detail['regon'] = app.regon
            detail['company_address'] = app.company_address
            detail['company_city'] = app.company_city
            detail['company_postal_code'] = app.company_postal_code
        
        return jsonify({
            'success': True,
            'data': detail
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Admin application detail error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania szczegółów aplikacji'}), 500


@sales_bp.route('/admin/api/application/<int:application_id>/status', methods=['POST'])
@require_module_access('sales')
def update_application_status(application_id):
    """Zmiana statusu aplikacji"""
    try:
        data = request.get_json()
        new_status = data.get('status')
        notes = data.get('notes')
        
        if not new_status:
            return jsonify({'success': False, 'message': 'Brak nowego statusu'}), 400
        
        application = ApplicationService.update_application_status(
            application_id=application_id,
            new_status=new_status,
            notes=notes
        )
        
        # Wyślij email do kandydata o zmianie statusu
        try:
            EmailService.send_status_update(application, new_status)
        except Exception as email_error:
            current_app.logger.error(f"Email sending failed: {str(email_error)}")
        
        return jsonify({
            'success': True,
            'message': 'Status zaktualizowany',
            'new_status': application.status
        }), 200
        
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Admin status update error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd aktualizacji statusu'}), 500


@sales_bp.route('/admin/api/application/<int:application_id>/note', methods=['POST'])
@require_module_access('sales')
def add_application_note(application_id):
    """Dodanie notatki do aplikacji"""
    try:
        data = request.get_json()
        note_text = data.get('note')
        
        if not note_text:
            return jsonify({'success': False, 'message': 'Brak treści notatki'}), 400
        
        app = SalesApplication.query.get_or_404(application_id)
        
        # Dodaj notatkę
        current_notes = json.loads(app.notes) if app.notes else []
        new_note = {
            'timestamp': datetime.utcnow().isoformat(),
            'author': session.get('user_email', 'admin'),
            'text': note_text
        }
        current_notes.append(new_note)
        
        app.notes = json.dumps(current_notes)
        app.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Notatka dodana', 'note': new_note}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Admin note add error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd dodawania notatki'}), 500


@sales_bp.route('/admin/api/application/<int:application_id>/nda')
@require_module_access('sales')
def download_nda(application_id):
    """Pobieranie pliku NDA"""
    try:
        filepath = ApplicationService.get_nda_file_path(application_id)
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'Plik NDA nie istnieje'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath)
        )
        
    except Exception as e:
        current_app.logger.error(f"NDA download error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania pliku'}), 500


@sales_bp.route('/admin/api/export')
@require_module_access('sales')
def export_admin_applications():
    """Eksport aplikacji do XLSX z wszystkimi polami"""
    try:
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '')
        is_b2b = request.args.get('is_b2b', '')
        
        # Buduj query
        query = SalesApplication.query
        
        if status_filter:
            query = query.filter_by(status=status_filter)
        
        if is_b2b == 'true':
            query = query.filter_by(is_b2b=True)
        elif is_b2b == 'false':
            query = query.filter_by(is_b2b=False)
        
        if search:
            pattern = f'%{search}%'
            query = query.filter(
                or_(
                    SalesApplication.first_name.ilike(pattern),
                    SalesApplication.last_name.ilike(pattern),
                    SalesApplication.email.ilike(pattern),
                    SalesApplication.phone.ilike(pattern),
                    SalesApplication.company_name.ilike(pattern),
                    SalesApplication.nip.ilike(pattern)
                )
            )
        
        applications = query.order_by(desc(SalesApplication.created_at)).all()
        
        # Tworzenie workbooka
        wb = Workbook()
        ws = wb.active
        ws.title = "Aplikacje Sales"
        
        # Nagłówki - nowa kolejność według specyfikacji
        headers = [
            'Data aplikacji', 'Imię', 'Nazwisko', 'Typ', 'Województwo',
            'Miejscowość działalności', 'Źródło',
            'E-mail', 'Telefon', 'Adres', 'Miasto', 'Kod pocztowy',
            'O sobie', 'Firma', 'NIP', 'REGON', 'Adres firmy',
            'Miasto firmy', 'Kod pocztowy firmy'
        ]
        
        # Stylizacja nagłówków
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Funkcja pomocnicza do formatowania telefonu
        def format_phone(phone_str):
            """
            Formatuje numer telefonu do formatu: 123 456 789
            Usuwa prefixy +48 lub 48 oraz wszystkie znaki specjalne
            """
            if not phone_str:
                return ''
            
            # Usuń wszystkie znaki nie będące cyframi
            digits = ''.join(filter(str.isdigit, phone_str))
            
            # Usuń prefix 48 jeśli występuje na początku
            if digits.startswith('48') and len(digits) > 9:
                digits = digits[2:]
            
            # Jeśli mamy dokładnie 9 cyfr, formatuj jako XXX XXX XXX
            if len(digits) == 9:
                return f'{digits[0:3]} {digits[3:6]} {digits[6:9]}'
            
            # Jeśli inna liczba cyfr, zwróć jak jest
            return digits
        
        # Funkcja pomocnicza do formatowania tekstu (pierwsza litera duża, reszta mała)
        def format_text(text):
            """
            Formatuje tekst: pierwsza litera duża, reszta małe
            Np. "JAN KOWALSKI" -> "Jan kowalski", "WARSZAWA" -> "Warszawa"
            """
            if not text:
                return ''
            return text.strip().capitalize()
        
        # Wypełnianie danymi
        for row_num, app in enumerate(applications, 2):
            # Kolumna 1: Data aplikacji (format YYYY-MM-DD)
            ws.cell(row=row_num, column=1, value=app.created_at.strftime('%Y-%m-%d') if app.created_at else '')
            
            # Kolumna 2: Imię (sformatowane)
            ws.cell(row=row_num, column=2, value=format_text(app.first_name))
            
            # Kolumna 3: Nazwisko (sformatowane)
            ws.cell(row=row_num, column=3, value=format_text(app.last_name))
            
            # Kolumna 4: Typ (B2B lub Zlecenie zamiast B2C)
            ws.cell(row=row_num, column=4, value='B2B' if app.is_b2b else 'Zlecenie')
            
            # Kolumna 5: Województwo (sformatowane)
            ws.cell(row=row_num, column=5, value=format_text(app.voivodeship))
            
            # Kolumna 6: Miejscowość działalności (sformatowane)
            ws.cell(row=row_num, column=6, value=format_text(app.business_location))
            
            # Kolumna 7: Źródło (zawsze "Sales")
            ws.cell(row=row_num, column=7, value='Sales')
            
            # Kolumna 8: E-mail
            ws.cell(row=row_num, column=8, value=app.email)
            
            # Kolumna 9: Telefon (sformatowany)
            ws.cell(row=row_num, column=9, value=format_phone(app.phone))
            
            # Kolumna 10: Adres (sformatowany)
            ws.cell(row=row_num, column=10, value=format_text(app.address))
            
            # Kolumna 11: Miasto (sformatowane)
            ws.cell(row=row_num, column=11, value=format_text(app.city))
            
            # Kolumna 12: Kod pocztowy
            ws.cell(row=row_num, column=12, value=app.postal_code)
            
            # Kolumna 13: O sobie
            ws.cell(row=row_num, column=13, value=app.about_text or '')
            
            # Kolumna 14: Firma (tylko dla B2B)
            ws.cell(row=row_num, column=14, value=app.company_name if app.is_b2b else '')
            
            # Kolumna 15: NIP (tylko dla B2B)
            ws.cell(row=row_num, column=15, value=app.nip if app.is_b2b else '')
            
            # Kolumna 16: REGON (tylko dla B2B)
            ws.cell(row=row_num, column=16, value=app.regon if app.is_b2b else '')
            
            # Kolumna 17: Adres firmy (tylko dla B2B)
            ws.cell(row=row_num, column=17, value=app.company_address if app.is_b2b else '')
            
            # Kolumna 18: Miasto firmy (tylko dla B2B, sformatowane)
            ws.cell(row=row_num, column=18, value=format_text(app.company_city) if app.is_b2b else '')
            
            # Kolumna 19: Kod pocztowy firmy (tylko dla B2B)
            ws.cell(row=row_num, column=19, value=app.company_postal_code if app.is_b2b else '')
        
        # Autosize kolumn
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Zapisz do bufora
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Przygotuj nazwę pliku
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'aplikacje_sales_{timestamp}.xlsx'
        
        # ============================================================================
        # POPRAWKA: Użyj make_response() zamiast send_file() dla BytesIO w Passenger
        # ============================================================================
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Export error: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Błąd eksportu: {str(e)}'}), 500


@sales_bp.route('/admin/api/export-applications')
@require_module_access('sales')
def export_applications_xlsx():
    """Eksport aplikacji do pliku XLSX (alias dla /admin/api/export)"""
    return export_admin_applications()