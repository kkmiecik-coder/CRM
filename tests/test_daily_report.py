# -*- coding: utf-8 -*-
"""
Agregat i eksport dziennego raportu produkcji.

Testy pilnują tego, na czym raport stoi i co już raz wywróciło się w projekcie:
filtr auto_skip/system, cofnięcia niewidoczne w netto, praca bez atrybucji,
oraz różnica między pustą komórką a zerem.
"""
import io
import os
import sys
from datetime import date, datetime, time, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from flask import Flask
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool

from extensions import db
from modules.production.models import (
    ProductionConfig, ProductionConfiguration, ProductionDevice, ProductionOrder,
    ProductionProduct, ProductionReworkLog, ProductionStationEvent,
    ProductionStationEventWorker, ProductionWorker, ProductionWorkerSession,
)
from modules.production.sawmill.models import (
    SawmillAudit, SawmillCounter, SawmillDelivery, SawmillLog, SawmillOrder,
    SawmillSpecies, SawmillSupplier,
)
from modules.production.services import daily_report_export, daily_report_service
from modules.users.models import User
from modules.calculator.models import Multiplier  # noqa: F401
from modules.clients.models import Client  # noqa: F401
import modules.quotes.models  # noqa: F401

_TABLES = [m.__table__ for m in (
    User, ProductionDevice, ProductionConfig, ProductionOrder, ProductionProduct,
    ProductionConfiguration, ProductionWorker, ProductionWorkerSession,
    ProductionStationEvent, ProductionStationEventWorker, ProductionReworkLog,
    SawmillSupplier, SawmillSpecies, SawmillCounter, SawmillDelivery,
    SawmillOrder, SawmillLog, SawmillAudit,
)]

# shipping_label_base64 jest typem MySQL-owym (LONGTEXT) — SQLite go nie zna.
# Ta sama podmiana co w tests/test_reports_service.py:51.
ProductionOrder.__table__.c.shipping_label_base64.type = db.Text()

# Data odniesienia. Stała, nie get_local_now(): raport kubełkuje po dobach,
# więc test liczony „od dziś" przechodziłby przez inne gałęzie w poniedziałek
# niż w sobotę.
PONIEDZIALEK = date(2026, 8, 10)


@pytest.fixture()
def app():
    app = Flask(__name__)
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite://'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'poolclass': StaticPool,
        'connect_args': {'check_same_thread': False},
    }
    db.init_app(app)
    with app.app_context():
        db.metadata.create_all(bind=db.engine, tables=_TABLES)
        yield app
        db.session.remove()


# ============================================================================
# POMOCNICZE
# ============================================================================

_licznik_zamowien = [0]


def _produkt(status='czeka_na_sklejanie', volume=0.5, quantity=10,
             deadline=None, wartosc=1000.0):
    """Zamówienie + pozycja. Własny licznik, bo baselinker_order_id jest UNIQUE."""
    _licznik_zamowien[0] += 1
    numer = _licznik_zamowien[0]
    order = ProductionOrder(baselinker_order_id=numer,
                            internal_order_number=f'26/{numer:05d}',
                            client_name='Klient Testowy')
    db.session.add(order)
    db.session.flush()
    produkt = ProductionProduct(
        order_id=order.id, short_product_id=f'26{numer:03d}_1',
        product_sequence_in_order=1,
        original_product_name='Blat', quantity=quantity, volume_m3=volume,
        total_value_net=wartosc,
        current_status=status, deadline_date=deadline,
        created_at=datetime.combine(PONIEDZIALEK, time(9, 0)))
    db.session.add(produkt)
    db.session.commit()
    return produkt


def _event(produkt, station, delta, kiedy, source='mobile', worker=None):
    ev = ProductionStationEvent(
        production_item_id=produkt.id, station_code=station, delta=delta,
        quantity_done_after=max(0, delta), created_at=kiedy, source=source)
    db.session.add(ev)
    db.session.flush()
    if worker is not None:
        db.session.add(ProductionStationEventWorker(
            event_id=ev.id, worker_id=worker.id, share=1.0))
    db.session.commit()
    return ev


def _pracownik(imie='Adam'):
    w = ProductionWorker(first_name=imie, last_name='Nowak')
    db.session.add(w)
    db.session.commit()
    return w


def _sesja(pracownik, station, dzien, od=time(8, 0), godzin=8):
    start = datetime.combine(dzien, od)
    sesja = ProductionWorkerSession(
        worker_id=pracownik.id, station_code=station,
        session_group=f'g-{station}-{dzien}',
        started_at=start, last_activity_at=start,
        ended_at=start + timedelta(hours=godzin), work_date=dzien)
    db.session.add(sesja)
    db.session.commit()
    return sesja


def _stanowisko(dane, kod):
    """Wiersz stanowiska z wyniku zbierz_dane()."""
    return next(s for s in dane['stanowiska'] if s['kod'] == kod)


def _wiersz_sumy(ws):
    """
    Numer wiersza SUMA w arkuszu „Stanowiska".

    Szukamy po etykiecie, a nie przez ws.max_row: pod tabelą stoi jeszcze
    przypis wyjaśniający sumę zamówień, więc ostatni wiersz arkusza to nie
    jest wiersz sumy.
    """
    for numer in range(ws.max_row, 1, -1):
        if ws.cell(row=numer, column=1).value == 'SUMA':
            return numer
    raise AssertionError('arkusz nie ma wiersza SUMA')


# ============================================================================
# PRZERÓB PER STANOWISKO
# ============================================================================

def test_przerob_per_stanowisko(app):
    with app.app_context():
        p = _produkt()
        _event(p, 'gluing', 6, datetime.combine(PONIEDZIALEK, time(10, 0)))

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        gluing = _stanowisko(dane, 'gluing')
        assert gluing['sztuki'] == 6
        assert gluing['m3'] == pytest.approx(3.0)          # 6 × 0.5
        assert gluing['wartosc_netto'] == pytest.approx(600.0)  # 1000 × 6/10


def test_zdarzenia_automatu_nie_licza_sie_do_przerobu(app):
    """
    complete_task() generuje eventy dla stanowisk POMIJANYCH (cut_to_size=False
    przeskakuje formatowanie i wykańczanie). Bez filtra formatowanie dostaje
    sztuki, których nikt nie tknął.
    """
    with app.app_context():
        p = _produkt()
        _event(p, 'formatting', 10, datetime.combine(PONIEDZIALEK, time(11, 0)),
               source='auto_skip')
        _event(p, 'finishing', 10, datetime.combine(PONIEDZIALEK, time(11, 0)),
               source='system')

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        assert _stanowisko(dane, 'formatting')['sztuki'] == 0
        assert _stanowisko(dane, 'finishing')['sztuki'] == 0


def test_wszystkie_stanowiska_obecne_takze_bez_pracy(app):
    """
    Arkusz ma stały układ wierszy — stanowisko bez pracy pokazuje zera,
    a nie znika. Kolejność 1:1 z procesem produkcyjnym.
    """
    with app.app_context():
        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        kody = [s['kod'] for s in dane['stanowiska']]
        assert kody == ['cutting', 'assembly', 'gluing', 'formatting',
                        'finishing', 'painting', 'packaging']
        assert all(s['sztuki'] == 0 for s in dane['stanowiska'])


def test_praca_z_sasiedniego_dnia_nie_wchodzi(app):
    """Granica doby: 23:59:59 należy do dnia, 00:00:00 do następnego."""
    with app.app_context():
        p = _produkt()
        _event(p, 'gluing', 3, datetime.combine(PONIEDZIALEK, time(23, 59, 59)))
        _event(p, 'gluing', 5,
               datetime.combine(PONIEDZIALEK + timedelta(days=1), time(0, 0, 0)))

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        assert _stanowisko(dane, 'gluing')['sztuki'] == 3


# ============================================================================
# KOLEJKI, TERMINY, COFNIĘCIA
# ============================================================================

def test_cofniecia_liczone_osobno_od_netto(app):
    """
    Sztuki liczymy netto, więc dzień „10 zrobione, 4 cofnięte" i dzień
    „6 zrobione bez problemów" dają identyczne 6. Bez osobnej liczby cofnięć
    te dwa dni są w raporcie nieodróżnialne.
    """
    with app.app_context():
        p = _produkt()
        _event(p, 'gluing', 10, datetime.combine(PONIEDZIALEK, time(10, 0)))
        _event(p, 'gluing', -4, datetime.combine(PONIEDZIALEK, time(14, 0)))

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        gluing = _stanowisko(dane, 'gluing')
        assert gluing['sztuki'] == 6        # netto
        assert gluing['cofniecia'] == 4     # liczba dodatnia, nie -4


def test_kolejka_liczona_z_biezacego_statusu(app):
    """
    Definicja 1:1 z panelem (reports_service.dni_zapasu_stanowisk:313-321):
    pozycja czeka jako CAŁOŚĆ, z pełnym quantity, nawet jeśli jest w połowie
    zrobiona. Zawyża, i tak ma być — inaczej mail i dashboard pokazałyby dwie
    różne kolejki.
    """
    with app.app_context():
        _produkt(status='czeka_na_sklejanie', quantity=10, volume=0.5)
        _produkt(status='czeka_na_sklejanie', quantity=4, volume=0.25)
        _produkt(status='czeka_na_pakowanie', quantity=7, volume=0.1)

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        gluing = _stanowisko(dane, 'gluing')
        assert gluing['kolejka_szt'] == 14
        assert gluing['kolejka_m3'] == pytest.approx(6.0)   # 10×0.5 + 4×0.25

        assert _stanowisko(dane, 'packaging')['kolejka_szt'] == 7
        assert _stanowisko(dane, 'cutting')['kolejka_szt'] == 0


def test_koszyki_terminow(app):
    """Granice koszyków wg reports_service._koszyk_terminu — jedna definicja."""
    with app.app_context():
        _produkt(deadline=PONIEDZIALEK - timedelta(days=1))   # po terminie
        _produkt(deadline=PONIEDZIALEK)                        # dziś
        _produkt(deadline=PONIEDZIALEK + timedelta(days=2))    # 1-2 dni
        _produkt(deadline=PONIEDZIALEK + timedelta(days=7))    # 3-7 dni
        _produkt(deadline=PONIEDZIALEK + timedelta(days=8))    # 8+
        _produkt(deadline=None)                                # bez terminu

        terminy = daily_report_service.zbierz_dane(PONIEDZIALEK)['terminy']

        assert terminy == {'po_terminie': 1, 'dzis': 1, '1_2_dni': 1,
                           '3_7_dni': 1, '8_dni_plus': 1, 'bez_terminu': 1}


def test_pozycje_spakowane_nie_licza_sie_do_terminow(app):
    """
    Koszyki opisują to, co ZOSTAŁO do zrobienia. Pozycja spakowana albo
    anulowana nie jest zaległością i nie ma czego pilnować w jej terminie.
    """
    with app.app_context():
        _produkt(status='spakowane', deadline=PONIEDZIALEK - timedelta(days=5))
        _produkt(status='anulowane', deadline=PONIEDZIALEK - timedelta(days=5))
        _produkt(status='czeka_na_sklejanie', deadline=PONIEDZIALEK - timedelta(days=5))

        terminy = daily_report_service.zbierz_dane(PONIEDZIALEK)['terminy']

        assert terminy['po_terminie'] == 1


def test_wstrzymane_pozycje_licza_sie_do_terminow(app):
    """
    Wstrzymana pozycja po terminie MA być widoczna. Panel trzyma ją jako osobny
    segment wykresu „Termin vs postęp" (reports_service.py:120-124) — gdyby mail
    ją pomijał, to samo zamówienie byłoby po terminie na ekranie i nieobecne
    w raporcie.
    """
    with app.app_context():
        _produkt(status='wstrzymane', deadline=PONIEDZIALEK - timedelta(days=3))

        terminy = daily_report_service.zbierz_dane(PONIEDZIALEK)['terminy']

        assert terminy['po_terminie'] == 1


# ============================================================================
# LUDZIE I LICZBY ZBIORCZE
# ============================================================================

def test_wiersz_pracownika(app):
    with app.app_context():
        adam = _pracownik('Adam')
        _sesja(adam, 'gluing', PONIEDZIALEK, godzin=4)
        p = _produkt()
        _event(p, 'gluing', 8, datetime.combine(PONIEDZIALEK, time(10, 0)),
               worker=adam)

        ludzie = daily_report_service.zbierz_dane(PONIEDZIALEK)['ludzie']

        assert ludzie['osoby'] == 1
        wiersz = ludzie['wiersze'][0]
        assert wiersz['nazwa'] == 'Adam Nowak'
        assert wiersz['sztuki'] == pytest.approx(8.0)
        assert wiersz['zdarzenia'] == 1
        assert wiersz['godziny'] == pytest.approx(4.0)


def test_praca_bez_atrybucji_trafia_do_nieprzypisanych(app):
    """
    Bez tego wiersza suma z arkusza „Ludzie" nie zgadza się z sumą z arkusza
    „Stanowiska" i raport wygląda na zepsuty. Zasada z worker_stats_service:20-25.
    """
    with app.app_context():
        p = _produkt()
        _event(p, 'gluing', 5, datetime.combine(PONIEDZIALEK, time(10, 0)))

        ludzie = daily_report_service.zbierz_dane(PONIEDZIALEK)['ludzie']

        assert ludzie['wiersze'] == []
        assert ludzie['nieprzypisane']['sztuki'] == pytest.approx(5.0)
        assert ludzie['nieprzypisane']['m3'] == pytest.approx(2.5)


def test_tempo_bez_atrybucji_jest_puste_a_nie_zerowe(app):
    """
    Pracownik z sesją, ale bez ani jednej odbitej sztuki, ma dostać tempo
    „—", nie „0,0 m³/h". Zero przy nazwisku to zarzut bezczynności
    postawiony na podstawie braku danych.
    """
    with app.app_context():
        adam = _pracownik('Adam')
        _sesja(adam, 'gluing', PONIEDZIALEK, godzin=8)

        ludzie = daily_report_service.zbierz_dane(PONIEDZIALEK)['ludzie']

        wiersz = ludzie['wiersze'][0]
        assert wiersz['godziny'] == pytest.approx(8.0)
        assert wiersz['tempo'] is None


def test_wykonanie_zbiorcze_dnia(app):
    """Sumy nagłówkowe: sztuki, m³, wartość, pozycje, zamówienia, cofnięcia."""
    with app.app_context():
        p1 = _produkt(quantity=10, volume=0.5, wartosc=1000.0)
        p2 = _produkt(quantity=10, volume=0.2, wartosc=500.0)
        _event(p1, 'gluing', 6, datetime.combine(PONIEDZIALEK, time(10, 0)))
        _event(p2, 'packaging', 4, datetime.combine(PONIEDZIALEK, time(11, 0)))
        _event(p2, 'packaging', -1, datetime.combine(PONIEDZIALEK, time(15, 0)))

        wykonanie = daily_report_service.zbierz_dane(PONIEDZIALEK)['wykonanie']

        assert wykonanie['sztuki'] == 9              # 6 + 4 - 1
        assert wykonanie['m3'] == pytest.approx(3.6)  # 6×0.5 + 3×0.2
        assert wykonanie['wartosc_netto'] == pytest.approx(750.0)
        assert wykonanie['pozycje'] == 2
        assert wykonanie['zamowienia'] == 2
        assert wykonanie['cofniecia'] == 1


def test_zakonczone_liczy_wylacznie_pakowanie(app):
    """
    „Zakończone" to wynik dnia, więc bierze tylko pakowanie — jedyne
    stanowisko, na którym sztuka pada raz. Suma wszystkich stanowisk
    (`wykonanie`) liczy tę samą sztukę tyle razy, przez ile stanowisk
    przeszła, i pod nazwą wyniku dnia zawyżała kilkukrotnie.
    """
    with app.app_context():
        p = _produkt(quantity=10, volume=0.5, wartosc=1000.0)
        # Jedna i ta sama sztuka na trzech stanowiskach tego samego dnia.
        _event(p, 'gluing', 4, datetime.combine(PONIEDZIALEK, time(8, 0)))
        _event(p, 'formatting', 4, datetime.combine(PONIEDZIALEK, time(10, 0)))
        _event(p, 'packaging', 4, datetime.combine(PONIEDZIALEK, time(14, 0)))

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        assert dane['zakonczone']['sztuki'] == 4
        assert dane['zakonczone']['m3'] == pytest.approx(2.0)        # 4 × 0.5
        assert dane['zakonczone']['wartosc_netto'] == pytest.approx(400.0)
        # Suma stanowisk widzi te same 4 sztuki trzy razy — i tak ma być,
        # bo to miara ruchu na hali.
        assert dane['wykonanie']['sztuki'] == 12


def test_zakonczone_odejmuje_cofniecia_z_pakowania(app):
    """
    Cofnięcie z pakowania obniża wynik dnia, bo produkt wrócił na halę.
    Sztuki są netto (suma delt), a `cofniecia` idą osobno wyłącznie po to,
    żeby mail mógł je pokazać w nawiasie.
    """
    with app.app_context():
        p = _produkt(quantity=10, volume=0.5, wartosc=1000.0)
        _event(p, 'packaging', 5, datetime.combine(PONIEDZIALEK, time(11, 0)))
        _event(p, 'packaging', -2, datetime.combine(PONIEDZIALEK, time(16, 0)))

        zakonczone = daily_report_service.zbierz_dane(PONIEDZIALEK)['zakonczone']

        assert zakonczone['sztuki'] == 3
        assert zakonczone['cofniecia'] == 2


def test_dzien_bez_ruchu_daje_zera_a_nie_wyjatek(app):
    """Raport idzie także w dniu bez produkcji — brak maila ma znaczyć awarię."""
    with app.app_context():
        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        assert dane['wykonanie']['sztuki'] == 0
        assert dane['ludzie']['osoby'] == 0
        assert dane['ludzie']['wiersze'] == []


# ============================================================================
# TRAKOWNIA
# ============================================================================

_licznik_klod = [0]


def _zlecenie_trakowni():
    """
    Dostawca → gatunek → dostawa → zlecenie. Pełny łańcuch, bo SawmillOrder ma
    delivery_id i species_id jako NOT NULL. Wzór: tests/test_sawmill_orders.py:68-80.
    """
    from decimal import Decimal

    supplier = SawmillSupplier(name='Tartak Testowy')
    species = SawmillSpecies(name='Dąb')
    db.session.add_all([supplier, species])
    db.session.flush()
    delivery = SawmillDelivery(supplier_id=supplier.id,
                               delivery_date=PONIEDZIALEK)
    db.session.add(delivery)
    db.session.flush()
    order = SawmillOrder(order_number='TRK/2026/001', delivery_id=delivery.id,
                         species_id=species.id,
                         declared_volume_m3=Decimal('80.000'))
    db.session.add(order)
    db.session.commit()
    return order


def _kloda(zlecenie, volume=1.5, kiedy=None, usunieta=False):
    """
    Kłoda mierzona na tablecie. measured_at, nie created_at — patrz test niżej.

    order_id i sequence_no są NOT NULL, więc kłoda zawsze wisi przy zleceniu,
    a numer w sekwencji nadaje własny licznik (para order_id+sequence_no jest
    UNIQUE).
    """
    _licznik_klod[0] += 1
    log = SawmillLog(
        order_id=zlecenie.id, sequence_no=_licznik_klod[0],
        mid_circumference_cm=120.0, length_cm=400.0,
        volume_m3=volume, is_deleted=usunieta,
        measured_at=kiedy or datetime.combine(PONIEDZIALEK, time(10, 0)))
    db.session.add(log)
    db.session.commit()
    return log


def test_blok_trakowni(app):
    with app.app_context():
        zlecenie = _zlecenie_trakowni()
        _kloda(zlecenie, volume=1.5)
        _kloda(zlecenie, volume=2.0)

        trakownia = daily_report_service.zbierz_dane(PONIEDZIALEK)['trakownia']

        assert trakownia['klody'] == 2
        assert trakownia['m3'] == pytest.approx(3.5)


def test_trakownia_liczy_po_measured_at_nie_created_at(app):
    """
    Tablet potrafi rano wysłać pomiary z wczorajszego popołudnia (kolejka
    offline). Mają się policzyć do dnia, w którym faktycznie powstały —
    ta sama zasada co w sawmill_dashboard_stats.
    """
    with app.app_context():
        zlecenie = _zlecenie_trakowni()
        _kloda(zlecenie, volume=1.0,
               kiedy=datetime.combine(PONIEDZIALEK - timedelta(days=1), time(16, 0)))
        _kloda(zlecenie, volume=3.0,
               kiedy=datetime.combine(PONIEDZIALEK, time(9, 0)))

        trakownia = daily_report_service.zbierz_dane(PONIEDZIALEK)['trakownia']

        assert trakownia['klody'] == 1
        assert trakownia['m3'] == pytest.approx(3.0)


def test_usuniete_klody_nie_licza_sie(app):
    with app.app_context():
        zlecenie = _zlecenie_trakowni()
        _kloda(zlecenie, volume=1.0)
        _kloda(zlecenie, volume=9.0, usunieta=True)

        trakownia = daily_report_service.zbierz_dane(PONIEDZIALEK)['trakownia']

        assert trakownia['klody'] == 1
        assert trakownia['m3'] == pytest.approx(1.0)


def test_modele_trakowni_uzywaja_czasu_lokalnego(app):
    """
    Kontener chodzi na UTC, reszta bazy zapisuje czas Europe/Warsaw. Dopóki
    trakownia używała datetime.now, jej wpisy były o 2 h wcześniejsze niż
    wszystko inne — a między 22:00 a północą wpadały do sąsiedniego dnia.
    """
    from modules.production.sawmill import models as sawmill_models

    kolumny = [
        sawmill_models.SawmillLog.__table__.c.created_at,
        sawmill_models.SawmillOrder.__table__.c.created_at,
        sawmill_models.SawmillAudit.__table__.c.created_at,
    ]
    # Po nazwie, nie po tożsamości obiektu: SQLAlchemy opakowuje callable
    # w ColumnDefault i porównanie `is` bywa zależne od wersji.
    for kolumna in kolumny:
        assert kolumna.default.arg.__name__ == 'get_local_now', (
            f'{kolumna} nadal używa czasu kontenera zamiast get_local_now')


def test_serwisy_trakowni_nie_nadpisuja_created_at(app, zegar):
    """
    Test wyżej patrzy WYŁĄCZNIE na definicję kolumny i przeszedłby także
    wtedy, gdyby ktoś z powrotem wstawił `created_at=datetime.now()` do
    konstruktora w add_log()/write_audit() — a to był właśnie znaleziony błąd:
    jawna wartość w konstruktorze unieważnia `default=` kolumny i SQLAlchemy
    nigdy jej nie woła.

    Dlatego tu jedziemy przez PRAWDZIWE serwisy, z zamrożonym zegarem
    lokalnym ustawionym na chwilę, w którą realny zegar kontenera nie ma jak
    trafić. Asercja czyta wartość PO commicie, czyli to, co faktycznie
    wylądowało w bazie.
    """
    from modules.production.sawmill.services.orders import add_log, write_audit

    chwila = zegar(datetime(2026, 1, 15, 3, 17, 42))

    with app.app_context():
        zlecenie = _zlecenie_trakowni()

        log = add_log(
            zlecenie,
            {'mid_circumference_cm': 120.0, 'length_cm': 400.0},
            measured_at=datetime.combine(PONIEDZIALEK, time(10, 0)))
        wpis = write_audit(zlecenie.id, 'log_create', log_id=log.id)
        db.session.commit()

        assert log.created_at == chwila, (
            'add_log() ustawia created_at jawnie zamiast zostawić default= kolumny')
        assert wpis.created_at == chwila, (
            'write_audit() ustawia created_at jawnie zamiast zostawić default= kolumny')
        # measured_at przychodzi Z ZEWNĄTRZ (czas z tabletu) i NIE ma się
        # równać zegarowi serwera — inaczej pomiar z kolejki offline
        # przypisałby się do doby wysyłki zamiast doby pomiaru.
        assert log.measured_at == datetime.combine(PONIEDZIALEK, time(10, 0))


# ============================================================================
# EKSPORT XLSX
# ============================================================================

def _pusty_raport():
    """Minimalny dict o kształcie kontraktu zbierz_dane() — bez bazy."""
    return {
        'dzien': PONIEDZIALEK,
        'zakonczone': {'sztuki': 0, 'm3': 0.0, 'wartosc_netto': 0.0,
                       'cofniecia': 0},
        'wykonanie': {'sztuki': 0, 'm3': 0.0, 'wartosc_netto': 0.0,
                      'pozycje': 0, 'zamowienia': 0, 'cofniecia': 0},
        'ludzie': {'osoby': 0, 'godziny': 0.0, 'pokrycie_proc': 0.0,
                   'wiersze': [], 'nieprzypisane': {'sztuki': 0.0, 'm3': 0.0}},
        'trakownia': {'klody': 0, 'm3': 0.0},
        'stanowiska': [{'kod': 'gluing', 'etykieta': 'Sklejanie', 'zamowienia': 3,
                        'sztuki': 6,
                        'm3': 3.0, 'wartosc_netto': 600.0, 'cofniecia': 1,
                        'kolejka_szt': 14, 'kolejka_m3': 6.0}],
        'terminy': {'po_terminie': 2, 'dzis': 1, '1_2_dni': 0,
                    '3_7_dni': 3, '8_dni_plus': 5, 'bez_terminu': 1},
    }


def test_naglowki_uzywaja_slowa_roznica():
    """Symbol delty jest zakazany w eksportach — czyta je księgowość."""
    wszystkie = daily_report_export.NAGLOWKI_STANOWISKA + daily_report_export.NAGLOWKI_LUDZIE
    assert not any('Δ' in h for h in wszystkie)


def test_plik_ma_trzy_arkusze_o_ustalonych_nazwach():
    wb = load_workbook(io.BytesIO(
        daily_report_export.build_daily_xlsx(_pusty_raport())))
    assert wb.sheetnames == ['Dzień', 'Stanowiska', 'Ludzie']


def test_arkusz_stanowiska_ma_naglowek_i_wiersze():
    wb = load_workbook(io.BytesIO(
        daily_report_export.build_daily_xlsx(_pusty_raport())))
    ws = wb['Stanowiska']

    assert [c.value for c in ws[1]] == list(daily_report_export.NAGLOWKI_STANOWISKA)
    assert ws.cell(row=2, column=1).value == 'Sklejanie'
    assert ws.cell(row=2, column=2).value == 3   # zamówienia
    assert ws.cell(row=2, column=3).value == 6   # sztuki


def test_wartosci_stanowiska_zaokraglane_tak_samo_jak_suma():
    """
    get_station_work_per_day() liczy wartosc_netto jako total_value_net * delta
    / quantity, a kolejka_m3 jako sumę volume_m3 * quantity — oba to surowe
    wyniki dzielenia zmiennoprzecinkowego, bez zaokrąglenia. Wiersz SUMA JEST
    zaokrąglany, więc bez zaokrąglenia wiersza pojedynczego stanowiska komórka
    stanowiska pokazywała więcej cyfr po przecinku niż komórka SUMY tej samej
    kolumny, mimo że dla „ładnych" liczb (6, 600.0, 3.0) różnicy nie widać —
    dlatego test używa wartości, która się realnie nie dzieli.
    """
    dane = _pusty_raport()
    dane['stanowiska'][0]['wartosc_netto'] = 1000.0 * 7 / 3   # 2333.3333333333335
    dane['stanowiska'][0]['m3'] = 1 / 3                        # 0.3333333333333333
    dane['stanowiska'][0]['kolejka_m3'] = 2 / 3                # 0.6666666666666666
    dane['trakownia'] = {'klody': 12, 'm3': 8.4}

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Stanowiska']

    assert ws.cell(row=2, column=4).value == pytest.approx(0.333)      # m³
    assert ws.cell(row=2, column=5).value == pytest.approx(2333.33)    # wartość netto
    assert ws.cell(row=2, column=8).value == pytest.approx(0.667)      # kolejka m³

    # Zaokrąglanie nie może zamienić pustych komórek trakowni w zera —
    # _zaokr() musi przepuszczać None bez rzucania TypeError (round(None, 2)
    # rzuciłby wprost).
    wiersz_trakowni = _wiersz_sumy(ws) - 1
    assert ws.cell(row=wiersz_trakowni, column=1).value == 'Trakownia'
    assert ws.cell(row=wiersz_trakowni, column=2).value is None
    assert ws.cell(row=wiersz_trakowni, column=5).value is None
    assert ws.cell(row=wiersz_trakowni, column=7).value is None
    assert ws.cell(row=wiersz_trakowni, column=8).value is None


def test_trakownia_ma_puste_komorki_kolejki_a_nie_zera():
    """
    Trakownia nie ma statusów kolejki. Zero znaczyłoby „policzone i wyszło
    zero", a prawda brzmi „nie dotyczy" — openpyxl zapisuje None jako pustą
    komórkę. Ta sama zasada co w sawmill/services/exports.py:106-112.

    Wiersz trakowni składa EKSPORT z bloku dane['trakownia'] — agregat nie
    udaje, że trakownia jest ósmym stanowiskiem.
    """
    dane = _pusty_raport()
    dane['trakownia'] = {'klody': 12, 'm3': 8.4}

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Stanowiska']
    wiersz_trakowni = _wiersz_sumy(ws) - 1   # wiersz nad SUMĄ

    assert ws.cell(row=wiersz_trakowni, column=1).value == 'Trakownia'
    assert ws.cell(row=wiersz_trakowni, column=3).value == 12    # kłody
    assert ws.cell(row=wiersz_trakowni, column=2).value is None  # zamówienia
    assert ws.cell(row=wiersz_trakowni, column=5).value is None  # wartość netto
    assert ws.cell(row=wiersz_trakowni, column=7).value is None  # kolejka szt.
    assert ws.cell(row=wiersz_trakowni, column=8).value is None  # kolejka m³


def test_arkusz_stanowiska_konczy_sie_wierszem_suma():
    """
    Suma obejmuje wyłącznie stanowiska produkcyjne. Trakownia mierzy surowiec
    przed wejściem na halę — doliczenie jej m³ liczyłoby ten sam materiał
    drugi raz.
    """
    dane = _pusty_raport()
    dane['trakownia'] = {'klody': 12, 'm3': 8.4}
    dane['stanowiska'].append({
        'kod': 'packaging', 'etykieta': 'Pakowanie', 'zamowienia': 2, 'sztuki': 4, 'm3': 1.0,
        'wartosc_netto': 200.0, 'cofniecia': 0,
        'kolejka_szt': 7, 'kolejka_m3': 0.7,
    })

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Stanowiska']

    suma = _wiersz_sumy(ws)
    assert ws.cell(row=suma, column=3).value == 10   # 6 + 4, bez trakowni
    assert ws.cell(row=suma, column=4).value == pytest.approx(4.0)  # bez 8.4


def test_arkusz_ludzie_ma_wiersz_nieprzypisane():
    """
    Bez tego wiersza suma arkusza „Ludzie" nie zgadza się z sumą arkusza
    „Stanowiska" i raport wygląda na zepsuty.
    """
    dane = _pusty_raport()
    dane['ludzie']['wiersze'] = [{
        'nazwa': 'Adam Nowak', 'stanowiska': 'Sklejanie', 'sztuki': 8.0,
        'm3': 4.0, 'zdarzenia': 3, 'godziny': 7.5, 'tempo': 0.533,
    }]
    dane['ludzie']['nieprzypisane'] = {'sztuki': 5.0, 'm3': 2.5}

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Ludzie']
    etykiety = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]

    assert 'Adam Nowak' in etykiety
    assert 'Nieprzypisane' in etykiety


def test_puste_tempo_zostaje_pusta_komorka():
    dane = _pusty_raport()
    dane['ludzie']['wiersze'] = [{
        'nazwa': 'Adam Nowak', 'stanowiska': '', 'sztuki': 0.0, 'm3': 0.0,
        'zdarzenia': 0, 'godziny': 8.0, 'tempo': None,
    }]

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Ludzie']

    assert ws.cell(row=2, column=7).value is None


def test_polskie_znaki_przechodza_bez_okaleczenia():
    """
    W repo jest safe_str(), które wycina diakrytyki — to legacy, nie wymóg
    openpyxl. Ten test pilnuje, żeby nikt go tu nie skopiował.
    """
    dane = _pusty_raport()
    dane['stanowiska'][0]['etykieta'] = 'Wykańczanie — dąb, jesion'

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))

    assert wb['Stanowiska'].cell(row=2, column=1).value == 'Wykańczanie — dąb, jesion'


def test_nazwa_pliku_jest_ascii():
    """
    Polskie znaki w nagłówku Content-Disposition rwą odpowiedź (WSGI koduje
    nagłówki w latin-1) — patrz tests/test_nda_filename_header.py.
    """
    nazwa = daily_report_export.nazwa_pliku(PONIEDZIALEK)

    assert nazwa == 'Raport_produkcji_2026-08-10.xlsx'
    nazwa.encode('ascii')   # rzuci UnicodeEncodeError, jeśli się zepsuje


def _bloki_arkusza_dzien(ws):
    """
    {'WYKONANIE': {'Sztuki (wykonane − cofnięte)': 342, ...}, ...} — arkusz czytany
    blokami, nie płaskim słownikiem: etykieta „m³" występuje DWA razy (raz
    w WYKONANIU, raz w TRAKOWNI) i płaska mapa cicho gubiłaby pierwszą.
    """
    bloki = {}
    biezacy = None
    for wiersz in range(1, ws.max_row + 1):
        etykieta = ws.cell(row=wiersz, column=1).value
        wartosc = ws.cell(row=wiersz, column=2).value
        if etykieta is None:
            continue
        if wartosc is None and etykieta.isupper():
            biezacy = etykieta
            bloki[biezacy] = {}
        elif biezacy is not None:
            bloki[biezacy][etykieta] = wartosc
    return bloki


def test_arkusz_dzien_ma_cztery_bloki_i_liczby_podsumowania():
    """
    Arkusz „Dzień" to pierwsza rzecz, którą widzi otwierający załącznik —
    dotąd nie był sprawdzany niczym poza własną nazwą.
    """
    dane = _pusty_raport()
    dane['zakonczone'] = {'sztuki': 58, 'm3': 1.24913,
                          'wartosc_netto': 9210.417, 'cofniecia': 2}
    dane['wykonanie'] = {'sztuki': 342, 'm3': 4.18742, 'wartosc_netto': 28450.126,
                         'pozycje': 51, 'zamowienia': 18, 'cofniecia': 4}
    dane['ludzie'] = {'osoby': 5, 'godziny': 37.5, 'pokrycie_proc': 92.0,
                      'wiersze': [], 'nieprzypisane': {'sztuki': 0.0, 'm3': 0.0}}
    dane['trakownia'] = {'klody': 12, 'm3': 8.4}

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    bloki = _bloki_arkusza_dzien(wb['Dzień'])

    assert set(bloki) == {'WYKONANIE', 'LUDZIE', 'TRAKOWNIA', 'ZOSTAŁO'}

    wykonanie = bloki['WYKONANIE']
    # Wynik dnia (pakowanie) i ruch na hali (suma stanowisk) stoją w osobnych
    # wierszach — jedna etykieta na obie liczby zawyżała wynik kilkukrotnie.
    assert wykonanie['Zakończone — sztuki (spakowane − cofnięte)'] == 58
    assert wykonanie['Zakończone — m³'] == pytest.approx(1.249)
    assert wykonanie['Zakończone — wartość netto (zł)'] == pytest.approx(9210.42)
    assert wykonanie['Ruch na stanowiskach (szt.)'] == 342
    assert wykonanie['Ruch na stanowiskach (m³)'] == pytest.approx(4.187)
    assert wykonanie['Pozycji dotkniętych'] == 51
    assert wykonanie['Zamówień dotkniętych'] == 18
    assert wykonanie['Cofnięcia (szt.)'] == 4
    # Ruch świadomie NIE ma odpowiednika w złotówkach: ta sama pozycja
    # doliczona na pięciu stanowiskach dałaby pięciokrotną cenę.
    assert 'Ruch na stanowiskach (zł)' not in wykonanie

    assert bloki['LUDZIE'] == {'Pracowników z pracą': 5, 'Osobogodziny': 37.5,
                               'Pokrycie atrybucją (%)': 92.0}

    assert bloki['TRAKOWNIA']['Kłody'] == 12
    assert bloki['TRAKOWNIA']['m³'] == pytest.approx(8.4)

    zostalo = bloki['ZOSTAŁO']
    # Kolejka z wiersza stanowiska w _pusty_raport(): 14 szt. / 6.0 m³.
    assert zostalo['W kolejce (szt.)'] == 14
    assert zostalo['W kolejce (m³)'] == pytest.approx(6.0)
    assert zostalo['Po terminie'] == 2
    assert zostalo['Termin dziś'] == 1
    assert zostalo['Termin za 1–2 dni'] == 0
    assert zostalo['Termin za 3–7 dni'] == 3
    assert zostalo['Termin za 8+ dni'] == 5
    assert zostalo['Bez terminu'] == 1


def test_arkusz_dzien_zostawia_puste_pokrycie_a_nie_zero():
    """
    Doba bez ŻADNEGO ruchu nie ma z czego policzyć pokrycia — agregat oddaje
    None, a arkusz ma zostawić PUSTĄ komórkę. Zero znaczyłoby „policzone
    i wyszło zero", czyli zarzut, że nikt się nie podpisał pod robotą.
    """
    dane = _pusty_raport()
    dane['ludzie']['pokrycie_proc'] = None

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    bloki = _bloki_arkusza_dzien(wb['Dzień'])

    assert bloki['LUDZIE']['Pokrycie atrybucją (%)'] is None


# ============================================================================
# UZGODNIENIE ARKUSZY
# ============================================================================

def test_suma_ludzi_zgadza_sie_z_suma_stanowisk(app):
    """
    Test uzgodnienia: kolumna „Sztuki (wkład)" w arkuszu „Ludzie" — RAZEM
    z wierszem „Nieprzypisane" — musi dać tyle samo, co wiersz SUMA
    w arkuszu „Stanowiska".

    To jedyna asercja, która łapie rozjazd MIĘDZY arkuszami: wkład liczy się
    przez udziały (share), przerób przez delty zdarzeń, i są to dwie
    niezależne drogi do tej samej liczby. Tolerancja bierze się z zaokrąglania
    wkładu do dwóch miejsc (worker_stats_service.MIEJSC_WKLADU).
    """
    with app.app_context():
        adam = _pracownik('Adam')
        bartek = _pracownik('Bartek')
        produkt = _produkt(quantity=20, volume=0.5, wartosc=2000.0)

        # 6 szt. w całości Adama.
        _event(produkt, 'gluing', 6,
               datetime.combine(PONIEDZIALEK, time(10, 0)), worker=adam)

        # 5 szt. rozdzielone po połowie — udział nie musi być całkowity,
        # a to właśnie na ułamkach uzgodnienie potrafi się rozjechać.
        zdarzenie = ProductionStationEvent(
            production_item_id=produkt.id, station_code='gluing', delta=5,
            quantity_done_after=11, source='mobile',
            created_at=datetime.combine(PONIEDZIALEK, time(11, 0)))
        db.session.add(zdarzenie)
        db.session.flush()
        db.session.add_all([
            ProductionStationEventWorker(event_id=zdarzenie.id,
                                         worker_id=adam.id, share=0.5),
            ProductionStationEventWorker(event_id=zdarzenie.id,
                                         worker_id=bartek.id, share=0.5),
        ])
        db.session.commit()

        # 4 szt. bez atrybucji — to one wypełniają wiersz „Nieprzypisane".
        _event(produkt, 'packaging', 4,
               datetime.combine(PONIEDZIALEK, time(12, 0)))

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)
        wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))

        arkusz_stanowisk = wb['Stanowiska']
        suma_stanowisk = arkusz_stanowisk.cell(
            row=_wiersz_sumy(arkusz_stanowisk), column=3).value

        arkusz_ludzi = wb['Ludzie']
        suma_ludzi = sum(
            arkusz_ludzi.cell(row=wiersz, column=3).value or 0
            for wiersz in range(2, arkusz_ludzi.max_row + 1))

        assert suma_stanowisk == 15          # 6 + 5 + 4
        assert suma_ludzi == pytest.approx(suma_stanowisk, abs=0.05)

        # Wiersz „Nieprzypisane" musi tu realnie coś nieść — inaczej test
        # przechodziłby także wtedy, gdyby eksport go w ogóle nie dopisywał.
        etykiety = [arkusz_ludzi.cell(row=wiersz, column=1).value
                    for wiersz in range(2, arkusz_ludzi.max_row + 1)]
        assert etykiety[-1] == 'Nieprzypisane'
        assert arkusz_ludzi.cell(
            row=arkusz_ludzi.max_row, column=3).value == pytest.approx(4.0)


# ============================================================================
# WYGLĄD ARKUSZA — kolory firmowe, formaty liczb, szerokości kolumn
# ============================================================================

def test_naglowki_maja_firmowy_grafit_i_bialy_tekst():
    """
    Nagłówki mają grafitowe tło (#1F2020 — --secondary-color z CSS aplikacji),
    a nie pomarańczowe: biały tekst na #ED6B24 daje kontrast ok. 2,8:1, czyli
    poniżej progu czytelności. Pomarańcz pracuje jako akcent.
    """
    wb = load_workbook(io.BytesIO(
        daily_report_export.build_daily_xlsx(_pusty_raport())))

    for nazwa in ('Stanowiska', 'Ludzie'):
        komorka = wb[nazwa].cell(row=1, column=1)
        assert komorka.fill.fgColor.rgb.endswith('1F2020'), nazwa
        assert komorka.font.color.rgb.endswith('FFFFFF'), nazwa
        assert komorka.font.bold is True, nazwa


def test_liczby_maja_jawny_format_w_notacji_ooxml():
    """
    Bez jawnego formatu Excel renderuje liczby wedle ustawień systemu, więc ten
    sam plik u dwóch osób wygląda inaczej.

    Kod formatu MUSI być w notacji en-US (przecinek = tysiące, kropka =
    dziesiętne) — Excel podmienia separatory przy wyświetlaniu. Polska notacja
    („# ##0,00") dałaby format, którego Excel nie rozpozna, więc ten test
    pilnuje właśnie kropki.
    """
    wb = load_workbook(io.BytesIO(
        daily_report_export.build_daily_xlsx(_pusty_raport())))
    ws = wb['Stanowiska']

    assert ws.cell(row=2, column=2).number_format == '#,##0'      # zamówienia
    assert ws.cell(row=2, column=3).number_format == '#,##0'      # sztuki
    assert ws.cell(row=2, column=4).number_format == '#,##0.000'  # m³
    assert ws.cell(row=2, column=5).number_format == '#,##0.00'   # wartość

    for format_ in ('#,##0', '#,##0.000', '#,##0.00'):
        assert '.' in format_ or ',' in format_
        assert ' ' not in format_


def test_kolorowanie_nie_zamienia_pustej_komorki_w_zero():
    """
    Wiersz trakowni dostaje własne tło, ale jego puste komórki mają POZOSTAĆ
    puste. Wypełnienie tła to atrybut stylu, nie wartość — gdyby kolorowanie
    przechodziło przez zapis wartości, „nie dotyczy" zamieniłoby się w zero.
    """
    dane = _pusty_raport()
    dane['trakownia'] = {'klody': 12, 'm3': 8.4}

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Stanowiska']
    wiersz_trakowni = _wiersz_sumy(ws) - 1

    assert ws.cell(row=wiersz_trakowni, column=1).value == 'Trakownia'
    assert ws.cell(row=wiersz_trakowni, column=1).fill.fgColor.rgb.endswith('F2F0ED')
    # Zamówienia, wartość netto i obie kolumny kolejki — trakownia ich nie ma.
    for kolumna in (2, 5, 7, 8):
        assert ws.cell(row=wiersz_trakowni, column=kolumna).value is None


def test_kolumna_rozszerza_sie_do_najdluzszej_zawartosci():
    """
    Nazwisko bywa dłuższe niż nagłówek „Pracownik", a lista stanowisk potrafi
    mieć trzy pozycje. Szerokość liczona z nagłówka ucinałaby jedno i drugie.
    """
    dane = _pusty_raport()
    dane['ludzie']['wiersze'] = [{
        'nazwa': 'Sylwester Rębiś-Zawadzki', 'stanowiska': 'Lakiernia, Pakowanie',
        'sztuki': 8.0, 'm3': 4.0, 'zdarzenia': 3, 'godziny': 7.5, 'tempo': 0.533,
    }]

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Ludzie']

    assert ws.column_dimensions['A'].width >= len('Sylwester Rębiś-Zawadzki')
    assert ws.column_dimensions['B'].width >= len('Lakiernia, Pakowanie')


def test_szerokosc_kolumny_ma_gorny_limit():
    """
    „Lakiernia, Pakowanie, Wykańczanie" przy każdym pracowniku rozpchałoby
    arkusz na pół ekranu. Limit trzyma tabelę w rozsądnej szerokości.
    """
    dane = _pusty_raport()
    dane['ludzie']['wiersze'] = [{
        'nazwa': 'X' * 200, 'stanowiska': '', 'sztuki': 0.0, 'm3': 0.0,
        'zdarzenia': 0, 'godziny': 0.0, 'tempo': None,
    }]

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))

    assert wb['Ludzie'].column_dimensions['A'].width <= 42


def test_wiersz_sumy_ma_akcent_i_pogrubienie():
    wb = load_workbook(io.BytesIO(
        daily_report_export.build_daily_xlsx(_pusty_raport())))
    ws = wb['Stanowiska']
    komorka = ws.cell(row=_wiersz_sumy(ws), column=1)

    assert komorka.value == 'SUMA'
    assert komorka.font.bold is True
    assert komorka.fill.fgColor.rgb.endswith('FBE1D3')
    assert komorka.border.top.color.rgb.endswith('ED6B24')


def test_etykieta_sztuk_tlumaczy_sie_sama():
    """
    „Sztuki (netto)" wymagało tłumaczenia przy każdym czytaniu raportu —
    działanie w nawiasie mówi to samo bez żargonu.
    """
    wb = load_workbook(io.BytesIO(
        daily_report_export.build_daily_xlsx(_pusty_raport())))
    ws = wb['Dzień']
    etykiety = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]

    assert 'Zakończone — sztuki (spakowane − cofnięte)' in etykiety
    assert 'Sztuki (netto)' not in etykiety
    # Sama „Sztuki" też nie wystarcza: arkusz podaje teraz dwie różne miary
    # sztuk (wynik dnia i ruch na stanowiskach), więc etykieta musi mówić,
    # o którą chodzi.
    assert 'Sztuki' not in etykiety


def test_suma_zamowien_nie_jest_suma_kolumny():
    """
    Jedno zamówienie zwykle przechodzi tego dnia przez kilka stanowisk, więc
    zsumowanie kolumny „Zamówienia" policzyłoby je wielokrotnie. W wierszu SUMA
    musi wylądować globalna liczba UNIKALNYCH zamówień z bloku „wykonanie".

    Tu: dwa stanowiska raportują po 3 i 2 zamówienia (suma kolumny = 5), ale
    realnie dotknięto tylko 4 różnych zamówień — bo jedno przeszło przez oba.
    """
    dane = _pusty_raport()
    dane['wykonanie']['zamowienia'] = 4
    dane['stanowiska'].append({
        'kod': 'packaging', 'etykieta': 'Pakowanie', 'zamowienia': 2,
        'sztuki': 4, 'm3': 1.0, 'wartosc_netto': 200.0, 'cofniecia': 0,
        'kolejka_szt': 7, 'kolejka_m3': 0.7,
    })

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Stanowiska']

    assert ws.cell(row=2, column=2).value == 3      # sklejanie
    assert ws.cell(row=3, column=2).value == 2      # pakowanie
    assert ws.cell(row=_wiersz_sumy(ws), column=2).value == 4   # nie 5


def test_zamowienia_per_stanowisko_licza_unikalne_zamowienia(app):
    """
    Dwie pozycje z tego samego zamówienia dotknięte na jednym stanowisku to
    jedno zamówienie, nie dwa. Zdarzenia automatu nie liczą się w ogóle.
    """
    with app.app_context():
        p1 = _produkt()
        p2 = _produkt()
        # Druga pozycja tego samego zamówienia co p1.
        p1b = ProductionProduct(
            order_id=p1.order_id,
            # short_product_id ma walidator wymuszający format N_S (models.py:325).
            short_product_id=f"{p1.short_product_id.split('_')[0]}_2",
            product_sequence_in_order=2, original_product_name='Blat',
            quantity=5, volume_m3=0.5, total_value_net=500.0,
            current_status='czeka_na_sklejanie',
            created_at=datetime.combine(PONIEDZIALEK, time(9, 0)))
        db.session.add(p1b)
        db.session.commit()

        kiedy = datetime.combine(PONIEDZIALEK, time(10, 0))
        _event(p1, 'gluing', 2, kiedy)
        _event(p1b, 'gluing', 3, kiedy)
        _event(p2, 'gluing', 1, kiedy)
        _event(p2, 'formatting', 4, kiedy, source='auto_skip')

        dane = daily_report_service.zbierz_dane(PONIEDZIALEK)

        assert _stanowisko(dane, 'gluing')['zamowienia'] == 2
        assert _stanowisko(dane, 'formatting')['zamowienia'] == 0


def test_przypis_wyjasnia_dlaczego_suma_zamowien_nie_jest_suma_kolumny():
    """
    Suma kolumny „Zamówienia" jest większa niż liczba w wierszu SUMA i bez
    wyjaśnienia wygląda to na błąd arkusza. Przypis stoi pod tabelą, bo tam
    czytelnik trafia zaraz po zobaczeniu rozbieżności.
    """
    dane = _pusty_raport()
    dane['wykonanie']['zamowienia'] = 4
    dane['stanowiska'].append({
        'kod': 'packaging', 'etykieta': 'Pakowanie', 'zamowienia': 2,
        'sztuki': 4, 'm3': 1.0, 'wartosc_netto': 200.0, 'cofniecia': 0,
        'kolejka_szt': 7, 'kolejka_m3': 0.7,
    })

    wb = load_workbook(io.BytesIO(daily_report_export.build_daily_xlsx(dane)))
    ws = wb['Stanowiska']
    ostatni = ws.cell(row=ws.max_row, column=1).value

    assert 'unikalnych zamówień' in ostatni
    assert 'nie suma kolumny' in ostatni
    # Przypis stoi POD wierszem SUMA, nie zamiast niego.
    etykiety = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert etykiety.index('SUMA') < ws.max_row - 1


def test_przypis_nie_rozpycha_kolumny():
    """
    Zdanie przypisu ma ponad 150 znaków. Gdyby autoszerokość je uwzględniała,
    pierwsza kolumna miałaby szerokość całego zdania i tabela stałaby się
    nieczytelna.
    """
    wb = load_workbook(io.BytesIO(
        daily_report_export.build_daily_xlsx(_pusty_raport())))

    assert wb['Stanowiska'].column_dimensions['A'].width <= 42
