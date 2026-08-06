# -*- coding: utf-8 -*-
"""Eksport XLSX i kontekst protokołu PDF."""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import io
from datetime import date, datetime
from decimal import Decimal

import pypdf
from openpyxl import load_workbook

from modules.production.sawmill.services.exports import (
    XLSX_HEADERS, build_orders_xlsx, build_protocol_context,
)
from modules.production.sawmill.services.orders import (
    add_log, compute_differences, order_totals,
)
from modules.production.sawmill.models import (
    STATUS_COMPLETED, SawmillDelivery, SawmillOrder, SawmillSpecies, SawmillSupplier,
)
from extensions import db
from tests.sawmill_fixtures import BASE, app, client  # noqa: F401

WIERSZ = {
    'order_number': 'TRK/2026/001', 'supplier_name': 'Tartak Nowak',
    'invoice_number': 'FV/2026/0451', 'delivery_date': '2026-08-05',
    'species': 'Dąb', 'declared_volume_m3': 80.0, 'measured_volume_m3': 76.34,
    'difference_m3': -3.66, 'difference_pct': -4.58, 'difference_value': -4392.0,
    'logs_count': 118, 'status': 'settled', 'agreed_volume_m3': 76.5,
}

POMIAR = {
    'mid_circumference_cm': Decimal('125.6'),
    'length_cm': Decimal('410.0'),
}


def test_naglowki_uzywaja_slowa_roznica():
    """Symbol delty jest zakazany także w eksportach — czyta je księgowość."""
    assert 'Różnica m³' in XLSX_HEADERS
    assert 'Różnica %' in XLSX_HEADERS
    assert 'Różnica zł' in XLSX_HEADERS
    assert not any('Δ' in h for h in XLSX_HEADERS)


def test_xlsx_ma_naglowek_i_wiersz():
    dane = build_orders_xlsx([WIERSZ])
    wb = load_workbook(io.BytesIO(dane))
    ws = wb.active
    assert [c.value for c in ws[1]] == list(XLSX_HEADERS)
    assert ws.cell(row=2, column=1).value == 'TRK/2026/001'


def test_xlsx_z_pusta_lista_ma_sam_naglowek():
    wb = load_workbook(io.BytesIO(build_orders_xlsx([])))
    assert wb.active.max_row == 1


def test_brak_ceny_daje_pusta_komorke_nie_zero():
    wiersz = dict(WIERSZ, difference_value=None, price_per_m3=None)
    wb = load_workbook(io.BytesIO(build_orders_xlsx([wiersz])))
    ws = wb.active
    kolumna = list(XLSX_HEADERS).index('Różnica zł') + 1
    assert ws.cell(row=2, column=kolumna).value is None


def _zlecenie_z_pomiarami(app, cena='1200.00', liczba_pomiarow=2):
    """Zlecenie z dostawcą (polskie znaki w nazwie i adresie) i kilkoma kłodami."""
    with app.app_context():
        supplier = SawmillSupplier(
            name='Tartak Żółć Sp. z o.o.', nip='1234567890',
            address_street='Leśna 5', address_zip='36-068', address_city='Bąchórz',
            contact_person='Łukasz Wąsik', phone='500600700', email='tartak@example.pl',
        )
        db.session.add(supplier)
        db.session.flush()
        delivery = SawmillDelivery(
            supplier_id=supplier.id, delivery_date=date(2026, 8, 5),
            invoice_number='FV/2026/0451', invoice_date=date(2026, 8, 4),
        )
        db.session.add(delivery)
        db.session.flush()
        species = SawmillSpecies.query.first()
        order = SawmillOrder(
            order_number='TRK/2026/001', delivery_id=delivery.id,
            species_id=species.id, declared_volume_m3=Decimal('1.200000'),
            price_per_m3=Decimal(cena) if cena else None,
        )
        db.session.add(order)
        db.session.flush()
        for _ in range(liczba_pomiarow):
            add_log(order, POMIAR, datetime(2026, 8, 5, 9, 31, 12), device_id='TRAK-1')
        # Pomiary idą przez add_log (tablet, status musi być otwarty) — status
        # `completed` ustawiamy dopiero PO nich, tak jak realny cykl życia
        # zlecenia (tablet kończy pomiary, potem `POST /complete`).
        order.status = STATUS_COMPLETED
        db.session.commit()
        return order.id


def test_build_protocol_context_liczy_srednie(app):
    oid = _zlecenie_z_pomiarami(app, liczba_pomiarow=3)
    with app.app_context():
        order = db.session.query(SawmillOrder).get(oid)
        count, volume = order_totals(order.id)
        logs = order.logs if hasattr(order, 'logs') else None
        from modules.production.sawmill.models import SawmillLog
        logs = (SawmillLog.query.filter_by(order_id=order.id, is_deleted=False)
                .order_by(SawmillLog.sequence_no).all())
        differences = compute_differences(order, volume, 5.0)

        ctx = build_protocol_context(order, order.delivery, logs, count, volume, differences)

        assert ctx['logs_count'] == 3
        assert ctx['avg_volume_m3'] == volume / 3
        # Sam wzór volume.py — kontrolna wartość ze specyfikacji (sekcja 5).
        assert round(float(logs[0].volume_m3), 6) == 0.514699
        assert ctx['avg_circumference_cm'] == Decimal('125.6')
        assert ctx['avg_length_cm'] == Decimal('410.0')
        assert ctx['status_label'] == 'Zakończone'


def test_build_protocol_context_brak_ceny_daje_none_w_roznicy_zl(app):
    oid = _zlecenie_z_pomiarami(app, cena=None, liczba_pomiarow=1)
    with app.app_context():
        order = db.session.query(SawmillOrder).get(oid)
        count, volume = order_totals(order.id)
        from modules.production.sawmill.models import SawmillLog
        logs = SawmillLog.query.filter_by(order_id=order.id).all()
        differences = compute_differences(order, volume, 5.0)
        ctx = build_protocol_context(order, order.delivery, logs, count, volume, differences)
        assert ctx['differences']['difference_value'] is None


def test_export_xlsx_reuzywa_filtrow_orders_list(client, app):
    """/export.xlsx ma zwrócić dokładnie to, co widać po zastosowaniu filtra statusu."""
    _zlecenie_z_pomiarami(app, liczba_pomiarow=1)
    r = client.get(BASE + '/export.xlsx?status=completed')
    assert r.status_code == 200
    assert r.headers['Content-Type'] == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    wb = load_workbook(io.BytesIO(r.data))
    ws = wb.active
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == 'TRK/2026/001'

    # Filtr, który nie pasuje do żadnego zlecenia, ma dać sam nagłówek.
    r2 = client.get(BASE + '/export.xlsx?status=new')
    wb2 = load_workbook(io.BytesIO(r2.data))
    assert wb2.active.max_row == 1


def test_protocol_pdf_renderuje_polskie_znaki(client, app):
    oid = _zlecenie_z_pomiarami(app, liczba_pomiarow=2)
    r = client.get(BASE + '/orders/{}/protocol.pdf'.format(oid))
    assert r.status_code == 200
    assert r.headers['Content-Type'] == 'application/pdf'
    assert r.data[:4] == b'%PDF'
    assert len(r.data) > 1000

    reader = pypdf.PdfReader(io.BytesIO(r.data))
    tekst = '\n'.join(page.extract_text() for page in reader.pages)
    assert 'Protokół pomiaru' in tekst
    assert 'Różnica' in tekst
    assert 'Żółć' in tekst or 'Bąchórz' in tekst  # polskie znaki dostawcy


def test_protocol_pdf_404_dla_nieznanego_zlecenia(client, app):
    r = client.get(BASE + '/orders/999999/protocol.pdf')
    assert r.status_code == 404
