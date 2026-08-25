# -*- coding: utf-8 -*-
"""
Eksporty i agregaty trakowni.

Zadanie 10 dało `sawmill_dashboard_stats` (kafelek dashboardu). Zadanie 13
dokłada eksport XLSX listy zleceń oraz kontekst szablonu protokołu PDF —
kontrakt API dla sesji Android idzie osobno, jako `docs/sawmill-mobile-api-contract.md`.
"""

import io
from datetime import datetime, time
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import func

from extensions import db
from modules.production.models import get_local_now
from modules.production.sawmill.models import (
    OPEN_STATUSES, STATUS_COMPLETED, STATUS_IN_PROGRESS, STATUS_NEW,
    STATUS_SETTLED, SawmillLog, SawmillOrder,
)


def sawmill_dashboard_stats():
    """
    Cztery liczby na kafelek dashboardu plus postęp realizacji.

    „Dziś" liczone po measured_at, NIE po created_at — tablet potrafi rano
    wysłać pomiary z wczorajszego popołudnia (kolejka offline), a mają się
    policzyć do dnia, w którym faktycznie powstały, inaczej statystyka
    wydajności traka jest bezwartościowa.
    """
    # Jedno odczytanie zegara na obie granice. Dwa osobne wywołania trafiały
    # w różne doby, gdy przebieg przypadł na przejście przez północ — a wtedy
    # `dzis` było z dnia poprzedniego, `jutro` z następnego i zakres wychodził
    # odwrotny (koniec wcześniej niż początek), czyli statystyka pusta.
    dzien = get_local_now().date()
    dzis = datetime.combine(dzien, time.min)
    jutro = datetime.combine(dzien, time.max)

    open_orders = SawmillOrder.query.filter(
        SawmillOrder.status.in_(OPEN_STATUSES)).count()
    to_settle = SawmillOrder.query.filter(
        SawmillOrder.status == STATUS_COMPLETED).count()

    dzis_row = (
        db.session.query(func.count(SawmillLog.id),
                         func.coalesce(func.sum(SawmillLog.volume_m3), 0))
        .filter(SawmillLog.is_deleted.is_(False))
        .filter(SawmillLog.measured_at >= dzis)
        .filter(SawmillLog.measured_at <= jutro)
        .one()
    )

    # Postęp realizacji otwartych zleceń — to widok admina, więc deklaracja
    # może się tu pojawić bez ryzyka; na tablet i tak nie idzie.
    otwarte = SawmillOrder.query.filter(SawmillOrder.status.in_(OPEN_STATUSES)).all()
    zadeklarowano = sum((Decimal(str(o.declared_volume_m3)) for o in otwarte), Decimal(0))
    zmierzono = Decimal(0)
    for order in otwarte:
        suma = (db.session.query(func.coalesce(func.sum(SawmillLog.volume_m3), 0))
                .filter(SawmillLog.order_id == order.id)
                .filter(SawmillLog.is_deleted.is_(False)).scalar())
        zmierzono += Decimal(str(suma))

    progress = float(zmierzono / zadeklarowano * 100) if zadeklarowano > 0 else 0.0

    return {
        'open_orders': open_orders,
        'logs_today': int(dzis_row[0]),
        'volume_today_m3': float(dzis_row[1]),
        'to_settle': to_settle,
        'progress_pct': round(min(progress, 100.0), 1),
    }


# ── Eksport XLSX ────────────────────────────────────────────────────────────

# W CAŁYM interfejsie i eksportach używamy słowa „Różnica", nigdy symbolu Δ —
# arkusz trafia do księgowości, która nie musi znać notacji technicznej.
# Wymaganie właściciela projektu, sprawdzane testem
# (test_naglowki_uzywaja_slowa_roznica w tests/test_sawmill_exports.py).
XLSX_HEADERS = (
    'Nr TRK', 'Dostawca', 'Faktura', 'Data dostawy', 'Gatunek',
    'm³ deklarowane', 'm³ zmierzone', 'Różnica m³', 'Różnica %', 'Różnica zł',
    'Kłód', 'm³ uzgodnione', 'Status',
)

# Kolejność 1:1 z XLSX_HEADERS — klucze pochodzą z payloadu `orders_list()`
# (serialize_order_for_panel), czyli DOKŁADNIE tego, co widzi użytkownik
# w tabeli panelu po zastosowaniu filtrów.
_XLSX_KEYS = (
    'order_number', 'supplier_name', 'invoice_number', 'delivery_date', 'species',
    'declared_volume_m3', 'measured_volume_m3', 'difference_m3', 'difference_pct',
    'difference_value', 'logs_count', 'agreed_volume_m3', 'status',
)

STATUS_LABELS = {
    STATUS_NEW: 'Nowe', STATUS_IN_PROGRESS: 'W trakcie',
    STATUS_COMPLETED: 'Zakończone', STATUS_SETTLED: 'Rozliczone',
}


def build_orders_xlsx(orders_payload):
    """
    Buduje arkusz z przefiltrowanej listy zleceń (payload `orders_list()`
    z panel_api.py — ten sam request, te same filtry, bez duplikowania ich
    parsowania). Zwraca bajty pliku .xlsx.

    Brakująca wartość (np. `difference_value` bez ceny za m³, albo wszystkie
    trzy różnice przy zleceniu jeszcze niezakończonym) zostaje w komórce jako
    `None` — openpyxl zapisuje to jako PUSTĄ komórkę, nie zero. Zero w
    kolumnie różnicy zł sugerowałoby księgowej, że różnica faktycznie wynosi
    0 zł, a nie że cena nie była znana albo że pomiar jeszcze trwa. Arkusz
    zachowuje się tu tak samo jak panel (myślnik), co jest celowe — te same
    dane nie mogą mówić w dwóch miejscach czegoś innego.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Trakownia'

    ws.append(list(XLSX_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row in orders_payload:
        ws.append([
            STATUS_LABELS.get(row.get(key), row.get(key)) if key == 'status'
            else row.get(key)
            for key in _XLSX_KEYS
        ])

    for idx, header in enumerate(XLSX_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = \
            max(12, len(header) + 2)
    ws.freeze_panes = 'A2'

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


# ── Protokół PDF ────────────────────────────────────────────────────────────

def build_protocol_context(order, delivery, logs, logs_count,
                           measured_volume_m3, differences):
    """
    Kontekst dla szablonu `sawmill/protocol_pdf.html` — dokument samodzielny
    (WeasyPrint nie ma dostępu do statycznych zasobów aplikacji, CSS jest
    inline w szablonie), idący jako załącznik do reklamacji u dostawcy.

    Podsumowanie podaje średni OBWÓD, nie średnicę — obwód jest tym, co
    pracownik faktycznie zmierzył, więc dostawca może go zweryfikować taśmą
    na tej samej kłodzie. Średnica byłaby wielkością pochodną i sugerowałaby
    pomiar, którego nikt nie wykonał.
    """
    obwody = []
    dlugosci = []
    for log in logs:
        obwody.append(Decimal(str(log.mid_circumference_cm)))
        dlugosci.append(Decimal(str(log.length_cm)))

    return {
        'order': order,
        'delivery': delivery,
        'logs': logs,
        'logs_count': logs_count,
        'measured_volume_m3': measured_volume_m3,
        'differences': differences,
        'avg_volume_m3': (measured_volume_m3 / logs_count) if logs_count else None,
        'avg_circumference_cm': (sum(obwody) / len(obwody)) if obwody else None,
        'avg_length_cm': (sum(dlugosci) / len(dlugosci)) if dlugosci else None,
        'status_label': STATUS_LABELS.get(order.status, order.status),
        # Data wygenerowania na stopce protokołu — kto i kiedy wydrukował
        # dokument idący do dostawcy, niezależnie od daty dostawy czy faktury.
        'generated_at': get_local_now(),
    }
