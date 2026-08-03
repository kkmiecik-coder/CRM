# modules/production/routers/api/products_api.py
"""
Products tab content + CRUD + task completion + station progress endpoints.
Extracted from api_routers.py.
"""

import json
import traceback
from datetime import datetime, date, timedelta
from flask import request, jsonify, render_template, current_app
from flask_login import login_required, current_user
from extensions import db
from sqlalchemy import and_, or_, func, distinct, cast, String, case

from . import api_bp, logger, ProductionItem, ProductionError, get_local_now
from .common_api import admin_required, ip_validation_required, _format_status, _validate_config_value
from modules.production.models import ProductionOrder, ProductionConfiguration


@api_bp.route('/complete-task', methods=['POST'])
@ip_validation_required
def complete_task():
    """
    POST /api/complete-task - Oznaczenie zadania jako wykonane (PRD Section 6.2)
    
    Body JSON zgodny z PRD:
    {
        "product_id": "25_05248_1",
        "station_code": "cutting"
    }
    
    Akcje:
    1. Sprawdza czy produkt istnieje i ma odpowiedni status
    2. Oznacza zadanie jako ukończone w modelu
    3. Automatycznie zmienia status na następny w workflow
    4. Loguje operację
    
    Autoryzacja: Brak (walidacja IP)
    Returns: JSON status operacji
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400
        
        product_id = data.get('product_id')
        station_code = data.get('station_code')
        
        if not product_id or not station_code:
            return jsonify({
                'success': False, 
                'error': 'Wymagane pola: product_id, station_code'
            }), 400
        
        # Walidacja station_code
        valid_stations = ['cutting', 'assembly', 'packaging']
        if station_code not in valid_stations:
            return jsonify({
                'success': False,
                'error': f'Nieprawidłowy station_code. Dozwolone: {valid_stations}'
            }), 400
        
        logger.info("API: Próba ukończenia zadania", extra={
            'product_id': product_id,
            'station_code': station_code,
            'client_ip': request.remote_addr
        })
        
        from ...models import ProductionItem

        # Znajdź produkt - preferuj record_id (PK) gdy obecny (disambig dla doróbek)
        record_id = data.get('record_id')
        if record_id is not None:
            product = ProductionItem.query.filter_by(id=int(record_id)).first()
        else:
            # Fallback do short_product_id (legacy)
            product = ProductionItem.query.filter_by(short_product_id=product_id).first()
        if not product:
            return jsonify({
                'success': False,
                'error': f'Produkt {product_id} nie znaleziony'
            }), 404

        # Sprawdź czy produkt jest w odpowiednim statusie dla danego stanowiska
        expected_status_map = {
            'cutting': 'czeka_na_wyciecie',
            'assembly': 'czeka_na_skladanie', 
            'packaging': 'czeka_na_pakowanie'
        }
        
        expected_status = expected_status_map[station_code]
        if product.current_status != expected_status:
            return jsonify({
                'success': False,
                'error': f'Produkt ma status "{product.current_status}", oczekiwano "{expected_status}"'
            }), 400
        
        # Ukończ zadanie używając metody z modelu
        old_status = product.current_status
        product.complete_task(station_code)
        
        db.session.commit()
        
        logger.info("API: Ukończono zadanie", extra={
            'product_id': product_id,
            'station_code': station_code,
            'old_status': old_status,
            'new_status': product.current_status,
            'client_ip': request.remote_addr
        })
        
        return jsonify({
            'success': True,
            'message': f'Zadanie {station_code} dla produktu {product_id} ukończone',
            'data': {
                'product_id': product.short_product_id,
                'old_status': old_status,
                'new_status': product.current_status,
                'completed_at': get_local_now().isoformat()
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error("API: Błąd ukończenia zadania", extra={
            'product_id': data.get('product_id') if 'data' in locals() else None,
            'station_code': data.get('station_code') if 'data' in locals() else None,
            'client_ip': request.remote_addr,
            'error': str(e)
        })
        
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# API ROUTERS - Toggle Product Marked Done (Checkbox na stanowiskach)
# ============================================================================


@api_bp.route('/toggle-product-done', methods=['POST'])
@ip_validation_required
def toggle_product_done():
    """
    POST /api/toggle-product-done - Zmiana stanu checkboxa produktu na stanowisku

    Body JSON:
    {
        "product_id": "25_05248_1",
        "station": "cutting",
        "is_done": true
    }

    Akcje:
    1. Waliduje station (cutting, assembly, gluing, formatting, finishing, packaging)
    2. Znajduje produkt po short_product_id
    3. Ustawia {station}_marked_done = is_done
    4. Zapisuje do bazy

    Autoryzacja: Walidacja IP (stanowiska)
    Returns: JSON status operacji
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        product_id = data.get('product_id')
        station = data.get('station')
        is_done = data.get('is_done')

        # Walidacja wymaganych pól
        if not product_id or not station or is_done is None:
            return jsonify({
                'success': False,
                'error': 'Wymagane pola: product_id, station, is_done'
            }), 400

        # Walidacja station
        valid_stations = ['cutting', 'assembly', 'gluing', 'formatting', 'finishing', 'packaging']
        if station not in valid_stations:
            return jsonify({
                'success': False,
                'error': f'Nieprawidłowy station. Dozwolone: {valid_stations}'
            }), 400

        # Konwersja is_done na boolean
        is_done = bool(is_done)

        logger.debug("API: Toggle product done", extra={
            'product_id': product_id,
            'station': station,
            'is_done': is_done,
            'client_ip': request.remote_addr
        })

        from ...models import ProductionItem

        # Znajdź produkt - preferuj record_id (PK) gdy obecny (disambig dla doróbek)
        record_id = data.get('record_id')
        if record_id is not None:
            product = ProductionItem.query.filter_by(id=int(record_id)).first()
        else:
            # Fallback do short_product_id (legacy)
            product = ProductionItem.query.filter_by(short_product_id=product_id).first()
        if not product:
            return jsonify({
                'success': False,
                'error': f'Produkt {product_id} nie znaleziony'
            }), 404

        # Ustaw odpowiednią kolumnę marked_done
        marked_done_field = f'{station}_marked_done'
        setattr(product, marked_done_field, is_done)
        product.updated_at = get_local_now()

        db.session.commit()

        logger.info("API: Product marked done toggled", extra={
            'product_id': product_id,
            'station': station,
            'is_done': is_done,
            'client_ip': request.remote_addr
        })

        return jsonify({
            'success': True,
            'product_id': product_id,
            'station': station,
            'is_done': is_done
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.error("API: Błąd toggle product done", extra={
            'product_id': data.get('product_id') if 'data' in locals() else None,
            'station': data.get('station') if 'data' in locals() else None,
            'client_ip': request.remote_addr,
            'error': str(e)
        })

        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# API ROUTERS - Update Quantity Done (Przyciski +/- na stanowiskach) - NOWY 2025-11
# ============================================================================


@api_bp.route('/update-quantity-done', methods=['POST'])
@ip_validation_required
def update_quantity_done():
    """
    POST /api/update-quantity-done - Zmiana ilości wykonanych sztuk na stanowisku

    Body JSON:
    {
        "product_id": "25_05248_1",
        "station": "cutting",
        "action": "increment" | "decrement" | "increment10" | "decrement10"
    }

    Akcje:
    - increment: quantity_done += 1 (max = quantity)
    - decrement: quantity_done -= 1 (min = 0)
    - increment10: quantity_done += 10 (max = quantity)
    - decrement10: quantity_done -= 10 (min = 0)

    Gdy quantity_done == quantity → ustawia {station}_completed_at
    Gdy quantity_done < quantity i było complete → czyści {station}_completed_at

    Autoryzacja: Walidacja IP (stanowiska)
    Returns: JSON ze zaktualizowanymi wartościami
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        product_id = data.get('product_id')
        station = data.get('station')
        action = data.get('action')

        # Walidacja wymaganych pól
        if not product_id or not station or not action:
            return jsonify({
                'success': False,
                'error': 'Wymagane pola: product_id, station, action'
            }), 400

        # Walidacja station
        valid_stations = ['cutting', 'assembly', 'gluing', 'formatting', 'finishing', 'painting', 'packaging']
        if station not in valid_stations:
            return jsonify({
                'success': False,
                'error': f'Nieprawidłowy station. Dozwolone: {valid_stations}'
            }), 400

        # Walidacja action
        valid_actions = ['increment', 'decrement', 'increment10', 'decrement10']
        if action not in valid_actions:
            return jsonify({
                'success': False,
                'error': f'Nieprawidłowa akcja. Dozwolone: {valid_actions}'
            }), 400

        logger.debug("API: Update quantity done", extra={
            'product_id': product_id,
            'station': station,
            'action': action,
            'client_ip': request.remote_addr
        })

        from ...models import ProductionItem

        # Znajdź produkt - preferuj record_id (PK) gdy obecny (disambig dla doróbek)
        record_id = data.get('record_id')
        if record_id is not None:
            product = ProductionItem.query.filter_by(id=int(record_id)).first()
        else:
            # Fallback do short_product_id (legacy)
            product = ProductionItem.query.filter_by(short_product_id=product_id).first()
        if not product:
            return jsonify({
                'success': False,
                'error': f'Produkt {product_id} nie znaleziony'
            }), 404

        # actor_user_id — bez tego praca z panelu nie ma autora w audycie
        actor_id = current_user.id if current_user.is_authenticated else None

        # Wykonaj akcję
        if action == 'increment':
            new_value = product.increment_quantity_done(
                station, 1, source='web', actor_user_id=actor_id)
        elif action == 'decrement':
            new_value = product.decrement_quantity_done(
                station, 1, source='web', actor_user_id=actor_id)
        elif action == 'increment10':
            new_value = product.increment_quantity_done(
                station, 10, source='web', actor_user_id=actor_id)
        elif action == 'decrement10':
            new_value = product.decrement_quantity_done(
                station, 10, source='web', actor_user_id=actor_id)

        db.session.commit()

        is_complete = (new_value == product.quantity)

        logger.info("API: Quantity done updated", extra={
            'product_id': product_id,
            'station': station,
            'action': action,
            'quantity_done': new_value,
            'quantity': product.quantity,
            'is_complete': is_complete,
            'client_ip': request.remote_addr
        })

        return jsonify({
            'success': True,
            'product_id': product_id,
            'station': station,
            'quantity_done': new_value,
            'quantity': product.quantity,
            'is_complete': is_complete
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.error("API: Błąd update quantity done", extra={
            'product_id': data.get('product_id') if 'data' in locals() else None,
            'station': data.get('station') if 'data' in locals() else None,
            'client_ip': request.remote_addr,
            'error': str(e)
        })

        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@api_bp.route('/products/<int:product_id>/reject', methods=['POST'])
def product_reject(product_id):
    """
    POST /api/products/<id>/reject - Cofnięcie sztuk produktu z formatowania.

    Body JSON: {
        quantity: int (>= 1),
        reason_category: 'wymiary'|'jakosc_sklejenia'|'jakosc_produktu'|'inne',
        station_code: 'formatting' (opcjonalne, MVP zawsze formatting)
    }

    Autoryzacja: brak (jak inne endpointy tabletów). Tablet wpisuje POST z widoku web.
    Returns: {original: serialized, rework: serialized, rework_log_id: int}
    """
    from ...services.rework_service import reject_product_quantity, RejectError

    data = request.get_json(silent=True) or {}
    quantity = data.get('quantity')
    reason_category = (data.get('reason_category') or '').strip()
    station_code = (data.get('station_code') or 'formatting').strip()

    try:
        original, rework, log_entry = reject_product_quantity(
            product_id=product_id,
            quantity=int(quantity) if quantity is not None else None,
            reason_category=reason_category,
            rejected_at_station=station_code,
            user_id=None,
            device_id=None,
        )
    except RejectError as e:
        return jsonify({'success': False, 'error': e.code, 'detail': e.message}), e.status
    except Exception as e:
        logger.error("Web API reject błąd", extra={'product_id': product_id}, exc_info=True)
        db.session.rollback()
        return jsonify({'success': False, 'error': 'reject_failed', 'detail': str(e)}), 500

    return jsonify({
        'success': True,
        'original': {
            'id': original.id,
            'short_product_id': original.short_product_id,
            'quantity': original.quantity,
            'quantity_done_formatting': original.quantity_done_formatting,
            'current_status': original.current_status,
        },
        'rework': {
            'id': rework.id,
            'short_product_id': rework.short_product_id,
            'quantity': rework.quantity,
            'current_status': rework.current_status,
            'is_rework': True,
            'original_product_id': original.id,
        },
        'rework_log_id': log_entry.id,
    }), 200


@api_bp.route('/get-cutting-progress', methods=['POST'])
@ip_validation_required
def get_cutting_progress():
    """
    POST /api/get-cutting-progress - Pobiera postęp wycinania dla listy produktów

    Używane przez stanowisko Składanie do synchronizacji live liczników wycinania.

    Body JSON:
    {
        "product_ids": ["25_05248_1", "25_05248_2", ...]
    }

    Returns: JSON z postępem wycinania dla każdego produktu
    {
        "success": true,
        "progress": {
            "25_05248_1": {
                "quantity_done_cutting": 5,
                "quantity": 10,
                "cutting_completed_at": "2025-12-17T14:30:00" | null
            },
            ...
        }
    }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        product_ids = data.get('product_ids', [])

        if not product_ids:
            return jsonify({'success': True, 'progress': {}}), 200

        if not isinstance(product_ids, list):
            return jsonify({'success': False, 'error': 'product_ids musi być listą'}), 400

        from ...models import ProductionItem

        products = ProductionItem.query.filter(
            ProductionItem.short_product_id.in_(product_ids)
        ).all()

        progress = {}
        for product in products:
            progress[product.short_product_id] = {
                'quantity_done_cutting': product.quantity_done_cutting or 0,
                'quantity': product.quantity or 1,
                'cutting_completed_at': product.cutting_completed_at.isoformat() if product.cutting_completed_at else None
            }

        return jsonify({
            'success': True,
            'progress': progress
        }), 200

    except Exception as e:
        logger.error("API: Błąd get-cutting-progress", extra={
            'client_ip': request.remote_addr,
            'error': str(e)
        })

        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@api_bp.route('/get-assembly-progress', methods=['POST'])
@ip_validation_required
def get_assembly_progress():
    """
    POST /api/get-assembly-progress - Pobiera postęp składania dla listy produktów

    Używane przez stanowisko Wycinanie do synchronizacji live liczników składania.

    Body JSON:
    {
        "product_ids": ["25_05248_1", "25_05248_2", ...]
    }

    Returns: JSON z postępem składania dla każdego produktu
    {
        "success": true,
        "progress": {
            "25_05248_1": {
                "quantity_done_assembly": 5,
                "quantity": 10,
                "assembly_completed_at": "2025-12-17T14:30:00" | null
            },
            ...
        }
    }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        product_ids = data.get('product_ids', [])

        if not product_ids:
            return jsonify({'success': True, 'progress': {}}), 200

        if not isinstance(product_ids, list):
            return jsonify({'success': False, 'error': 'product_ids musi być listą'}), 400

        from ...models import ProductionItem

        products = ProductionItem.query.filter(
            ProductionItem.short_product_id.in_(product_ids)
        ).all()

        progress = {}
        for product in products:
            progress[product.short_product_id] = {
                'quantity_done_assembly': product.quantity_done_assembly or 0,
                'quantity': product.quantity or 1,
                'assembly_completed_at': product.assembly_completed_at.isoformat() if product.assembly_completed_at else None
            }

        return jsonify({
            'success': True,
            'progress': progress
        }), 200

    except Exception as e:
        logger.error("API: Błąd get-assembly-progress", extra={
            'client_ip': request.remote_addr,
            'error': str(e)
        })

        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# API ROUTERS - PRD Section 6.3 (CRON i Synchronizacja)
# ============================================================================


@api_bp.route('/products-tab-content', methods=['GET'])
@login_required  
def products_tab_content():
    """
    Endpoint zwracający zawartość taba produktów - NAPRAWIONY
    BUGFIX: Usuwa limit 100 produktów, zwraca wszystkie produkty z parsowanymi polami
    """
    try:
        # Pobierz podstawowe parametry
        status_filter = request.args.get('status', 'all')
        search_query = request.args.get('search', '')
        load_all = request.args.get('load_all', 'true').lower() == 'true'

        # Tryb widoku: active (default) | archive | all
        # active  → ukrywa zamówienia, w których WSZYSTKIE pozycje mają status 'spakowane'
        # archive → pokazuje wyłącznie te zamówienia
        # all     → bez filtra
        view_mode = request.args.get('view', 'active').lower()
        if view_mode not in ('active', 'archive', 'all'):
            view_mode = 'active'

        # Filtr zakresu dat zakończenia (działa tylko w archive/all, na packaging_completed_at)
        completed_from_raw = request.args.get('completed_from', '').strip()
        completed_to_raw = request.args.get('completed_to', '').strip()
        completed_from_dt = None
        completed_to_dt = None
        if completed_from_raw:
            try:
                completed_from_dt = datetime.strptime(completed_from_raw, '%Y-%m-%d')
            except ValueError:
                completed_from_dt = None
        if completed_to_raw:
            try:
                completed_to_dt = datetime.strptime(completed_to_raw, '%Y-%m-%d') + timedelta(days=1)
            except ValueError:
                completed_to_dt = None

        # Pobierz produkty z bazy danych - BEZ LIMITU
        # Zawsze joinujemy ProductionOrder — potrzebne do filtrowania i wyszukiwania po polach order-level
        products_query = ProductionItem.query.join(ProductionOrder)

        # Subquery: numery zamówień, w których KAŻDA pozycja ma status 'spakowane'
        # (z agregacją MAX/MIN dat na poziomie zamówienia, do filtrowania po dacie zakończenia)
        if view_mode in ('active', 'archive'):
            order_agg_q = db.session.query(
                ProductionOrder.internal_order_number.label('ion'),
                func.max(ProductionItem.packaging_completed_at).label('order_completed_at'),
                func.sum(case((ProductionItem.current_status != 'spakowane', 1), else_=0)).label('non_packed_cnt')
            ).join(ProductionItem, ProductionItem.order_id == ProductionOrder.id).filter(
                ProductionOrder.internal_order_number.isnot(None)
            ).group_by(
                ProductionOrder.internal_order_number
            )

            # Zamówienie należy do archiwum, gdy:
            #  - WSZYSTKIE pozycje mają status 'spakowane' (zakończone produkcyjnie), LUB
            #  - WSZYSTKIE pozycje mają status 'anulowane' (całe zamówienie anulowane)
            fully_packed_cond = func.sum(
                case((ProductionItem.current_status != 'spakowane', 1), else_=0)
            ) == 0
            fully_cancelled_cond = func.sum(
                case((ProductionItem.current_status != 'anulowane', 1), else_=0)
            ) == 0
            archived_cond = or_(fully_packed_cond, fully_cancelled_cond)

            # archive: tylko zamówienia archiwalne + opcjonalnie zakres dat zakończenia (na poziomie zamówienia)
            if view_mode == 'archive':
                order_agg_q = order_agg_q.having(archived_cond)
                if completed_from_dt is not None:
                    order_agg_q = order_agg_q.having(
                        func.max(ProductionItem.packaging_completed_at) >= completed_from_dt
                    )
                if completed_to_dt is not None:
                    order_agg_q = order_agg_q.having(
                        func.max(ProductionItem.packaging_completed_at) < completed_to_dt
                    )
                matching_orders_subq = order_agg_q.subquery()
                products_query = products_query.filter(
                    ProductionOrder.internal_order_number.in_(
                        db.session.query(matching_orders_subq.c.ion)
                    )
                )
            else:  # active
                archived_subq = order_agg_q.having(archived_cond).subquery()
                products_query = products_query.filter(
                    ~ProductionOrder.internal_order_number.in_(
                        db.session.query(archived_subq.c.ion)
                    )
                )

        # Filtrowanie po statusie
        if status_filter and status_filter != 'all':
            products_query = products_query.filter(ProductionItem.current_status == status_filter)
        
        # Wyszukiwanie - bezpieczne sprawdzenie atrybutów
        if search_query:
            search_pattern = f"%{search_query}%"
            search_conditions = []
            
            if hasattr(ProductionItem, 'original_product_name'):
                search_conditions.append(ProductionItem.original_product_name.ilike(search_pattern))
            if hasattr(ProductionItem, 'short_product_id'):
                search_conditions.append(ProductionItem.short_product_id.ilike(search_pattern))
            # order-level fields via ProductionOrder (join already present)
            search_conditions.append(ProductionOrder.client_name.ilike(search_pattern))
            search_conditions.append(ProductionOrder.internal_order_number.ilike(search_pattern))
            search_conditions.append(ProductionOrder.client_order_number.ilike(search_pattern))
            search_conditions.append(ProductionOrder.quote_number.ilike(search_pattern))
            search_conditions.append(cast(ProductionOrder.baselinker_order_id, String).ilike(search_pattern))

            if search_conditions:
                products_query = products_query.filter(or_(*search_conditions))

        # Obsługa parametrów sortowania
        sort_by = request.args.get('sort_by', 'priority_rank')  # ZMIANA: domyślnie priority_rank
        sort_order = request.args.get('sort_order', 'asc' if sort_by == 'priority_rank' else 'desc')
        
        # Sortowanie z obsługą parametrów
        sort_column = None
        if sort_by == 'priority_rank' and hasattr(ProductionItem, 'priority_rank'):
            sort_column = ProductionItem.priority_rank
        elif sort_by == 'created_at' and hasattr(ProductionItem, 'created_at'):
            sort_column = ProductionItem.created_at
        elif sort_by == 'deadline_date' and hasattr(ProductionItem, 'deadline_date'):
            sort_column = ProductionItem.deadline_date
            
        if sort_column is not None:
            if sort_order == 'desc':
                products_query = products_query.order_by(sort_column.desc())
            else:
                products_query = products_query.order_by(sort_column.asc())
        else:
            # Fallback - domyślnie priority_rank ASC
            if hasattr(ProductionItem, 'priority_rank'):
                products_query = products_query.order_by(ProductionItem.priority_rank.asc())
            else:
                products_query = products_query.order_by(ProductionItem.id.desc())
        
        # ZMIANA: Pobierz WSZYSTKIE produkty (usuń limit)
        products = products_query.all()

        logger.info(f"Pobranych produktów: {len(products)} (bez limitu, view={view_mode})")

        # Pre-agregacja po zamówieniu: MAX(packaging_completed_at) i MIN(created_at)
        # potrzebne do pola order_completed_at oraz do statystyk archiwum.
        order_aggregates = {}
        for p in products:
            ion = p.order.internal_order_number if p.order else None
            if not ion:
                continue
            agg = order_aggregates.setdefault(ion, {'completed_at': None, 'created_at': None})
            pkg = getattr(p, 'packaging_completed_at', None)
            if pkg is not None and (agg['completed_at'] is None or pkg > agg['completed_at']):
                agg['completed_at'] = pkg
            crt = getattr(p, 'created_at', None)
            if crt is not None and (agg['created_at'] is None or crt < agg['created_at']):
                agg['created_at'] = crt

        # Renderuj HTML template (osobny komponent dla widoku archiwum)
        template_name = (
            'components/archive-tab-content.html'
            if view_mode == 'archive'
            else 'components/products-tab-content.html'
        )
        html_content = render_template(template_name)
        
        # Przygotuj dane produktów z bezpiecznym dostępem do atrybutów
        products_data = []
        for product in products:
            # Bezpieczne pobieranie wartości z fallback
            def get_attr(obj, attr_name, default=None):
                return getattr(obj, attr_name, default) if hasattr(obj, attr_name) else default
            
            # Oblicz dni do deadline
            days_to_deadline = None
            if hasattr(product, 'deadline_date') and product.deadline_date:
                try:
                    deadline = product.deadline_date
                    if isinstance(deadline, str):
                        deadline = datetime.strptime(deadline, '%Y-%m-%d').date()
                    days_to_deadline = (deadline - date.today()).days
                except:
                    days_to_deadline = None
            
            # Bezpieczne pobieranie objętości
            volume_m3 = 0.0
            try:
                vol = get_attr(product, 'volume_m3', 0)
                volume_m3 = float(vol) if vol is not None else 0.0
            except (ValueError, TypeError):
                volume_m3 = 0.0
            
            # Bezpieczne pobieranie wartości netto
            total_value_net = 0.0
            try:
                val = get_attr(product, 'total_value_net', 0)
                total_value_net = float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                total_value_net = 0.0
            
            # Bezpieczne pobieranie ceny jednostkowej
            unit_price_net = 0.0
            try:
                price = get_attr(product, 'unit_price_net', 0)
                unit_price_net = float(price) if price is not None else 0.0
            except (ValueError, TypeError):
                unit_price_net = 0.0
            
            # Bezpieczne pobieranie parsowanych wymiarów
            parsed_length_cm = 0.0
            parsed_width_cm = 0.0
            parsed_thickness_cm = 0.0
            try:
                if get_attr(product, 'parsed_length_cm') is not None:
                    parsed_length_cm = float(get_attr(product, 'parsed_length_cm', 0))
                if get_attr(product, 'parsed_width_cm') is not None:
                    parsed_width_cm = float(get_attr(product, 'parsed_width_cm', 0))
                if get_attr(product, 'parsed_thickness_cm') is not None:
                    parsed_thickness_cm = float(get_attr(product, 'parsed_thickness_cm', 0))
            except (ValueError, TypeError):
                pass
            
            product_dict = {
                # Podstawowe dane
                'id': product.id,
                'short_product_id': get_attr(product, 'short_product_id', ''),
                'original_product_name': get_attr(product, 'original_product_name', ''),
                'current_status': get_attr(product, 'current_status', 'czeka_na_wyciecie'),
                'priority_rank': get_attr(product, 'priority_rank', None),
                'priority_manual_override': get_attr(product, 'priority_manual_override', False),

                # Wymiary i wartości
                'volume_m3': volume_m3,
                'total_value_net': total_value_net,
                'unit_price_net': unit_price_net,
                
                # Deadline
                'deadline_date': get_attr(product, 'deadline_date').isoformat() if get_attr(product, 'deadline_date') else None,
                'days_until_deadline': days_to_deadline,
                
                # Dane klienta
                'client_name': (product.order.client_name if product.order else None) or '',
                'client_email': (product.order.client_email if product.order else None) or '',
                'client_phone': (product.order.client_phone if product.order else None) or '',
                'delivery_address': (product.order.delivery_address if product.order else None) or '',

                # Dane zamówienia
                'internal_order_number': (product.order.internal_order_number if product.order else None) or '',
                'baselinker_order_id': product.order.baselinker_order_id if product.order else None,
                'baselinker_product_id': get_attr(product, 'baselinker_product_id', ''),
                'product_sequence_in_order': get_attr(product, 'product_sequence_in_order', 1),
                'total_products_in_order': db.session.query(func.count(ProductionItem.id)).filter(
                    ProductionItem.order_id == product.order_id
                ).scalar() if product.order_id else 1,

                # POPRAWIONE: Specyfikacja produktu - PARSOWANE POLA z bazy danych
                'parsed_wood_species': product.configuration.species if product.configuration else None,
                'parsed_technology': product.configuration.technology if product.configuration else None,
                'parsed_wood_class': product.configuration.wood_class if product.configuration else None,
                'parsed_length_cm': parsed_length_cm,
                'parsed_width_cm': parsed_width_cm,
                'parsed_thickness_cm': parsed_thickness_cm,
                'parsed_finish_state': get_attr(product, 'parsed_finish_state', None),
                'parsed_edge_processing': get_attr(product, 'parsed_edge_processing', False),
                'cut_to_size': bool(get_attr(product, 'cut_to_size', True)),
                'shape': get_attr(product, 'shape', None),

                # Produkcja - statusy i czasami
                'cutting_started_at': get_attr(product, 'cutting_started_at').isoformat() if get_attr(product, 'cutting_started_at') else None,
                'cutting_completed_at': get_attr(product, 'cutting_completed_at').isoformat() if get_attr(product, 'cutting_completed_at') else None,
                'cutting_duration_minutes': get_attr(product, 'cutting_duration_minutes', None),
                'assembly_started_at': get_attr(product, 'assembly_started_at').isoformat() if get_attr(product, 'assembly_started_at') else None,
                'assembly_completed_at': get_attr(product, 'assembly_completed_at').isoformat() if get_attr(product, 'assembly_completed_at') else None,
                'assembly_duration_minutes': get_attr(product, 'assembly_duration_minutes', None),
                'packaging_started_at': get_attr(product, 'packaging_started_at').isoformat() if get_attr(product, 'packaging_started_at') else None,
                'packaging_completed_at': get_attr(product, 'packaging_completed_at').isoformat() if get_attr(product, 'packaging_completed_at') else None,
                'packaging_duration_minutes': get_attr(product, 'packaging_duration_minutes', None),

                # Logistyka
                'logistics_completed_at': product.order.logistics_completed_at.isoformat() if product.order and product.order.logistics_completed_at else None,
                'painting_completed_at': get_attr(product, 'painting_completed_at').isoformat() if get_attr(product, 'painting_completed_at') else None,
                'override_delivery_method': product.order.override_delivery_method if product.order else None,
                'is_personal_pickup': product.order.is_personal_pickup if product.order else None,

                # Przypisani pracownicy
                'cutting_assigned_worker_id': get_attr(product, 'cutting_assigned_worker_id', None),
                'assembly_assigned_worker_id': get_attr(product, 'assembly_assigned_worker_id', None),
                'packaging_assigned_worker_id': get_attr(product, 'packaging_assigned_worker_id', None),
                
                # Notatki i problemy
                'production_notes': get_attr(product, 'production_notes', ''),
                'quality_issues': get_attr(product, 'quality_issues', ''),
                
                # Timestampy
                'created_at': product.created_at.isoformat() if hasattr(product, 'created_at') and product.created_at else None,
                'updated_at': product.updated_at.isoformat() if hasattr(product, 'updated_at') and product.updated_at else None,
                'sync_source': product.order.sync_source if product.order else None,

                # Załączniki
                'attachment_file_name': product.order.attachment_file_name if product.order else None,
                'attachment_file_url': product.order.attachment_file_url if product.order else None,

                # NOWE: Quantity fields (2025-11)
                'quantity': get_attr(product, 'quantity', 1),
                'quantity_done_cutting': get_attr(product, 'quantity_done_cutting', 0),
                'quantity_done_assembly': get_attr(product, 'quantity_done_assembly', 0),
                'quantity_done_gluing': get_attr(product, 'quantity_done_gluing', 0),
                'quantity_done_formatting': get_attr(product, 'quantity_done_formatting', 0),
                'quantity_done_finishing': get_attr(product, 'quantity_done_finishing', 0),
                'quantity_done_painting': get_attr(product, 'quantity_done_painting', 0),
                'quantity_done_packaging': get_attr(product, 'quantity_done_packaging', 0),

                # Dodatkowe pola z zamówienia
                'client_order_number': product.order.client_order_number if product.order else None,
                'quote_number': product.order.quote_number if product.order else None,
                'order_notes': product.order.order_notes if product.order else None,

                # Priorytet ręczny (gwiazdka)
                'is_priority': get_attr(product, 'is_priority', False),

                # Unique identifier dla frontend
                'unique_id': f"{get_attr(product, 'short_product_id', '')}-{product.id}"
            }

            # Data zakończenia zamówienia (MAX packaging_completed_at po wszystkich pozycjach
            # tego samego internal_order_number) — używane głównie w widoku archiwum.
            ion = product_dict['internal_order_number']
            agg = order_aggregates.get(ion) if ion else None
            product_dict['order_completed_at'] = (
                agg['completed_at'].isoformat() if agg and agg['completed_at'] else None
            )

            products_data.append(product_dict)
        
        # Statystyki — w widoku 'active' liczone per niespakowana sztuka
        # (remaining = quantity - quantity_done_packaging). Pozycje w pełni spakowane
        # są ignorowane mimo że należą do nieukończonych jeszcze zamówień. W widoku
        # 'archive' (wszystko spakowane) pokazujemy pełne wartości historyczne.
        if view_mode == 'archive':
            total_volume = sum(p['volume_m3'] * p['quantity'] for p in products_data)
            total_value = sum(p['total_value_net'] for p in products_data)
            total_quantity = sum(p['quantity'] for p in products_data)
            total_count = len(products_data)
            stats_breakdown_source = products_data
        else:
            total_volume = 0.0
            total_value = 0.0
            total_quantity = 0
            total_count = 0
            stats_breakdown_source = []
            for p in products_data:
                if p.get('current_status') in ('spakowane', 'anulowane'):
                    continue
                qty = int(p.get('quantity') or 0)
                done = int(p.get('quantity_done_packaging') or 0)
                remaining = qty - done
                if remaining <= 0:
                    continue
                ratio = remaining / qty if qty else 0
                total_volume += float(p.get('volume_m3') or 0) * remaining
                total_value += float(p.get('total_value_net') or 0) * ratio
                total_quantity += remaining
                total_count += 1
                stats_breakdown_source.append(p)

        # Oblicz pilne produkty (deadline <= 3 dni) — tylko z policzonych pozycji
        urgent_count = 0
        for p in stats_breakdown_source:
            if p['days_until_deadline'] is not None and p['days_until_deadline'] <= 3:
                urgent_count += 1

        # Breakdown statusów
        status_breakdown = {}
        for p in stats_breakdown_source:
            status = p['current_status']
            status_breakdown[status] = status_breakdown.get(status, 0) + 1

        stats_data = {
            'total_count': total_count,
            'total_quantity': total_quantity,
            'total_volume': round(total_volume, 4),
            'total_value': round(total_value, 2),
            'urgent_count': urgent_count,
            'status_breakdown': status_breakdown
        }

        # Statystyki specyficzne dla widoku archiwum:
        # liczba spakowanych zamówień + średni czas realizacji w dniach
        # (od MIN(created_at) do MAX(packaging_completed_at) per zamówienie).
        if view_mode == 'archive':
            durations_days = []
            for ion, agg in order_aggregates.items():
                if agg['completed_at'] and agg['created_at']:
                    delta = agg['completed_at'] - agg['created_at']
                    if delta.total_seconds() >= 0:
                        durations_days.append(delta.total_seconds() / 86400.0)
            avg_realization_days = (
                round(sum(durations_days) / len(durations_days), 1)
                if durations_days else None
            )
            stats_data['archive'] = {
                'orders_count': len(order_aggregates),
                'avg_realization_days': avg_realization_days
            }
        
        # Przygotuj opcje filtrów z produktów
        filters_data = {
            'wood_species': list(set(p['parsed_wood_species'] for p in products_data if p['parsed_wood_species'])),
            'technologies': list(set(p['parsed_technology'] for p in products_data if p['parsed_technology'])),
            'wood_classes': list(set(p['parsed_wood_class'] for p in products_data if p['parsed_wood_class'])),
            'thicknesses': list(set(f"{p['parsed_thickness_cm']}cm" for p in products_data if p['parsed_thickness_cm'] and p['parsed_thickness_cm'] > 0)),
            'statuses': list(set(p['current_status'] for p in products_data if p['current_status']))
        }
        
        # Sortuj opcje filtrów
        filters_data['wood_species'].sort()
        filters_data['technologies'].sort()
        filters_data['wood_classes'].sort()
        filters_data['thicknesses'].sort(key=lambda x: float(x.replace('cm', '')))
        filters_data['statuses'].sort()
        
        logger.info(f"Statystyki: {stats_data}")
        logger.info(f"Opcje filtrów: gatunki={len(filters_data['wood_species'])}, technologie={len(filters_data['technologies'])}")
        
        return jsonify({
            'success': True,
            'html': html_content,
            'initial_data': {
                'products': products_data,
                'stats': stats_data,
                'filters': filters_data,
                'total_count': len(products_data),
                'load_all': load_all
            },
            'products_count': len(products),
            'debug_info': {
                'status_filter': status_filter,
                'search_query': search_query,
                'products_returned': len(products_data),
                'load_all': load_all,
                'query_without_limit': True,
                'parsed_fields_included': True,
                'filters_data_included': True
            }
        })
        
    except Exception as e:
        # Szczegółowe logowanie błędu
        error_traceback = traceback.format_exc()
        logger.error(f"Błąd endpoint products-tab-content: {str(e)}")
        logger.error(f"Traceback: {error_traceback}")
        
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': error_traceback if current_app.debug else None
        }), 500


# Dla paginowanych zapytań

@api_bp.route('/products-paginated', methods=['GET'])
@login_required
def products_paginated():
    """
    Endpoint dla paginowanych produktów - dla bardzo dużych zbiorów danych
    """
    try:
        # Parametry paginacji
        page = int(request.args.get('page', 1))
        per_page = min(int(request.args.get('per_page', 50)), 800)  # Max 800 na stronę
        
        # Filtry
        status_filter = request.args.get('status', 'all')
        search_query = request.args.get('search', '')
        
        # Query builder — join ProductionOrder needed for order-level field searches
        products_query = ProductionItem.query.join(ProductionOrder)

        if status_filter and status_filter != 'all':
            products_query = products_query.filter(ProductionItem.current_status == status_filter)

        if search_query:
            search_pattern = f"%{search_query}%"
            search_conditions = []

            if hasattr(ProductionItem, 'original_product_name'):
                search_conditions.append(ProductionItem.original_product_name.ilike(search_pattern))
            if hasattr(ProductionItem, 'short_product_id'):
                search_conditions.append(ProductionItem.short_product_id.ilike(search_pattern))
            # order-level fields via ProductionOrder (join already present)
            search_conditions.append(ProductionOrder.client_name.ilike(search_pattern))
            search_conditions.append(ProductionOrder.internal_order_number.ilike(search_pattern))
            search_conditions.append(ProductionOrder.client_order_number.ilike(search_pattern))
            search_conditions.append(ProductionOrder.quote_number.ilike(search_pattern))
            search_conditions.append(cast(ProductionOrder.baselinker_order_id, String).ilike(search_pattern))

            if search_conditions:
                products_query = products_query.filter(or_(*search_conditions))

        # Sortowanie
        products_query = products_query.order_by(ProductionItem.priority_rank.asc())
        
        # Paginacja
        paginated = products_query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        # Przygotuj dane
        products_data = []
        for product in paginated.items:
            # Używamy tej samej logiki co w głównym endpoint
            product_dict = {
                'id': product.id,
                'short_product_id': getattr(product, 'short_product_id', ''),
                'original_product_name': getattr(product, 'original_product_name', ''),
                'current_status': getattr(product, 'current_status', 'nieznany'),
                'priority_rank': getattr(product, 'priority_rank', None),
                'unique_id': f"{getattr(product, 'short_product_id', '')}-{product.id}"
                # ... reszta pól
            }
            products_data.append(product_dict)
        
        return jsonify({
            'success': True,
            'products': products_data,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': paginated.total,
                'pages': paginated.pages,
                'has_next': paginated.has_next,
                'has_prev': paginated.has_prev
            }
        })
        
    except Exception as e:
        logger.error(f"Błąd endpoint products-paginated: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# API ADMIN - Update Quantity Done (Panel admina - bez restrykcji IP)
# ============================================================================


@api_bp.route('/admin/update-quantity-done', methods=['POST'])
@login_required
def admin_update_quantity_done():
    """
    POST /api/admin/update-quantity-done - Zmiana ilości wykonanych sztuk z panelu admina

    Body JSON:
    {
        "product_id": "25_05248_1",
        "station": "cutting",
        "action": "increment" | "decrement" | "increment10" | "decrement10" | "set",
        "value": 5  // tylko dla action="set"
    }

    Akcje:
    - increment: quantity_done += 1 (max = quantity)
    - decrement: quantity_done -= 1 (min = 0)
    - increment10: quantity_done += 10 (max = quantity)
    - decrement10: quantity_done -= 10 (min = 0)
    - set: quantity_done = value (0 <= value <= quantity)

    Autoryzacja: login_required (panel admina)
    Returns: JSON ze zaktualizowanymi wartościami
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        product_id = data.get('product_id')
        station = data.get('station')
        action = data.get('action')
        value = data.get('value')

        # Walidacja wymaganych pól
        if not product_id or not station or not action:
            return jsonify({
                'success': False,
                'error': 'Wymagane pola: product_id, station, action'
            }), 400

        # Walidacja station
        valid_stations = ['cutting', 'assembly', 'gluing', 'formatting', 'finishing', 'packaging']
        if station not in valid_stations:
            return jsonify({
                'success': False,
                'error': f'Nieprawidłowy station. Dozwolone: {valid_stations}'
            }), 400

        # Walidacja action
        valid_actions = ['increment', 'decrement', 'increment10', 'decrement10', 'set']
        if action not in valid_actions:
            return jsonify({
                'success': False,
                'error': f'Nieprawidłowa akcja. Dozwolone: {valid_actions}'
            }), 400

        logger.info("API Admin: Update quantity done", extra={
            'product_id': product_id,
            'station': station,
            'action': action,
            'value': value,
            'user_id': current_user.id,
            'user_email': current_user.email
        })

        from ...models import ProductionItem

        # Znajdź produkt - preferuj record_id (PK) gdy obecny (disambig dla doróbek)
        record_id = data.get('record_id')
        if record_id is not None:
            product = ProductionItem.query.filter_by(id=int(record_id)).first()
        else:
            # Fallback do short_product_id (legacy)
            product = ProductionItem.query.filter_by(short_product_id=product_id).first()
        if not product:
            return jsonify({
                'success': False,
                'error': f'Produkt {product_id} nie znaleziony'
            }), 404

        # Akcja z panelu admina — autor: zalogowany user
        actor_kwargs = {
            'actor_user_id': current_user.id if current_user.is_authenticated else None,
            'source': 'admin',
        }

        # Wykonaj akcję
        if action == 'increment':
            new_value = product.increment_quantity_done(station, 1, **actor_kwargs)
        elif action == 'decrement':
            new_value = product.decrement_quantity_done(station, 1, **actor_kwargs)
        elif action == 'increment10':
            new_value = product.increment_quantity_done(station, 10, **actor_kwargs)
        elif action == 'decrement10':
            new_value = product.decrement_quantity_done(station, 10, **actor_kwargs)
        elif action == 'set':
            # Ustaw konkretną wartość
            if value is None:
                return jsonify({
                    'success': False,
                    'error': 'Akcja "set" wymaga pola "value"'
                }), 400
            value = int(value)
            if value < 0 or value > product.quantity:
                return jsonify({
                    'success': False,
                    'error': f'Wartość musi być między 0 a {product.quantity}'
                }), 400
            new_value = product.set_quantity_done(station, value, **actor_kwargs)

        db.session.commit()

        is_complete = (new_value == product.quantity)

        # Zbierz wszystkie wartości quantity_done
        all_quantity_done = {
            'quantity_done_cutting': product.quantity_done_cutting or 0,
            'quantity_done_assembly': product.quantity_done_assembly or 0,
            'quantity_done_gluing': product.quantity_done_gluing or 0,
            'quantity_done_formatting': product.quantity_done_formatting or 0,
            'quantity_done_finishing': product.quantity_done_finishing or 0,
            'quantity_done_packaging': product.quantity_done_packaging or 0
        }

        logger.info("API Admin: Quantity done updated", extra={
            'product_id': product_id,
            'station': station,
            'action': action,
            'quantity_done': new_value,
            'quantity': product.quantity,
            'is_complete': is_complete
        })

        return jsonify({
            'success': True,
            'product_id': product_id,
            'station': station,
            'quantity_done': new_value,
            'quantity': product.quantity,
            'is_complete': is_complete,
            'all_quantity_done': all_quantity_done
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.error("API Admin: Błąd update quantity done", extra={
            'product_id': data.get('product_id') if 'data' in locals() else None,
            'station': data.get('station') if 'data' in locals() else None,
            'error': str(e)
        })
        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500


# ============================================================================
# API ADMIN - Baselinker Order Comparison & Update
# ============================================================================


@api_bp.route('/admin/compare-baselinker-order', methods=['POST'])
@login_required
def admin_compare_baselinker_order():
    """
    POST /api/admin/compare-baselinker-order - Porównuje zamówienie z danymi z Baselinker

    Body JSON:
    {
        "baselinker_order_id": 25208907
    }

    Returns: JSON ze strukturą zmian do przeglądu
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        baselinker_order_id = data.get('baselinker_order_id')
        if not baselinker_order_id:
            return jsonify({
                'success': False,
                'error': 'Wymagane pole: baselinker_order_id'
            }), 400

        # Konwertuj na int
        try:
            baselinker_order_id = int(baselinker_order_id)
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'baselinker_order_id musi być liczbą'
            }), 400

        logger.info("API Admin: Porównanie z Baselinker", extra={
            'baselinker_order_id': baselinker_order_id,
            'user_id': current_user.id
        })

        from ...services.sync_service import BaselinkerSyncService

        sync_service = BaselinkerSyncService()
        result = sync_service.compare_order_with_baselinker(baselinker_order_id)

        return jsonify(result), 200

    except Exception as e:
        logger.error("API Admin: Błąd porównania z Baselinker", extra={
            'baselinker_order_id': data.get('baselinker_order_id') if 'data' in locals() else None,
            'error': str(e)
        })
        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500



@api_bp.route('/admin/apply-baselinker-changes', methods=['POST'])
@login_required
def admin_apply_baselinker_changes():
    """
    POST /api/admin/apply-baselinker-changes - Aplikuje zmiany z Baselinker do bazy

    Body JSON:
    {
        "baselinker_order_id": 25208907,
        "changes": {
            "products_to_add": [...],
            "products_to_remove": [...],
            "products_to_update": [...],
            "order_level": [...]
        }
    }

    Returns: JSON z wynikiem operacji
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        baselinker_order_id = data.get('baselinker_order_id')
        changes = data.get('changes')

        if not baselinker_order_id or not changes:
            return jsonify({
                'success': False,
                'error': 'Wymagane pola: baselinker_order_id, changes'
            }), 400

        # Konwertuj na int
        try:
            baselinker_order_id = int(baselinker_order_id)
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'baselinker_order_id musi być liczbą'
            }), 400

        logger.info("API Admin: Aplikowanie zmian z Baselinker", extra={
            'baselinker_order_id': baselinker_order_id,
            'user_id': current_user.id,
            'changes_summary': {
                'to_add': len(changes.get('products_to_add', [])),
                'to_remove': len(changes.get('products_to_remove', [])),
                'to_update': len(changes.get('products_to_update', []))
            }
        })

        from ...services.sync_service import BaselinkerSyncService

        sync_service = BaselinkerSyncService()
        result = sync_service.apply_baselinker_changes(baselinker_order_id, changes)

        return jsonify(result), 200

    except Exception as e:
        db.session.rollback()
        logger.error("API Admin: Błąd aplikowania zmian z Baselinker", extra={
            'baselinker_order_id': data.get('baselinker_order_id') if 'data' in locals() else None,
            'error': str(e)
        })
        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500



@api_bp.route('/products/bulk-action', methods=['POST'])
@login_required
def bulk_action():
    """
    POST /production/api/products/bulk-action
    
    Wykonuje masowe operacje na produktach
    
    Body (JSON):
    {
        "action": "update_status|update_priority|export|delete",
        "product_ids": [1, 2, 3, ...],
        "parameters": {
            "new_status": "czeka_na_wyciecie",
            "new_priority": 150,
            "export_format": "excel"
        }
    }
    
    Returns: JSON z rezultatem operacji
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400
        
        action = data.get('action')
        product_ids = data.get('product_ids', [])
        parameters = data.get('parameters', {})
        
        # Walidacja
        if not action or not product_ids:
            return jsonify({'success': False, 'error': 'Wymagane: action i product_ids'}), 400
        
        valid_actions = ['update_status', 'update_priority', 'export', 'delete']
        if action not in valid_actions:
            return jsonify({'success': False, 'error': f'Nieprawidłowa akcja. Dostępne: {valid_actions}'}), 400
        
        from ...models import ProductionItem
        
        # Pobierz produkty
        products = ProductionItem.query.filter(ProductionItem.id.in_(product_ids)).all()
        
        if not products:
            return jsonify({'success': False, 'error': 'Nie znaleziono produktów'}), 404
        
        results = {
            'success': True,
            'action': action,
            'processed_count': 0,
            'failed_count': 0,
            'errors': []
        }
        
        # Wykonaj akcję na każdym produkcie
        for product in products:
            try:
                if action == 'update_status':
                    new_status = parameters.get('new_status')
                    if new_status and hasattr(ProductionItem, 'current_status'):
                        product.current_status = new_status
                        results['processed_count'] += 1
                
                elif action == 'update_priority':
                    new_priority = parameters.get('new_priority')
                    if new_priority is not None:
                        product.priority_rank = int(new_priority)
                        results['processed_count'] += 1
                
                elif action == 'delete':
                    # Tylko admin może usuwać
                    if not (hasattr(current_user, 'role') and current_user.role.lower() in ['admin', 'administrator']):
                        results['errors'].append(f'Brak uprawnień do usunięcia produktu {product.id}')
                        results['failed_count'] += 1
                        continue
                    
                    db.session.delete(product)
                    results['processed_count'] += 1
                
                elif action == 'export':
                    # Export będzie obsłużony w osobnym endpoincie
                    results['processed_count'] += 1
                
            except Exception as e:
                logger.error(f"Błąd bulk action dla produktu {product.id}", extra={'error': str(e)})
                results['errors'].append(f'Błąd produktu {product.id}: {str(e)}')
                results['failed_count'] += 1
        
        # Zapisz zmiany dla akcji modyfikujących
        if action in ['update_status', 'update_priority', 'delete']:
            db.session.commit()
        
        logger.info("Bulk action wykonana", extra={
            'user_id': current_user.id,
            'action': action,
            'product_count': len(product_ids),
            'processed': results['processed_count'],
            'failed': results['failed_count']
        })
        
        return jsonify(results)
        
    except Exception as e:
        db.session.rollback()
        logger.error("Błąd bulk action", extra={
            'user_id': current_user.id if current_user.is_authenticated else None,
            'error': str(e)
        })
        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500


# 3. EXPORT PRODUKTÓW ENDPOINT

@api_bp.route('/products/export', methods=['POST'])
@login_required
def export_products():
    """
    POST /production/api/products/export

    Generuje export produktów w różnych formatach (Excel, CSV, PDF)

    Body (JSON):
    {
        "format": "excel|csv|pdf",
        "product_ids": [1, 2, 3] | "all" | "filtered",
        "filters": {...},
        "report_type": "full|summary"  // dla PDF
    }

    Returns: Plik do pobrania
    """
    try:
        import io
        import csv
        from datetime import datetime
        from flask import send_file, make_response

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400

        export_format = data.get('format', 'excel').lower()
        product_selection = data.get('product_ids', 'all')
        filters = data.get('filters', {})
        report_type = data.get('report_type', 'full')

        # Walidacja formatu
        valid_formats = ['excel', 'csv', 'pdf']
        if export_format not in valid_formats:
            return jsonify({'success': False, 'error': f'Nieprawidłowy format. Dostępne: {valid_formats}'}), 400

        from ...models import ProductionItem

        # Buduj query na podstawie selekcji
        query = ProductionItem.query

        if isinstance(product_selection, list):
            query = query.filter(ProductionItem.id.in_(product_selection))
        elif product_selection == "filtered":
            if filters.get('status'):
                query = query.filter(ProductionItem.current_status == filters['status'])
            if filters.get('search'):
                search_term = f"%{filters['search']}%"
                query = query.filter(
                    db.or_(
                        ProductionItem.short_product_id.ilike(search_term),
                        ProductionItem.original_product_name.ilike(search_term)
                    )
                )

        # Sortuj po priorytecie (MariaDB nie obsługuje NULLS LAST)
        # NULL-e na końcu: COALESCE(priority_rank, 999999)
        query = query.order_by(
            func.coalesce(ProductionItem.priority_rank, 999999).asc(),
            ProductionItem.created_at.desc()
        )
        products = query.all()

        if not products:
            return jsonify({'success': False, 'error': 'Brak produktów do eksportu'}), 404

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # =====================================================================
        # EXPORT CSV
        # =====================================================================
        if export_format == 'csv':
            return _export_csv(products, timestamp)

        # =====================================================================
        # EXPORT EXCEL (z openpyxl)
        # =====================================================================
        elif export_format == 'excel':
            return _export_excel(products, timestamp)

        # =====================================================================
        # EXPORT PDF (z reportlab)
        # =====================================================================
        elif export_format == 'pdf':
            return _export_pdf(products, timestamp, report_type)

    except Exception as e:
        logger.error("Błąd eksportu produktów", extra={
            'user_id': current_user.id if current_user.is_authenticated else None,
            'error': str(e),
            'traceback': traceback.format_exc()
        })
        return jsonify({
            'success': False,
            'error': f'Błąd eksportu: {str(e)}'
        }), 500



def _export_csv(products, timestamp):
    """Eksport do CSV"""
    import io
    import csv
    from flask import make_response

    output = io.StringIO()

    fieldnames = ['ID', 'Zamówienie', 'Nr klienta', 'Nazwa', 'Status', 'Priorytet',
                  'Ilość', 'Gatunek', 'Technologia', 'Klasa', 'Wymiary',
                  'Objętość m³', 'Klient', 'Data utworzenia']

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for p in products:
        writer.writerow({
            'ID': p.short_product_id,
            'Zamówienie': p.order.internal_order_number if p.order else None,
            'Nr klienta': (p.order.client_order_number if p.order else None) or '',
            'Nazwa': p.original_product_name,
            'Status': _format_status(p.current_status),
            'Priorytet': p.priority_rank or '-',
            'Ilość': p.quantity,
            'Gatunek': (p.configuration.species if p.configuration else None) or '',
            'Technologia': (p.configuration.technology if p.configuration else None) or '',
            'Klasa': (p.configuration.wood_class if p.configuration else None) or '',
            'Wymiary': f"{p.parsed_length_cm or 0}×{p.parsed_width_cm or 0}×{p.parsed_thickness_cm or 0}",
            'Objętość m³': float(p.volume_m3) if p.volume_m3 else 0,
            'Klient': (p.order.client_name if p.order else None) or '',
            'Data utworzenia': p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else ''
        })

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=raport_produkcji_{timestamp}.csv'

    logger.info("Export CSV wygenerowany", extra={
        'user_id': current_user.id,
        'products_count': len(products)
    })

    return response



def _export_excel(products, timestamp):
    """Eksport do Excel z pełnymi danymi analitycznymi w wielu arkuszach"""
    import io
    from collections import defaultdict

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.label import DataLabelList
    except ImportError:
        logger.warning("openpyxl niedostępne, używam CSV jako fallback")
        return _export_csv(products, timestamp)

    wb = Workbook()

    # =========================================================================
    # STYLE WSPÓLNE
    # =========================================================================
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_blue = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    status_colors = {
        'czeka_na_wyciecie': 'FFF3E0',
        'czeka_na_skladanie': 'E3F2FD',
        'czeka_na_sklejanie': 'F3E5F5',
        'czeka_na_formatowanie': 'E8F5E9',
        'czeka_na_wykanczanie': 'FFF8E1',
        'czeka_na_pakowanie': 'E0F7FA',
        'spakowane': 'C8E6C9',
        'anulowane': 'FFCDD2',
        'wstrzymane': 'CFD8DC'
    }

    # =========================================================================
    # PRZYGOTOWANIE DANYCH ANALITYCZNYCH
    # =========================================================================
    total_volume = sum(float(p.volume_m3 or 0) * (p.quantity or 1) for p in products)
    total_qty = sum(p.quantity for p in products)

    # Agregacje
    status_stats = defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0})
    species_stats = defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0})
    technology_stats = defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0})
    wood_class_stats = defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0})
    thickness_stats = defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0})

    for p in products:
        vol = float(p.volume_m3 or 0)
        qty = p.quantity or 1

        status_stats[p.current_status]['count'] += 1
        status_stats[p.current_status]['qty'] += qty
        status_stats[p.current_status]['volume'] += vol * qty

        species = (p.configuration.species if p.configuration else None) or 'Nieokreślony'
        species_stats[species]['count'] += 1
        species_stats[species]['qty'] += qty
        species_stats[species]['volume'] += vol * qty

        tech = (p.configuration.technology if p.configuration else None) or 'Nieokreślona'
        technology_stats[tech]['count'] += 1
        technology_stats[tech]['qty'] += qty
        technology_stats[tech]['volume'] += vol * qty

        wood_class = (p.configuration.wood_class if p.configuration else None) or 'Nieokreślona'
        wood_class_stats[wood_class]['count'] += 1
        wood_class_stats[wood_class]['qty'] += qty
        wood_class_stats[wood_class]['volume'] += vol * qty

        thickness = f"{float(p.parsed_thickness_cm):.1f} cm" if p.parsed_thickness_cm else 'Nieokreślona'
        thickness_stats[thickness]['count'] += 1
        thickness_stats[thickness]['qty'] += qty
        thickness_stats[thickness]['volume'] += vol * qty

    # =========================================================================
    # ARKUSZ 1: PODSUMOWANIE
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Podsumowanie"

    # Tytuł
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = f"RAPORT PRODUKCJI - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_summary['A1'].font = Font(bold=True, size=18, color="2E7D32")
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[1].height = 40

    # Główne podsumowanie
    ws_summary['A3'] = "PODSUMOWANIE OGÓLNE"
    ws_summary['A3'].font = Font(bold=True, size=14, color="1565C0")

    summary_data = [
        ['Parametr', 'Wartość'],
        ['Liczba pozycji', len(products)],
        ['Łączna ilość sztuk', total_qty],
        ['Łączna objętość (m³)', round(total_volume, 3)],
    ]

    for row_idx, row_data in enumerate(summary_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 5:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = number_alignment if col_idx == 2 and row_idx > 5 else cell_alignment

    # Rozkład statusów
    ws_summary['A11'] = "ROZKŁAD WEDŁUG STATUSÓW"
    ws_summary['A11'].font = Font(bold=True, size=14, color="1565C0")

    status_headers = ['Status', 'Pozycje', 'Sztuki', 'Objętość (m³)', '% pozycji']
    for col_idx, header in enumerate(status_headers, 1):
        cell = ws_summary.cell(row=13, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.border = thin_border
        cell.alignment = header_alignment

    row_idx = 14
    for status, data in sorted(status_stats.items(), key=lambda x: x[1]['count'], reverse=True):
        pct = (data['count'] / len(products) * 100) if products else 0
        row_data = [_format_status(status), data['count'], data['qty'], round(data['volume'], 3), round(pct, 1)]
        status_color = status_colors.get(status, 'FFFFFF')
        row_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = number_alignment if col_idx > 1 else cell_alignment
        row_idx += 1

    # Szerokości kolumn
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 12

    # =========================================================================
    # ARKUSZ 2: ANALIZA GATUNKÓW
    # =========================================================================
    ws_species = wb.create_sheet("Gatunki drewna")
    ws_species['A1'] = "ANALIZA WEDŁUG GATUNKÓW DREWNA"
    ws_species['A1'].font = Font(bold=True, size=14, color="2E7D32")

    species_headers = ['Gatunek', 'Pozycje', 'Sztuki', 'Objętość (m³)', '% objętości']
    for col_idx, header in enumerate(species_headers, 1):
        cell = ws_species.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 4
    for species, data in sorted(species_stats.items(), key=lambda x: x[1]['volume'], reverse=True):
        pct = (data['volume'] / total_volume * 100) if total_volume else 0
        row_data = [species, data['count'], data['qty'], round(data['volume'], 3), round(pct, 1)]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_species.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = number_alignment if col_idx > 1 else cell_alignment
        row_idx += 1

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_species.column_dimensions[col].width = 18

    # =========================================================================
    # ARKUSZ 3: ANALIZA TECHNOLOGII
    # =========================================================================
    ws_tech = wb.create_sheet("Technologie")
    ws_tech['A1'] = "ANALIZA WEDŁUG TECHNOLOGII"
    ws_tech['A1'].font = Font(bold=True, size=14, color="2E7D32")

    tech_headers = ['Technologia', 'Pozycje', 'Sztuki', 'Objętość (m³)', '% objętości']
    for col_idx, header in enumerate(tech_headers, 1):
        cell = ws_tech.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 4
    for tech, data in sorted(technology_stats.items(), key=lambda x: x[1]['volume'], reverse=True):
        pct = (data['volume'] / total_volume * 100) if total_volume else 0
        row_data = [tech, data['count'], data['qty'], round(data['volume'], 3), round(pct, 1)]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_tech.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = number_alignment if col_idx > 1 else cell_alignment
        row_idx += 1

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_tech.column_dimensions[col].width = 18

    # =========================================================================
    # ARKUSZ 4: ANALIZA KLAS DREWNA
    # =========================================================================
    ws_class = wb.create_sheet("Klasy drewna")
    ws_class['A1'] = "ANALIZA WEDŁUG KLAS DREWNA"
    ws_class['A1'].font = Font(bold=True, size=14, color="2E7D32")

    class_headers = ['Klasa', 'Pozycje', 'Sztuki', 'Objętość (m³)', '% objętości']
    for col_idx, header in enumerate(class_headers, 1):
        cell = ws_class.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 4
    for wood_class, data in sorted(wood_class_stats.items(), key=lambda x: x[1]['volume'], reverse=True):
        pct = (data['volume'] / total_volume * 100) if total_volume else 0
        row_data = [wood_class, data['count'], data['qty'], round(data['volume'], 3), round(pct, 1)]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_class.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = number_alignment if col_idx > 1 else cell_alignment
        row_idx += 1

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_class.column_dimensions[col].width = 18

    # =========================================================================
    # ARKUSZ 5: ANALIZA GRUBOŚCI
    # =========================================================================
    ws_thick = wb.create_sheet("Grubości")
    ws_thick['A1'] = "ANALIZA WEDŁUG GRUBOŚCI"
    ws_thick['A1'].font = Font(bold=True, size=14, color="2E7D32")

    thick_headers = ['Grubość', 'Pozycje', 'Sztuki', 'Objętość (m³)', '% objętości']
    for col_idx, header in enumerate(thick_headers, 1):
        cell = ws_thick.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 4
    # Sortuj po grubości numerycznie
    sorted_thickness = sorted(thickness_stats.items(),
                              key=lambda x: float(x[0].replace(' cm', '').replace('Nieokreślona', '0')))
    for thickness, data in sorted_thickness:
        pct = (data['volume'] / total_volume * 100) if total_volume else 0
        row_data = [thickness, data['count'], data['qty'], round(data['volume'], 3), round(pct, 1)]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_thick.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = number_alignment if col_idx > 1 else cell_alignment
        row_idx += 1

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_thick.column_dimensions[col].width = 18

    # =========================================================================
    # ARKUSZ 6: LISTA PRODUKTÓW (szczegółowa)
    # =========================================================================
    ws_products = wb.create_sheet("Lista produktów")

    ws_products.merge_cells('A1:O1')
    ws_products['A1'] = f"SZCZEGÓŁOWA LISTA PRODUKTÓW - {len(products)} pozycji"
    ws_products['A1'].font = Font(bold=True, size=14, color="2E7D32")
    ws_products['A1'].alignment = Alignment(horizontal="center")

    headers = ['Lp.', 'ID Produktu', 'Zamówienie', 'Nr Baselinker', 'Nr klienta',
               'Nazwa produktu', 'Status', 'Priorytet', 'Ilość', 'Gatunek',
               'Technologia', 'Klasa', 'Wymiary (cm)', 'Objętość (m³)', 'Klient']

    for col, header in enumerate(headers, 1):
        cell = ws_products.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, product in enumerate(products, 1):
        row = 3 + idx
        status_color = status_colors.get(product.current_status, 'FFFFFF')
        row_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

        data = [
            idx,
            product.short_product_id,
            product.order.internal_order_number if product.order else None,
            (product.order.baselinker_order_id if product.order else None) or '',
            (product.order.client_order_number if product.order else None) or '',
            (product.original_product_name[:50] + '...') if len(product.original_product_name or '') > 50 else (product.original_product_name or ''),
            _format_status(product.current_status),
            product.priority_rank or '-',
            product.quantity,
            (product.configuration.species if product.configuration else None) or '',
            (product.configuration.technology if product.configuration else None) or '',
            (product.configuration.wood_class if product.configuration else None) or '',
            f"{product.parsed_length_cm or 0}×{product.parsed_width_cm or 0}×{product.parsed_thickness_cm or 0}",
            float(product.volume_m3) if product.volume_m3 else 0,
            (product.order.client_name if product.order else None) or ''
        ]

        for col, value in enumerate(data, 1):
            cell = ws_products.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.fill = row_fill
            # Kolumny wyrównane do prawej: Lp.(1), Nr Baselinker(4), Priorytet(8), Ilość(9), Objętość(14)
            if col in [1, 4, 8, 9, 14]:
                cell.alignment = number_alignment
                if col == 14:
                    cell.number_format = '0.000'
            else:
                cell.alignment = cell_alignment

    column_widths = [5, 14, 12, 14, 14, 40, 18, 10, 8, 12, 14, 8, 18, 12, 25]
    for col, width in enumerate(column_widths, 1):
        ws_products.column_dimensions[get_column_letter(col)].width = width

    # =========================================================================
    # ZAPISZ I ZWRÓĆ
    # =========================================================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Export Excel wygenerowany (pełny raport)", extra={
        'user_id': current_user.id,
        'products_count': len(products),
        'sheets': ['Podsumowanie', 'Gatunki', 'Technologie', 'Klasy', 'Grubości', 'Lista produktów']
    })

    from flask import Response
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = f'attachment; filename=raport_produkcji_{timestamp}.xlsx'
    return response



def _export_pdf(products, timestamp, report_type='full'):
    """Eksport do PDF — czytelny raport produkcji"""
    import io
    import os
    from collections import defaultdict

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, PageBreak, KeepTogether)
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        logger.error("reportlab niedostępne")
        return jsonify({'success': False, 'error': 'Biblioteka PDF niedostępna'}), 501

    # Font z polskimi znakami
    font_name = 'Helvetica'
    for path in ['/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                 '/usr/share/fonts/dejavu/DejaVuSans.ttf',
                 'C:/Windows/Fonts/arial.ttf']:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('DejaVuSans', path))
                font_name = 'DejaVuSans'
                break
            except Exception:
                pass

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            rightMargin=1.2*cm, leftMargin=1.2*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)

    # -- Kolory i style --
    C_PRIMARY = colors.HexColor('#1a1a2e')
    C_HEADER = colors.HexColor('#2c5530')
    C_SECTION = colors.HexColor('#334155')
    C_BORDER = colors.HexColor('#e2e5ea')
    C_ALT_ROW = colors.HexColor('#f8f9fa')
    C_MUTED = colors.HexColor('#64748b')

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('T', parent=styles['Heading1'], fontName=font_name,
                             fontSize=18, textColor=C_PRIMARY, alignment=TA_LEFT, spaceAfter=2)
    s_subtitle = ParagraphStyle('ST', parent=styles['Normal'], fontName=font_name,
                                fontSize=9, textColor=C_MUTED, spaceAfter=14)
    s_section = ParagraphStyle('S', parent=styles['Heading2'], fontName=font_name,
                               fontSize=11, textColor=C_SECTION, spaceBefore=14, spaceAfter=6,
                               borderWidth=0, leading=14)
    s_footer = ParagraphStyle('F', parent=styles['Normal'], fontName=font_name,
                              fontSize=7, textColor=C_MUTED, alignment=TA_RIGHT)

    def _table_style(header_color=C_HEADER):
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('LINEBELOW', (0, 0), (-1, 0), 1, header_color),
            ('LINEBELOW', (0, 1), (-1, -2), 0.3, C_BORDER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, C_ALT_ROW]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ])

    # -- Dane analityczne --
    total_volume = sum(float(p.volume_m3 or 0) * (p.quantity or 1) for p in products)
    total_qty = sum(p.quantity or 1 for p in products)
    total_value = sum(float(p.total_value_net or 0) for p in products)

    buckets = {
        'status': defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0}),
        'species': defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0}),
        'technology': defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0}),
        'wood_class': defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0}),
        'thickness': defaultdict(lambda: {'count': 0, 'qty': 0, 'volume': 0.0}),
    }

    for p in products:
        vol = float(p.volume_m3 or 0)
        qty = p.quantity or 1
        for key, field in [('status', p.current_status),
                           ('species', (p.configuration.species if p.configuration else None) or 'Nieokreslony'),
                           ('technology', (p.configuration.technology if p.configuration else None) or 'Nieokreslona'),
                           ('wood_class', (p.configuration.wood_class if p.configuration else None) or 'Nieokreslona'),
                           ('thickness', f"{float(p.parsed_thickness_cm):.1f} cm" if p.parsed_thickness_cm else 'Nieokreslona')]:
            buckets[key][field]['count'] += 1
            buckets[key][field]['qty'] += qty
            buckets[key][field]['volume'] += vol * qty

    elements = []

    # ===================== STRONA 1: PODSUMOWANIE =====================
    elements.append(Paragraph("Raport produkcji — WoodPower", s_title))
    elements.append(Paragraph(
        f"{datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
        f"{len(products)} pozycji  |  {total_qty} szt.  |  "
        f"{total_volume:.3f} m\u00b3  |  "
        f"{total_value:,.2f} PLN".replace(',', ' '),
        s_subtitle))

    # -- KPI --
    kpi_data = [
        ['Pozycje', 'Sztuki', 'Objetosc m\u00b3', 'Wartosc netto'],
        [str(len(products)), str(total_qty), f"{total_volume:.3f}", f"{total_value:,.2f} PLN".replace(',', ' ')]
    ]
    kpi = Table(kpi_data, colWidths=[6*cm, 5*cm, 5*cm, 6.5*cm])
    kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), C_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, 0), 1, C_HEADER),
    ]))
    elements.append(kpi)
    elements.append(Spacer(1, 0.6*cm))

    # -- Statusy --
    elements.append(Paragraph("Rozklad wedlug stanowisk", s_section))
    rows = [['Stanowisko', 'Pozycje', 'Sztuki', 'Objetosc m\u00b3', '% obj.']]
    for status, d in sorted(buckets['status'].items(), key=lambda x: -x[1]['volume']):
        pct = (d['volume'] / total_volume * 100) if total_volume else 0
        rows.append([_format_status(status), str(d['count']), str(d['qty']),
                     f"{d['volume']:.3f}", f"{pct:.1f}%"])
    t = Table(rows, colWidths=[6.5*cm, 3*cm, 3*cm, 4.5*cm, 3*cm])
    t.setStyle(_table_style())
    elements.append(KeepTogether([t]))
    elements.append(Spacer(1, 0.4*cm))

    # -- Helper dla tabel analitycznych --
    def _analysis_section(title, data_dict, label='Kategoria'):
        s_rows = [[label, 'Poz.', 'Szt.', 'Obj. m\u00b3', '% obj.']]
        for name, d in sorted(data_dict.items(), key=lambda x: -x[1]['volume']):
            pct = (d['volume'] / total_volume * 100) if total_volume else 0
            s_rows.append([name[:25], str(d['count']), str(d['qty']),
                          f"{d['volume']:.3f}", f"{pct:.1f}%"])
        tbl = Table(s_rows, colWidths=[6.5*cm, 2.5*cm, 2.5*cm, 4*cm, 3*cm])
        tbl.setStyle(_table_style(C_SECTION))
        return KeepTogether([Paragraph(title, s_section), tbl, Spacer(1, 0.3*cm)])

    elements.append(_analysis_section("Gatunki drewna", buckets['species'], 'Gatunek'))
    elements.append(_analysis_section("Technologie", buckets['technology'], 'Technologia'))
    elements.append(_analysis_section("Klasy drewna", buckets['wood_class'], 'Klasa'))
    elements.append(_analysis_section("Grubosci", buckets['thickness'], 'Grubosc'))

    # ===================== STRONA 2+: LISTA PRODUKTÓW =====================
    elements.append(PageBreak())
    elements.append(Paragraph("Lista produktow", s_section))

    hdr = ['Lp.', 'ID', 'Zamowienie', 'Nazwa', 'Status', 'Prio.', 'Szt.', 'Gatunek', 'Wymiary', 'Obj. m\u00b3']
    tdata = [hdr]
    for i, p in enumerate(products, 1):
        name = (p.original_product_name or '')[:35]
        if len(p.original_product_name or '') > 35:
            name += '...'
        dims = f"{p.parsed_length_cm or 0}x{p.parsed_width_cm or 0}x{p.parsed_thickness_cm or 0}"
        tdata.append([
            str(i), p.short_product_id or '', (p.order.internal_order_number if p.order else None) or '',
            Paragraph(name, ParagraphStyle('cell', fontName=font_name, fontSize=7, leading=9)),
            _format_status(p.current_status),
            str(p.priority_rank) if p.priority_rank else '-',
            str(p.quantity or 1),
            ((p.configuration.species if p.configuration else None) or '')[:12],
            dims,
            f"{float(p.volume_m3 or 0):.3f}"
        ])

    cw = [0.9*cm, 2.2*cm, 2.2*cm, 5.5*cm, 2.8*cm, 1*cm, 0.9*cm, 2.2*cm, 3.2*cm, 1.8*cm]
    mt = Table(tdata, colWidths=cw, repeatRows=1, splitInRow=1)
    mt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), C_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (5, 0), (6, -1), 'CENTER'),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('LINEBELOW', (0, 0), (-1, 0), 1, C_HEADER),
        ('LINEBELOW', (0, 1), (-1, -2), 0.3, C_BORDER),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, C_ALT_ROW]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(mt)

    # Stopka
    elements.append(Spacer(1, 0.8*cm))
    elements.append(Paragraph(
        f"WoodPower CRM | {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}",
        s_footer))

    doc.build(elements)
    output.seek(0)

    from flask import Response
    response = Response(output.getvalue(), mimetype='application/pdf')
    response.headers['Content-Disposition'] = f'attachment; filename=raport_produkcji_{timestamp}.pdf'
    return response



@api_bp.route('/products/<int:product_id>/history', methods=['GET'])
@login_required
def product_history(product_id):
    """
    GET /production/api/products/<id>/history

    Zwraca autorów domknięcia stanowisk (flow) i scaloną oś czasu (history).
    Modal robi jedno żądanie i zasila z niego obie sekcje.
    """
    try:
        from ...services.product_history_service import build_product_history

        if not ProductionItem.query.get(product_id):
            return jsonify({'success': False, 'error': 'Nie znaleziono produktu'}), 404

        return jsonify({'success': True, **build_product_history(product_id)}), 200

    except Exception as e:
        logger.error("API: Błąd pobierania historii produktu", extra={
            'product_id': product_id,
            'error': str(e),
        })
        return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/products/filters-data', methods=['GET'])
@login_required
def get_filters_data():
    """
    GET /production/api/products/filters-data
    
    Zwraca unikalne wartości dla dropdownów filtrów
    
    Returns: JSON z listami unikalnych wartości
    """
    try:
        from ...models import ProductionItem
        from sqlalchemy import func, distinct
        
        # Pobierz unikalne wartości dla filtrów
        
        # Statusy - z enum w modelu
        statuses = [
            {'value': 'czeka_na_wyciecie', 'label': 'Czeka na wycięcie'},
            {'value': 'czeka_na_skladanie', 'label': 'Czeka na składanie'},
            {'value': 'czeka_na_pakowanie', 'label': 'Czeka na pakowanie'},
            {'value': 'spakowane', 'label': 'Spakowane'},
            {'value': 'wstrzymane', 'label': 'Wstrzymane'}
        ]
        
        # Gatunki drewna
        wood_species_query = db.session.query(distinct(ProductionConfiguration.species))\
                                      .filter(ProductionConfiguration.species.isnot(None))\
                                      .filter(ProductionConfiguration.species != 'unknown')\
                                      .all()
        wood_species = [{'value': item[0], 'label': item[0]} for item in wood_species_query]

        # Technologie
        technology_query = db.session.query(distinct(ProductionConfiguration.technology))\
                                    .filter(ProductionConfiguration.technology.isnot(None))\
                                    .filter(ProductionConfiguration.technology != 'unknown')\
                                    .all()
        technologies = [{'value': item[0], 'label': item[0]} for item in technology_query]

        # Klasy drewna
        wood_class_query = db.session.query(distinct(ProductionConfiguration.wood_class))\
                                    .filter(ProductionConfiguration.wood_class.isnot(None))\
                                    .filter(ProductionConfiguration.wood_class != 'unknown')\
                                    .all()
        wood_classes = [{'value': item[0], 'label': item[0]} for item in wood_class_query]
        
        # Zakres priorytetów
        priority_stats = db.session.query(
            func.min(ProductionItem.priority_rank),
            func.max(ProductionItem.priority_rank),
            func.avg(ProductionItem.priority_rank)
        ).filter(ProductionItem.priority_rank.isnot(None)).first()
        
        priority_range = {
            'min': int(priority_stats[0]) if priority_stats[0] else 0,
            'max': int(priority_stats[1]) if priority_stats[1] else 200,
            'avg': int(priority_stats[2]) if priority_stats[2] else 100
        }
        
        # Ostatnie 30 dni dla date picker
        from datetime import datetime, timedelta
        date_suggestions = {
            'today': datetime.now().date().isoformat(),
            'yesterday': (datetime.now() - timedelta(days=1)).date().isoformat(),
            'week_ago': (datetime.now() - timedelta(days=7)).date().isoformat(),
            'month_ago': (datetime.now() - timedelta(days=30)).date().isoformat()
        }
        
        filters_data = {
            'statuses': statuses,
            'wood_species': wood_species,
            'technologies': technologies,
            'wood_classes': wood_classes,
            'priority_range': priority_range,
            'date_suggestions': date_suggestions,
            'total_products': ProductionItem.query.count()
        }
        
        logger.info("Pobrano dane filtrów", extra={
            'user_id': current_user.id,
            'wood_species_count': len(wood_species),
            'technologies_count': len(technologies),
            'total_products': filters_data['total_products']
        })
        
        return jsonify({
            'success': True,
            'filters_data': filters_data
        })
        
    except Exception as e:
        logger.error("Błąd pobierania danych filtrów", extra={
            'user_id': current_user.id if current_user.is_authenticated else None,
            'error': str(e)
        })
        return jsonify({
            'success': False,
            'error': f'Błąd pobierania filtrów: {str(e)}'
        }), 500


# 5. UPDATE PRIORYTETU POJEDYNCZEGO PRODUKTU

@api_bp.route('/products/<int:product_id>/priority', methods=['PUT'])
@login_required
def update_single_product_priority(product_id):
    """
    PUT /production/api/products/<id>/priority
    
    Aktualizuje priorytet pojedynczego produktu - ZMODYFIKOWANY DLA priority_rank
    
    Body (JSON):
    {
        "priority": 5  // priority_rank (1,2,3,4... gdzie 1 = najwyższy)
    }
    
    Returns: JSON z rezultatem
    """
    try:
        data = request.get_json()
        if not data or 'priority' not in data:
            return jsonify({'success': False, 'error': 'Wymagany parametr: priority'}), 400
        
        new_priority = data['priority']
        
        # Walidacja priority_rank (1,2,3,4...)
        if not isinstance(new_priority, int) or new_priority < 1:
            return jsonify({'success': False, 'error': 'Priority rank musi być liczbą >= 1'}), 400
        
        from ...models import ProductionItem
        
        product = ProductionItem.query.get_or_404(product_id)
        old_priority = product.priority_rank
        
        # Używaj metody lock_priority() z modelu zamiast bezpośredniego ustawienia
        product.lock_priority(new_priority)
        db.session.commit()
        
        logger.info("Zaktualizowano priorytet produktu", extra={
            'user_id': current_user.id,
            'product_id': product_id,
            'product_short_id': product.short_product_id,
            'old_priority_rank': old_priority,
            'new_priority_rank': new_priority
        })
        
        return jsonify({
            'success': True,
            'message': 'Priorytet zaktualizowany',
            'product_id': product_id,
            'old_priority_rank': old_priority,
            'new_priority_rank': new_priority
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error("Błąd aktualizacji priorytetu", extra={
            'user_id': current_user.id if current_user.is_authenticated else None,
            'product_id': product_id,
            'error': str(e)
        })
        return jsonify({
            'success': False,
            'error': f'Błąd aktualizacji priorytetu: {str(e)}'
        }), 500



def _serialize_production_item(item, today=None):
    """
    Serializuje pojedynczy ProductionItem do słownika używanego przez endpointy listy i szczegółów.

    Współdzielone między /products-filtered (wiele rekordów) i /products/<id>/details (jeden rekord)
    żeby oba endpointy zwracały spójny kształt danych (w szczególności pola stacji gluing/formatting/finishing,
    które modal szczegółów sprawdza przez hasOwnProperty).
    """
    if today is None:
        today = date.today()

    days_to_deadline = None
    deadline_date = getattr(item, 'deadline_date', None)
    if deadline_date:
        days_to_deadline = (deadline_date - today).days

    product_data = {
        'id': item.id,
        'short_product_id': getattr(item, 'short_product_id', f'ID-{item.id}'),
        'internal_order_number': (item.order.internal_order_number if item.order else None) or '',
        'product_name': getattr(item, 'original_product_name', getattr(item, 'product_name', 'Brak nazwy')),
        'original_product_name': getattr(item, 'original_product_name', getattr(item, 'product_name', 'Brak nazwy')),
        'client_name': (item.order.client_name if item.order else None) or '',
        'current_status': getattr(item, 'current_status', 'unknown'),
        'priority_rank': getattr(item, 'priority_rank', None),
        'volume_m3': float(getattr(item, 'volume_m3', 0) or 0),
        'total_value_net': float(getattr(item, 'total_value_net', getattr(item, 'total_value', 0)) or 0),
        'total_value': float(getattr(item, 'total_value_net', getattr(item, 'total_value', 0)) or 0),
        'order_date': getattr(item, 'order_date', getattr(item, 'created_at', None)),
        'deadline_date': deadline_date.isoformat() if deadline_date else None,
        'days_to_deadline': days_to_deadline,
        'days_until_deadline': days_to_deadline,
        'is_overdue': days_to_deadline < 0 if days_to_deadline is not None else False,
        'baselinker_order_id': item.order.baselinker_order_id if item.order else None,
        'created_at': getattr(item, 'created_at', None),
        'attachment_file_name': item.order.attachment_file_name if item.order else None,
        'attachment_file_url': item.order.attachment_file_url if item.order else None,
        'parsed_wood_species': item.configuration.species if item.configuration else None,
        'parsed_technology': item.configuration.technology if item.configuration else None,
        'parsed_wood_class': item.configuration.wood_class if item.configuration else None,
        'parsed_finish_state': getattr(item, 'parsed_finish_state', None),
        'cut_to_size': bool(getattr(item, 'cut_to_size', True)),
        'shape': getattr(item, 'shape', None),
        'parsed_width_cm': getattr(item, 'parsed_width_cm', None),
        'parsed_length_cm': getattr(item, 'parsed_length_cm', None),
        'parsed_thickness_cm': getattr(item, 'parsed_thickness_cm', None),
        'client_email': (item.order.client_email if item.order else None) or '',
        'client_phone': (item.order.client_phone if item.order else None) or '',
        'product_sequence_in_order': getattr(item, 'product_sequence_in_order', None),
        'total_products_in_order': db.session.query(db.func.count(ProductionItem.id)).filter(
            ProductionItem.order_id == item.order_id
        ).scalar() if item.order_id else None,
        # Przepływ produkcji - wszystkie 6 stanowisk
        'cutting_started_at': getattr(item, 'cutting_started_at', None),
        'cutting_completed_at': getattr(item, 'cutting_completed_at', None),
        'cutting_duration_minutes': getattr(item, 'cutting_duration_minutes', None),
        'assembly_started_at': getattr(item, 'assembly_started_at', None),
        'assembly_completed_at': getattr(item, 'assembly_completed_at', None),
        'assembly_duration_minutes': getattr(item, 'assembly_duration_minutes', None),
        'gluing_started_at': getattr(item, 'gluing_started_at', None),
        'gluing_completed_at': getattr(item, 'gluing_completed_at', None),
        'gluing_duration_minutes': getattr(item, 'gluing_duration_minutes', None),
        'formatting_started_at': getattr(item, 'formatting_started_at', None),
        'formatting_completed_at': getattr(item, 'formatting_completed_at', None),
        'formatting_duration_minutes': getattr(item, 'formatting_duration_minutes', None),
        'finishing_started_at': getattr(item, 'finishing_started_at', None),
        'finishing_completed_at': getattr(item, 'finishing_completed_at', None),
        'finishing_duration_minutes': getattr(item, 'finishing_duration_minutes', None),
        'packaging_started_at': getattr(item, 'packaging_started_at', None),
        'packaging_completed_at': getattr(item, 'packaging_completed_at', None),
        'packaging_duration_minutes': getattr(item, 'packaging_duration_minutes', None),
        'logistics_completed_at': item.order.logistics_completed_at.isoformat() if item.order and item.order.logistics_completed_at else None,
        'override_delivery_method': item.order.override_delivery_method if item.order else None,
        'is_personal_pickup': item.order.is_personal_pickup if item.order else False,
        'quantity': getattr(item, 'quantity', 1),
        'quantity_done_cutting': getattr(item, 'quantity_done_cutting', 0),
        'quantity_done_assembly': getattr(item, 'quantity_done_assembly', 0),
        'quantity_done_gluing': getattr(item, 'quantity_done_gluing', 0),
        'quantity_done_formatting': getattr(item, 'quantity_done_formatting', 0),
        'quantity_done_finishing': getattr(item, 'quantity_done_finishing', 0),
        'quantity_done_packaging': getattr(item, 'quantity_done_packaging', 0),
        'client_order_number': item.order.client_order_number if item.order else None,
        'quote_number': item.order.quote_number if item.order else None,
        'unit_price_net': float(getattr(item, 'unit_price_net', 0) or 0),
        'updated_at': item.updated_at.isoformat() if getattr(item, 'updated_at', None) else None,
        'sync_source': item.order.sync_source if item.order else None,
        'delivery_address': (item.order.delivery_address if item.order else None) or '',
        'order_notes': item.order.order_notes if item.order else None,
    }

    if product_data['order_date'] and hasattr(product_data['order_date'], 'isoformat'):
        product_data['order_date'] = product_data['order_date'].isoformat()
    if product_data['created_at'] and hasattr(product_data['created_at'], 'isoformat'):
        product_data['created_at'] = product_data['created_at'].isoformat()

    station_fields = ['cutting', 'assembly', 'gluing', 'formatting', 'finishing', 'packaging']
    for station in station_fields:
        started_field = f'{station}_started_at'
        completed_field = f'{station}_completed_at'
        if product_data.get(started_field) and hasattr(product_data[started_field], 'isoformat'):
            product_data[started_field] = product_data[started_field].isoformat()
        if product_data.get(completed_field) and hasattr(product_data[completed_field], 'isoformat'):
            product_data[completed_field] = product_data[completed_field].isoformat()

    return product_data


@api_bp.route('/products-filtered', methods=['GET', 'POST'])
@login_required
def products_filtered():
    """
    API endpoint dla filtrowania produktów - ZMODYFIKOWANY DLA priority_rank
    """
    try:
        # Pobierz parametry filtrów
        status_filter = request.args.get('status', 'all')
        search_query = request.args.get('search', '').strip()
        page = int(request.args.get('page', 1))
        per_page = min(int(request.args.get('per_page', 50)), 200)
        # ZMIANA: domyślnie sortuj po priority_rank zamiast priority_score
        sort_by = request.args.get('sort_by', 'priority_rank')
        # ZMIANA: dla priority_rank używamy ASC (1,2,3...)
        default_order = 'asc' if sort_by == 'priority_rank' else 'desc'
        sort_order = request.args.get('sort_order', default_order)
        
        # Rozpocznij query od wszystkich produktów — join ProductionOrder dla pól order-level
        query = ProductionItem.query.join(ProductionOrder)

        # Filtrowanie po statusie
        if status_filter and status_filter != 'all':
            query = query.filter(ProductionItem.current_status == status_filter)

        # Wyszukiwanie - bez zmian
        if search_query:
            search_pattern = f"%{search_query}%"
            search_conditions = []

            if hasattr(ProductionItem, 'original_product_name'):
                search_conditions.append(ProductionItem.original_product_name.ilike(search_pattern))
            if hasattr(ProductionItem, 'short_product_id'):
                search_conditions.append(ProductionItem.short_product_id.ilike(search_pattern))
            # order-level fields via ProductionOrder (join already present)
            search_conditions.append(ProductionOrder.internal_order_number.ilike(search_pattern))
            search_conditions.append(ProductionOrder.client_name.ilike(search_pattern))
            # Wyszukiwanie po numerze zamówienia klienta (np. "1617/2025")
            search_conditions.append(ProductionOrder.client_order_number.ilike(search_pattern))
            # Wyszukiwanie po numerze wyceny (np. "226/04/26/W")
            search_conditions.append(ProductionOrder.quote_number.ilike(search_pattern))
            # Wyszukiwanie po numerze Baselinker — Integer, cast do String
            search_conditions.append(cast(ProductionOrder.baselinker_order_id, String).ilike(search_pattern))

            if search_conditions:
                query = query.filter(or_(*search_conditions))
        
        # ZMIANA: Sortowanie - priority_rank jako główne sortowanie
        sort_column = None
        if sort_by == 'priority_rank' and hasattr(ProductionItem, 'priority_rank'):
            sort_column = ProductionItem.priority_rank
        elif sort_by == 'deadline_date' and hasattr(ProductionItem, 'deadline_date'):
            sort_column = ProductionItem.deadline_date
        elif sort_by == 'created_at' and hasattr(ProductionItem, 'created_at'):
            sort_column = ProductionItem.created_at
        elif sort_by == 'short_product_id' and hasattr(ProductionItem, 'short_product_id'):
            sort_column = ProductionItem.short_product_id
        
        if sort_column is not None:
            if sort_order == 'desc':
                query = query.order_by(sort_column.desc())
            else:
                query = query.order_by(sort_column.asc())
        else:
            # ZMIANA: domyślne sortowanie po priority_rank ASC zamiast ID
            if hasattr(ProductionItem, 'priority_rank'):
                query = query.order_by(ProductionItem.priority_rank.asc())
            else:
                query = query.order_by(ProductionItem.id.desc())
        
        # Paginacja - bez zmian
        paginated = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        # Przygotuj dane produktów - delegacja do wspólnego helpera
        today = date.today()
        products_data = [_serialize_production_item(item, today) for item in paginated.items]
        
        # ZMIANA: Statystyki oparte na priority_rank
        stats = {
            'total_filtered': paginated.total,
            'total_volume': sum(float(p['volume_m3']) * p.get('quantity', 1) for p in products_data),
            'total_value': sum(float(p['total_value']) for p in products_data),
            # ZMIANA: avg_priority_rank zamiast avg_priority_score
            'avg_priority_rank': sum(float(p['priority_rank'] or 0) for p in products_data) / len(products_data) if products_data else 0,
            'avg_priority': sum(float(p['priority_rank'] or 0) for p in products_data) / len(products_data) if products_data else 0,
            'overdue_count': len([p for p in products_data if p['days_to_deadline'] is not None and p['days_to_deadline'] < 0])
        }
        
        # Informacje o paginacji - bez zmian
        pagination_info = {
            'page': paginated.page,
            'per_page': paginated.per_page,
            'total': paginated.total,
            'pages': paginated.pages,
            'has_prev': paginated.has_prev,
            'has_next': paginated.has_next,
            'prev_num': paginated.prev_num,
            'next_num': paginated.next_num
        }
        
        return jsonify({
            'success': True,
            'products': products_data,
            'pagination': pagination_info,
            'stats': stats,
            'filters_applied': {
                'status': status_filter,
                'search': search_query,
                'sort_by': sort_by,
                'sort_order': sort_order
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Błąd filtrowania produktów: {str(e)}',
            'products': [],
            'pagination': {'page': 1, 'pages': 1, 'total': 0},
            'stats': {},
            'traceback': traceback.format_exc()
        }), 500


@api_bp.route('/products/<int:product_id>/details', methods=['GET'])
@login_required
def product_details(product_id):
    """
    Zwraca pełne dane pojedynczego produktu (te same pola co /products-filtered).

    Używane przez modal szczegółów w products-module.js - zastępuje wcześniejsze
    pobieranie całej strony listy (domyślnie 50 rekordów), które gubiło produkty
    poza pierwszą stroną i nadpisywało cache frontendu.
    """
    try:
        item = ProductionItem.query.get(product_id)
        if not item:
            return jsonify({
                'success': False,
                'error': f'Nie znaleziono produktu o ID {product_id}'
            }), 404

        return jsonify({
            'success': True,
            'product': _serialize_production_item(item)
        })
    except Exception as e:
        logger.exception('Błąd pobierania szczegółów produktu', extra={'product_id': product_id})
        return jsonify({
            'success': False,
            'error': f'Błąd pobierania szczegółów produktu: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500


@api_bp.route('/update-priority', methods=['POST'])
@login_required
def update_priority():
    """
    API endpoint dla aktualizacji priorytetów produktów (priority_rank)

    Formaty:
    - Batch: {"products": [{"id": 1, "priority_rank": 3}, ...]}
    - Single: {"product_id": 1, "priority_rank": 3}
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Brak danych JSON'}), 400
        
        updated_products = []
        
        if 'products' in data:
            # ZMIANA: Batch update dla drag & drop (używa priority_rank)
            products_data = data.get('products', [])
            
            for product_data in products_data:
                product_id = product_data.get('id')
                new_priority_rank = product_data.get('priority_rank')

                if product_id is None or new_priority_rank is None:
                    continue

                product = ProductionItem.query.get(product_id)
                if product:
                    product.priority_rank = new_priority_rank
                    product.priority_manual_override = True  # Drag&drop = manual
                    updated_products.append({
                        'id': product_id,
                        'new_priority_rank': new_priority_rank
                    })
        
        elif 'product_id' in data:
            product_id = data.get('product_id')
            new_priority_rank = data.get('priority_rank')

            if product_id is None or new_priority_rank is None:
                return jsonify({'success': False, 'error': 'Wymagane: product_id i priority_rank'}), 400

            product = ProductionItem.query.get(product_id)
            if not product:
                return jsonify({'success': False, 'error': f'Produkt {product_id} nie znaleziony'}), 404

            product.priority_rank = new_priority_rank
            product.priority_manual_override = True
            updated_products.append({'id': product_id, 'new_priority_rank': new_priority_rank})
        
        else:
            return jsonify({'success': False, 'error': 'Wymagane: product_id+priority_rank LUB products'}), 400
        
        # Zapisz zmiany
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Zaktualizowano priorytety {len(updated_products)} produktów',
            'updated_count': len(updated_products),
            'updated_products': updated_products
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Błąd aktualizacji priorytetów: {str(e)}'
        }), 500

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================



# ============================================================================
# ERROR HANDLERS
# ============================================================================


@api_bp.route('/products/<int:product_id>/notes', methods=['PUT'])
@login_required
def update_product_notes(product_id):
    """
    PUT /production/api/products/<id>/notes
    
    Aktualizuje notatki produkcyjne dla produktu
    
    Body JSON:
    {
        "notes": "Nowa treść notatki"
    }
    
    Returns: JSON z rezultatem operacji
    """
    try:
        from ...models import ProductionItem
        
        # Pobierz dane z request
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False, 
                'error': 'Brak danych JSON'
            }), 400
        
        new_notes = data.get('notes', '').strip()
        
        # Znajdź produkt
        product = ProductionItem.query.get_or_404(product_id)
        
        # Zapisz stare notatki dla logowania
        old_notes = product.production_notes or ''
        
        # Aktualizuj notatki
        product.production_notes = new_notes
        product.updated_at = get_local_now()
        
        db.session.commit()
        
        logger.info("API: Zaktualizowano notatki produktu", extra={
            'product_id': product_id,
            'short_product_id': product.short_product_id,
            'user_id': current_user.id,
            'old_notes_length': len(old_notes),
            'new_notes_length': len(new_notes),
            'client_ip': request.remote_addr
        })
        
        return jsonify({
            'success': True,
            'message': 'Notatki zostały zaktualizowane',
            'data': {
                'product_id': product_id,
                'short_product_id': product.short_product_id,
                'notes': new_notes,
                'updated_at': product.updated_at.isoformat()
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error("API: Błąd aktualizacji notatek", extra={
            'product_id': product_id,
            'user_id': current_user.id,
            'error': str(e),
            'client_ip': request.remote_addr
        })
        
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@api_bp.route('/products/<int:product_id>/order-products', methods=['GET'])
@login_required  
def get_order_products(product_id):
    """
    GET /production/api/products/<id>/order-products
    
    Zwraca wszystkie produkty z tego samego zamówienia Baselinker
    dla paska produktów w modal
    
    Returns: JSON z listą produktów z zamówienia
    """
    try:
        from ...models import ProductionItem
        
        # Znajdź produkt
        product = ProductionItem.query.get_or_404(product_id)

        product_baselinker_id = product.order.baselinker_order_id if product.order else None
        if not product_baselinker_id:
            return jsonify({
                'success': True,
                'products': [_format_product_for_navigation(product)],
                'total_count': 1,
                'current_product_id': product_id
            }), 200

        # Znajdź wszystkie produkty z tego zamówienia
        order_products = ProductionItem.query.join(ProductionOrder).filter(
            ProductionOrder.baselinker_order_id == product_baselinker_id
        ).order_by(ProductionItem.product_sequence_in_order.asc()).all()

        # Formatuj produkty dla nawigacji
        formatted_products = []
        for p in order_products:
            formatted_products.append(_format_product_for_navigation(p))

        logger.info("API: Pobrano produkty zamówienia", extra={
            'product_id': product_id,
            'baselinker_order_id': product_baselinker_id,
            'products_count': len(formatted_products),
            'user_id': current_user.id
        })

        return jsonify({
            'success': True,
            'products': formatted_products,
            'total_count': len(formatted_products),
            'current_product_id': product_id,
            'baselinker_order_id': product_baselinker_id
        }), 200
        
    except Exception as e:
        logger.error("API: Błąd pobierania produktów zamówienia", extra={
            'product_id': product_id,
            'user_id': current_user.id,
            'error': str(e)
        })
        
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/order-timeline/<int:baselinker_order_id>', methods=['GET'])
@login_required
def get_order_timeline(baselinker_order_id):
    """
    GET /production/api/order-timeline/<baselinker_order_id>

    Zwraca linię czasu stanowisk produkcyjnych i status zamówienia
    dla bloku 'Zamówienie' w modalu szczegółów wyceny.
    """
    try:
        from sqlalchemy.orm import joinedload
        from ...models import ProductionProduct, ProductionOrder
        from ...services.order_timeline_service import (
            build_timeline_payload, order_status_badge,
        )

        products = (ProductionProduct.query
                    .join(ProductionOrder)
                    .options(joinedload(ProductionProduct.configuration))
                    .filter(ProductionOrder.baselinker_order_id == baselinker_order_id)
                    .order_by(ProductionProduct.product_sequence_in_order.asc())
                    .all())

        active = [p for p in products if p.current_status != 'anulowane']
        if not active:
            return jsonify({'success': True, 'has_data': False}), 200

        return jsonify({
            'success': True,
            'has_data': True,
            'order_status': order_status_badge(products),
            'stations': build_timeline_payload(products),
        }), 200

    except Exception as e:
        logger.error("API: Błąd budowy linii czasu zamówienia", extra={
            'user_id': current_user.id,
            'baselinker_order_id': baselinker_order_id,
            'error': str(e),
        })
        return jsonify({'success': False, 'has_data': False, 'error': str(e)}), 500


def _format_product_for_navigation(product):
    """
    Formatuje produkt dla nawigacji w modal (pasek produktów)
    
    Format: [Gatunek] [Technologia] [Klasa] [Wymiary] [Wykończenie]
    """
    # Buduj specyfikację produktu
    spec_parts = []
    
    # Gatunek drewna
    if product.configuration and product.configuration.species:
        spec_parts.append(product.configuration.species)

    # Technologia
    if product.configuration and product.configuration.technology:
        spec_parts.append(product.configuration.technology)

    # Klasa drewna
    if product.configuration and product.configuration.wood_class:
        spec_parts.append(product.configuration.wood_class)
    
    # Wymiary
    dimensions = []
    if product.parsed_width_cm:
        dimensions.append(f"{int(product.parsed_width_cm)}")
    if product.parsed_thickness_cm:
        dimensions.append(f"{int(product.parsed_thickness_cm)}")
    if product.parsed_length_cm:
        dimensions.append(f"{int(product.parsed_length_cm)}")
    
    if dimensions:
        spec_parts.append("×".join(dimensions))
    
    # Wykończenie
    if product.parsed_finish_state:
        spec_parts.append(product.parsed_finish_state)
    
    # Połącz specyfikację
    specification = " ".join(spec_parts) if spec_parts else product.original_product_name
    
    # Wymiary jako osobne pole
    dim_parts = []
    if product.parsed_width_cm:
        dim_parts.append(f"{int(product.parsed_width_cm)}")
    if product.parsed_thickness_cm:
        dim_parts.append(f"{int(product.parsed_thickness_cm)}")
    if product.parsed_length_cm:
        dim_parts.append(f"{int(product.parsed_length_cm)}")
    dimensions_str = "×".join(dim_parts) + " cm" if dim_parts else ''

    return {
        'id': product.id,
        'short_product_id': product.short_product_id,
        'specification': specification,
        'dimensions': dimensions_str,
        'sequence': product.product_sequence_in_order,
        'status': product.current_status,
        'priority': product.priority_rank or 100
    }


# ============================================================================
# API ROUTERS - NOWE ENDPOINTY - ENHANCED PRIORITY SYSTEM
# ============================================================================


@api_bp.route('/recalculate-all-priorities', methods=['POST'])
@login_required
def reset_all_priorities():
    """
    POST /api/recalculate-all-priorities - Reset wszystkich priorytetów
    
    Endpoint dla admina do resetowania wszystkich priorytetów (przycisk w UI).
    Ustawia priority_manual_override = FALSE dla wszystkich produktów
    i wywołuje pełne przeliczenie priorytetów.
    
    Body (opcjonalny):
    {
        "confirm_reset": true  // Potwierdzenie operacji (wymagane)
    }
    
    Autoryzacja: admin
    Returns: JSON z szczegółowym raportem przeliczenia
    """
    try:
        data = request.get_json() or {}
        confirm_reset = data.get('confirm_reset', False)
        
        # Wymagaj potwierdzenia dla bezpieczeństwa
        if not confirm_reset:
            return jsonify({
                'success': False,
                'error': 'Wymagane potwierdzenie reset operacji (confirm_reset: true)'
            }), 400
        
        logger.info("API: Rozpoczęcie reset wszystkich priorytetów", extra={
            'user_id': current_user.id,
            'endpoint': 'recalculate-all-priorities',
            'client_ip': request.remote_addr
        })
        
        from ...services.priority_service import get_priority_calculator
        from ...models import ProductionItem
        
        # Sprawdź ile produktów ma manual override przed resetem
        manual_overrides_before = ProductionItem.query.filter_by(priority_manual_override=True).count()
        
        # Resetuj wszystkie manual overrides
        updated_count = db.session.query(ProductionItem)\
                                .filter_by(priority_manual_override=True)\
                                .update({'priority_manual_override': False})
        db.session.commit()
        
        logger.info("API: Zresetowano manual overrides", extra={
            'manual_overrides_reset': updated_count,
            'user_id': current_user.id
        })
        
        # Wywołaj pełne przeliczenie priorytetów
        priority_calculator = get_priority_calculator()
        calculation_result = priority_calculator.recalculate_all_priorities()
        
        if calculation_result.get('success'):
            logger.info("API: Reset priorytetów zakończony pomyślnie", extra={
                'user_id': current_user.id,
                'products_updated': calculation_result.get('products_updated', 0),
                'calculation_duration': calculation_result.get('calculation_duration', '00:00:00'),
                'manual_overrides_reset': updated_count
            })
            
            return jsonify({
                'success': True,
                'message': f'Zresetowano priorytety {calculation_result.get("products_updated", 0)} produktów',
                'data': {
                    'reset_performed_at': get_local_now().isoformat(),
                    'reset_by': current_user.id,
                    'manual_overrides_reset': updated_count,
                    'manual_overrides_before': manual_overrides_before,
                    
                    'priority_recalculation': {
                        'products_updated': calculation_result.get('products_updated', 0),
                        'calculation_duration': calculation_result.get('calculation_duration', '00:00:00'),
                        'weeks_processed': calculation_result.get('weeks_processed', 0),
                        'algorithm': 'payment_date_weekly_grouping'
                    },
                    
                    'statistics': calculation_result.get('statistics', {}),
                    'performance_metrics': calculation_result.get('performance_metrics', {})
                }
            }), 200
        else:
            # Rollback manual overrides jeśli przeliczenie nie powiodło się
            db.session.rollback()
            
            logger.error("API: Błąd przeliczenia po reset", extra={
                'user_id': current_user.id,
                'error': calculation_result.get('error', 'Unknown error')
            })
            
            return jsonify({
                'success': False,
                'error': f'Błąd przeliczenia priorytetów: {calculation_result.get("error", "Unknown error")}',
                'data': {
                    'reset_rolled_back': True,
                    'manual_overrides_preserved': manual_overrides_before
                }
            }), 500
        
    except Exception as e:
        db.session.rollback()
        logger.error("API: Błąd reset priorytetów", extra={
            'user_id': current_user.id,
            'error': str(e),
            'client_ip': request.remote_addr
        })
        
        return jsonify({
            'success': False,
            'error': str(e),
            'data': {
                'reset_rolled_back': True
            }
        }), 500



@api_bp.route('/products/<int:product_id>/set-manual-priority', methods=['POST'])
@login_required
def set_manual_product_priority(product_id):
    """
    POST /api/products/<id>/set-manual-priority - Ręczne ustawienie priorytetu
    
    Endpoint dla admina do ustawienia ręcznego priorytetu dla konkretnego produktu.
    Sprawdza czy numer nie jest zajęty i ustawia manual_override = TRUE.
    
    Body (JSON):
    {
        "priority_rank": 5,       // Wymagane: numer priorytetu (1-1000)
        "reason": "Pilne zlecenie" // Opcjonalne: powód zmiany
    }
    
    Autoryzacja: admin
    Returns: JSON z rezultatem operacji
    """
    try:
        data = request.get_json()
        if not data or 'priority_rank' not in data:
            return jsonify({
                'success': False,
                'error': 'Wymagany parametr: priority_rank (liczba 1-1000)'
            }), 400
        
        priority_rank = data.get('priority_rank')
        reason = data.get('reason', '').strip()
        
        # Walidacja priority_rank
        if not isinstance(priority_rank, int) or priority_rank < 1 or priority_rank > 1000:
            return jsonify({
                'success': False,
                'error': 'priority_rank musi być liczbą między 1 a 1000'
            }), 400
        
        from ...models import ProductionItem
        
        # Znajdź produkt
        product = ProductionItem.query.get_or_404(product_id)
        
        # Sprawdź czy rank nie jest już zajęty przez inny produkt
        existing_product = ProductionItem.query.filter(
            ProductionItem.priority_rank == priority_rank,
            ProductionItem.priority_manual_override == True,
            ProductionItem.id != product_id
        ).first()
        
        if existing_product:
            return jsonify({
                'success': False,
                'error': f'Priorytet {priority_rank} jest już zajęty przez produkt {existing_product.short_product_id}',
                'conflict_product': {
                    'id': existing_product.id,
                    'short_product_id': existing_product.short_product_id,
                    'product_name': existing_product.original_product_name
                }
            }), 409
        
        # Zapisz stare wartości dla logowania
        old_priority_rank = product.priority_rank
        old_manual_override = product.priority_manual_override
        
        # Ustaw nowy priorytet używając metody z modelu
        product.lock_priority(priority_rank)
        
        # Dodaj informację o powodzie zmiany (jeśli model to obsługuje)
        if hasattr(product, 'priority_change_reason') and reason:
            product.priority_change_reason = reason
        
        product.updated_at = get_local_now()
        db.session.commit()
        
        logger.info("API: Ustawiono ręczny priorytet produktu", extra={
            'user_id': current_user.id,
            'product_id': product_id,
            'product_short_id': product.short_product_id,
            'old_priority_rank': old_priority_rank,
            'new_priority_rank': priority_rank,
            'reason': reason,
            'client_ip': request.remote_addr
        })
        
        return jsonify({
            'success': True,
            'message': f'Ustawiono priorytet {priority_rank} dla produktu {product.short_product_id}',
            'data': {
                'product_id': product_id,
                'short_product_id': product.short_product_id,
                'old_priority': {
                    'rank': old_priority_rank,
                    'manual_override': old_manual_override
                },
                'new_priority': {
                    'rank': product.priority_rank,
                    'manual_override': product.priority_manual_override
                },
                'reason': reason,
                'set_by': current_user.id,
                'set_at': product.updated_at.isoformat()
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error("API: Błąd ustawienia ręcznego priorytetu", extra={
            'user_id': current_user.id,
            'product_id': product_id,
            'error': str(e),
            'client_ip': request.remote_addr
        })
        
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@api_bp.route('/priority-statistics', methods=['GET'])
@login_required  # Nie musi być admin - może wszyscy użytkownicy
def get_priority_statistics():
    """
    GET /api/priority-statistics - Statystyki systemu priorytetów
    
    Zwraca statystyki nowego systemu priorytetów dla UI i monitoringu.
    
    Query params:
        include_details: true|false - czy dołączyć szczegóły (default: false)
        
    Autoryzacja: login_required
    Returns: JSON ze statystykami priorytetów
    """
    try:
        include_details = request.args.get('include_details', 'false').lower() == 'true'
        
        logger.info("API: Pobieranie statystyk priorytetów", extra={
            'user_id': current_user.id,
            'include_details': include_details,
            'endpoint': 'priority-statistics'
        })
        
        from ...models import ProductionItem
        from sqlalchemy import func, desc
        from datetime import datetime, timedelta
        
        # Podstawowe statystyki
        total_products = ProductionItem.query.count()
        products_in_queue = ProductionItem.query.filter(
            ProductionItem.current_status.in_([
                'czeka_na_wyciecie', 'czeka_na_skladanie', 'czeka_na_pakowanie', 'w_realizacji'
            ])
        ).count()
        
        # Manual overrides
        manual_overrides_count = ProductionItem.query.filter_by(priority_manual_override=True).count()
        
        # Statystyki payment_date — pole order-level, query bezpośrednio na ProductionOrder
        payment_date_stats = db.session.query(
            func.count(ProductionOrder.id).label('total'),
            func.count(ProductionOrder.payment_date).label('with_payment_date'),
            func.min(ProductionOrder.payment_date).label('oldest_payment'),
            func.max(ProductionOrder.payment_date).label('newest_payment')
        ).first()

        # Rozkład po tygodniach
        weekly_distribution = []
        if include_details and payment_date_stats.with_payment_date > 0:
            # Grupuj po tygodniach — join ProductionItem żeby filtrować po current_status
            weekly_query = db.session.query(
                func.year(ProductionOrder.payment_date).label('year'),
                func.week(ProductionOrder.payment_date).label('week'),
                func.count(ProductionItem.id).label('count')
            ).join(ProductionItem, ProductionItem.order_id == ProductionOrder.id).filter(
                ProductionOrder.payment_date.isnot(None),
                ProductionItem.current_status.in_([
                    'czeka_na_wyciecie', 'czeka_na_skladanie', 'czeka_na_pakowanie', 'w_realizacji'
                ])
            ).group_by(
                func.year(ProductionOrder.payment_date),
                func.week(ProductionOrder.payment_date)
            ).order_by(
                func.year(ProductionOrder.payment_date).asc(),
                func.week(ProductionOrder.payment_date).asc()
            ).limit(10).all()
            
            for year, week, count in weekly_query:
                weekly_distribution.append({
                    'year': year,
                    'week': week,
                    'week_label': f'{year}-W{week:02d}',
                    'products_count': count
                })
        
        # Rozkład priority_rank
        priority_rank_stats = db.session.query(
            func.min(ProductionItem.priority_rank).label('min_rank'),
            func.max(ProductionItem.priority_rank).label('max_rank'),
            func.avg(ProductionItem.priority_rank).label('avg_rank'),
            func.count(ProductionItem.priority_rank).label('products_with_rank')
        ).filter(ProductionItem.priority_rank.isnot(None)).first()
        
        # Ostatnia aktualizacja priorytetów (przybliżone)
        last_priority_update = None
        try:
            # Szukaj ostatniego produktu z aktualną datą updated_at
            last_updated_product = ProductionItem.query.filter(
                ProductionItem.updated_at.isnot(None)
            ).order_by(ProductionItem.updated_at.desc()).first()
            
            if last_updated_product:
                last_priority_update = last_updated_product.updated_at.isoformat()
        except:
            pass
        
        # Przygotuj response
        statistics_data = {
            'system_overview': {
                'total_products': total_products,
                'products_in_queue': products_in_queue,
                'manual_overrides_count': manual_overrides_count,
                'manual_override_percentage': round((manual_overrides_count / max(total_products, 1)) * 100, 1),
                'algorithm': 'payment_date_weekly_grouping_v2'
            },
            
            'payment_date_coverage': {
                'total_products': payment_date_stats.total,
                'with_payment_date': payment_date_stats.with_payment_date,
                'coverage_percentage': round((payment_date_stats.with_payment_date / max(payment_date_stats.total, 1)) * 100, 1),
                'date_range': {
                    'oldest': payment_date_stats.oldest_payment.isoformat() if payment_date_stats.oldest_payment else None,
                    'newest': payment_date_stats.newest_payment.isoformat() if payment_date_stats.newest_payment else None
                }
            },
            
            'priority_ranking': {
                'min_rank': priority_rank_stats.min_rank,
                'max_rank': priority_rank_stats.max_rank,
                'avg_rank': round(priority_rank_stats.avg_rank, 1) if priority_rank_stats.avg_rank else None,
                'products_with_rank': priority_rank_stats.products_with_rank,
                'ranking_coverage': round((priority_rank_stats.products_with_rank / max(products_in_queue, 1)) * 100, 1)
            },

            'system_info': {
                'priority_system_version': '2.0_rank_only',
                'uses_priority_score': False,
                'uses_priority_rank': True,
                'supports_unlimited_products': True
            },
            
            'last_updated': last_priority_update,
            'statistics_generated_at': get_local_now().isoformat()
        }
        
        # Dodaj szczegóły jeśli requested
        if include_details:
            statistics_data['weekly_distribution'] = weekly_distribution
            
            # Top manual overrides
            manual_override_products = ProductionItem.query.filter_by(
                priority_manual_override=True
            ).order_by(ProductionItem.priority_rank.asc()).limit(10).all()
            
            statistics_data['manual_overrides_details'] = [
                {
                    'id': p.id,
                    'short_product_id': p.short_product_id,
                    'priority_rank': p.priority_rank,
                    'current_status': p.current_status,
                    'manual_override': p.priority_manual_override,
                    'updated_at': p.updated_at.isoformat() if p.updated_at else None
                }
                for p in manual_override_products
            ]
        
        logger.info("API: Statystyki priorytetów pobrane", extra={
            'user_id': current_user.id,
            'total_products': total_products,
            'products_in_queue': products_in_queue,
            'manual_overrides': manual_overrides_count,
            'include_details': include_details
        })
        
        return jsonify({
            'success': True,
            'statistics': statistics_data
        }), 200
        
    except Exception as e:
        logger.error("API: Błąd pobierania statystyk priorytetów", extra={
            'user_id': current_user.id,
            'error': str(e),
            'client_ip': request.remote_addr
        })
        
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    

@api_bp.route('/set-priority', methods=['POST'])
@admin_required
def set_product_priority():
    """
    POST /production/api/set-priority

    Ustawia lub usuwa flagę is_priority dla produktu/produktów.

    Request JSON:
        {
            "product_id": int,           # ID produktu (wymagane jeśli nie ma order_number)
            "order_number": str,         # Numer zamówienia (opcjonalne - dla całego zamówienia)
            "is_priority": bool,         # True = włącz, False = wyłącz
            "mode": "product" | "order"  # Tryb: pojedynczy produkt lub całe zamówienie
        }

    Returns:
        JSON: Status operacji i lista zaktualizowanych produktów
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'success': False,
                'error': 'Brak danych w żądaniu'
            }), 400

        is_priority = data.get('is_priority', True)
        mode = data.get('mode', 'product')
        product_id = data.get('product_id')
        order_number = data.get('order_number')

        updated_products = []

        if mode == 'order' and order_number:
            # Aktualizuj wszystkie produkty w zamówieniu
            products = (
                ProductionItem.query
                .join(ProductionOrder, ProductionItem.order_id == ProductionOrder.id)
                .filter(ProductionOrder.internal_order_number == order_number)
                .all()
            )

            if not products:
                return jsonify({
                    'success': False,
                    'error': f'Nie znaleziono produktów dla zamówienia {order_number}'
                }), 404

            for product in products:
                product.is_priority = is_priority
                product.updated_at = get_local_now()
                updated_products.append({
                    'id': product.id,
                    'short_product_id': product.short_product_id,
                    'is_priority': product.is_priority
                })

            db.session.commit()

            logger.info("Ustawiono priorytet dla zamówienia", extra={
                'order_number': order_number,
                'is_priority': is_priority,
                'products_count': len(products),
                'user_id': current_user.id
            })

        elif product_id:
            # Aktualizuj pojedynczy produkt
            product = ProductionItem.query.get(product_id)

            if not product:
                return jsonify({
                    'success': False,
                    'error': f'Nie znaleziono produktu o ID {product_id}'
                }), 404

            product.is_priority = is_priority
            product.updated_at = get_local_now()

            db.session.commit()

            updated_products.append({
                'id': product.id,
                'short_product_id': product.short_product_id,
                'is_priority': product.is_priority
            })

            logger.info("Ustawiono priorytet dla produktu", extra={
                'product_id': product_id,
                'short_product_id': product.short_product_id,
                'is_priority': is_priority,
                'user_id': current_user.id
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Wymagane jest product_id lub order_number z mode=order'
            }), 400

        return jsonify({
            'success': True,
            'message': f'Zaktualizowano priorytet dla {len(updated_products)} produktów',
            'updated_products': updated_products,
            'is_priority': is_priority
        })

    except Exception as e:
        db.session.rollback()
        logger.error("Błąd ustawiania priorytetu", extra={
            'error': str(e),
            'traceback': traceback.format_exc()
        })

        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500



@api_bp.route('/get-order-products-count/<order_number>')
@login_required
def get_order_products_count(order_number: str):
    """
    GET /production/api/get-order-products-count/<order_number>

    Pobiera liczbę produktów w zamówieniu (używane do decyzji o wyświetlaniu tooltipa).

    Returns:
        JSON: Liczba produktów w zamówieniu
    """
    try:
        count = (
            ProductionItem.query
            .join(ProductionOrder, ProductionItem.order_id == ProductionOrder.id)
            .filter(ProductionOrder.internal_order_number == order_number)
            .count()
        )

        return jsonify({
            'success': True,
            'order_number': order_number,
            'products_count': count
        })

    except Exception as e:
        logger.error("Błąd pobierania liczby produktów", extra={
            'order_number': order_number,
            'error': str(e)
        })

        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500
