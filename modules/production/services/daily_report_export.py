# -*- coding: utf-8 -*-
"""
Eksport dziennego raportu produkcji do XLSX.

Warstwa czysto prezentacyjna: przyjmuje dict z daily_report_service.zbierz_dane()
i zwraca bajty pliku. Zero SQL, zero Flaska — dzięki temu układ arkusza da się
testować bez bazy, na ręcznie złożonym diccie.

TRZY ZASADY, KTÓRE WYGLĄDAJĄ NA DROBIAZG, A NIE SĄ:

1. Pusta komórka to NIE zero. `None` znaczy „nie dotyczy" (trakownia nie ma
   kolejki, pracownik bez atrybucji nie ma tempa), zero znaczy „policzone
   i wyszło zero". openpyxl zapisuje None jako pustą komórkę i o to chodzi.
   Kolorowanie tła NIE może tego zatrzeć — komórka z wypełnieniem, ale bez
   wartości, dalej jest pusta.

2. W nagłówkach piszemy „Różnica", nigdy Δ. Arkusz trafia do księgowości,
   która nie musi znać notacji technicznej. Wymaganie właściciela projektu,
   pilnowane testem.

3. Kolor rozdziela, nie ocenia. Nagłówki, bloki i sumy mają własne tło, ale
   żadna liczba nie jest podświetlana za „zły" wynik — raport ma pokazywać,
   co zrobiono i co zostało, bez stawiania diagnoz. Decyzja właściciela.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .station_catalog import STATION_LABELS

# ── Paleta firmowa ──────────────────────────────────────────────────────────
#
# Dwa kolory marki WoodPower, wzięte z CSS aplikacji (--primary-color
# i --secondary-color, m.in. modules/partner_academy/static/css/recruitment.css):
# pomarańcz #ED6B24 i grafit #1F2020. Reszta to rozjaśnienia pomarańczu na
# bieli — 10% i 20% krycia — więc paleta trzyma się jednego odcienia zamiast
# dobierać kolory „na oko".
#
# Nagłówki mają grafitowe tło, a nie pomarańczowe: biały tekst na #ED6B24 daje
# kontrast ok. 2,8:1, czyli poniżej progu czytelności. Pomarańcz pracuje jako
# akcent (tytuł, krawędzie, tło bloków), gdzie kontrast nie jest krytyczny.
# Całość jest rozróżnialna także po wydruku w skali szarości.
_GRAFIT = '1F2020'          # --secondary-color
_POMARANCZ = 'ED6B24'       # --primary-color
_POMARANCZ_10 = 'FDF0E9'    # 10% pomarańczu na bieli — nagłówki bloków
_POMARANCZ_20 = 'FBE1D3'    # 20% — wiersz SUMA
_BEZ = 'F2F0ED'             # trakownia: inny pion niż hala, więc neutralny
_SZARY_TEKST = '666666'     # --text-gray, wiersz „Nieprzypisane"

_TLO_NAGLOWKA = PatternFill('solid', fgColor=_GRAFIT)
_TLO_BLOKU = PatternFill('solid', fgColor=_POMARANCZ_10)
_TLO_SUMY = PatternFill('solid', fgColor=_POMARANCZ_20)
_TLO_TRAKOWNI = PatternFill('solid', fgColor=_BEZ)

_FONT_NAGLOWKA = Font(bold=True, color='FFFFFF')
_FONT_TYTULU = Font(bold=True, size=14, color=_POMARANCZ)
_FONT_BLOKU = Font(bold=True, color=_GRAFIT)
_FONT_NIEPRZYPISANE = Font(italic=True, color=_SZARY_TEKST)

_KRESKA_POMARANCZ = Side(style='thin', color=_POMARANCZ)
_OBRAMOWANIE_SUMY = Border(top=_KRESKA_POMARANCZ)

# ── Formaty liczbowe ────────────────────────────────────────────────────────
#
# Kody formatu zapisuje się ZAWSZE w notacji en-US — przecinek jako separator
# tysięcy, kropka jako dziesiętny — bo tak wymaga OOXML. Excel podmienia je przy
# wyświetlaniu na separatory z ustawień systemu, więc w polskiej instalacji
# '#,##0.00' pokaże się jako „89 817,51". Zapisanie tu polskiej notacji
# ('# ##0,00') dałoby format, którego Excel nie rozpozna.
_FMT_SZTUKI = '#,##0'
_FMT_M3 = '#,##0.000'
_FMT_ZL = '#,##0.00'
_FMT_DZIESIETNY = '#,##0.0'

NAGLOWKI_STANOWISKA = (
    'Stanowisko', 'Zamówienia', 'Sztuki', 'm³', 'Wartość netto', 'Cofnięcia',
    'W kolejce (szt.)', 'W kolejce (m³)',
)

_KLUCZE_STANOWISKA = (
    'etykieta', 'zamowienia', 'sztuki', 'm3', 'wartosc_netto', 'cofniecia',
    'kolejka_szt', 'kolejka_m3',
)

# Kolumna (1-indeks) → format liczbowy. Kolumna 1 to tekst, więc jej nie ma.
_FORMATY_STANOWISKA = {
    2: _FMT_SZTUKI, 3: _FMT_SZTUKI, 4: _FMT_M3, 5: _FMT_ZL,
    6: _FMT_SZTUKI, 7: _FMT_SZTUKI, 8: _FMT_M3,
}

# Klucze z _KLUCZE_STANOWISKA, które trzeba zaokrąglić, i do ilu miejsc.
# m3/kolejka_m3/wartosc_netto to surowe wyniki dzielenia zmiennoprzecinkowego
# (get_station_work_per_day: total_value_net * delta / quantity) — bez tego
# komórka pojedynczego stanowiska pokazuje więcej cyfr niż wiersz SUMA pod
# spodem, który JEST zaokrąglany. Ten sam poziom precyzji co w wierszu SUMA
# i w wierszu trakowni, żeby wszystkie trzy się zgadzały.
_ZAOKRAGLENIA_STANOWISKA = {'m3': 3, 'wartosc_netto': 2, 'kolejka_m3': 3}

NAGLOWKI_LUDZIE = (
    'Pracownik', 'Stanowiska', 'Sztuki (wkład)', 'm³ (wkład)',
    'Zdarzenia', 'Godziny', 'm³/h',
)

_KLUCZE_LUDZIE = (
    'nazwa', 'stanowiska', 'sztuki', 'm3', 'zdarzenia', 'godziny', 'tempo',
)

# „Sztuki (wkład)" bywają ułamkowe: przy dwóch osobach na jednej sztuce każda
# dostaje udział 0,5 — stąd format dziesiętny, nie całkowity.
_FORMATY_LUDZIE = {
    3: _FMT_DZIESIETNY, 4: _FMT_M3, 5: _FMT_SZTUKI,
    6: _FMT_DZIESIETNY, 7: _FMT_M3,
}

# Etykiety koszyków terminów — kolejność od najpilniejszego.
_ETYKIETY_TERMINOW = (
    ('po_terminie', 'Po terminie'),
    ('dzis', 'Termin dziś'),
    ('1_2_dni', 'Termin za 1–2 dni'),
    ('3_7_dni', 'Termin za 3–7 dni'),
    ('8_dni_plus', 'Termin za 8+ dni'),
    ('bez_terminu', 'Bez terminu'),
)

# Przypis pod tabelą stanowisk. Bez niego suma kolumny „Zamówienia" (91 dla
# przykładowego dnia) nie zgadza się z liczbą w wierszu SUMA (83) i wygląda to
# na błąd arkusza, a jest jedyną poprawną odpowiedzią — to samo zamówienie
# przechodzi tego dnia przez kilka stanowisk.
_PRZYPIS_ZAMOWIENIA = (
    'Zamówienia w wierszu SUMA to liczba unikalnych zamówień, a nie suma '
    'kolumny — jedno zamówienie przechodzi zwykle przez kilka stanowisk '
    'tego samego dnia i w każdym z nich liczy się osobno.'
)

# Granice szerokości kolumny. Dolna, żeby wąskie kolumny liczbowe nie zlewały
# się z nagłówkiem; górna, bo w arkuszu „Ludzie" kolumna stanowisk potrafi mieć
# „Lakiernia, Pakowanie, Wykańczanie" i bez limitu rozpycha arkusz na pół ekranu.
_MIN_SZEROKOSC = 11
_MAKS_SZEROKOSC = 42


def nazwa_pliku(dzien):
    """
    Nazwa załącznika — czysto ASCII.

    Polskie znaki w nagłówku Content-Disposition rwą odpowiedź (WSGI koduje
    nagłówki w latin-1); w repo jest na to osobny test regresyjny.
    """
    return f'Raport_produkcji_{dzien.isoformat()}.xlsx'


def _dlugosc_wyswietlana(komorka):
    """
    Ile znaków zajmie komórka po sformatowaniu.

    openpyxl nie ma auto-fitu, bo szerokość liczy Excel dopiero przy otwarciu
    pliku — musimy oszacować ją sami. Dla liczb z formatem nie wystarczy
    len(str(wartosc)): 89817.51 to osiem znaków w Pythonie, ale „89 817,51"
    to dziewięć na ekranie. Separator tysięcy dokłada po jednym znaku na
    każde trzy cyfry części całkowitej.
    """
    wartosc = komorka.value
    if wartosc is None:
        return 0
    if isinstance(wartosc, bool) or not isinstance(wartosc, (int, float)):
        return len(str(wartosc))

    # Miejsca po kropce w kodzie formatu — w notacji OOXML kropka jest
    # separatorem dziesiętnym, a przecinek tysięcznym (patrz komentarz przy
    # _FMT_*), więc liczymy po kropce, nie po przecinku.
    format_ = komorka.number_format or ''
    miejsca = len(format_.split('.')[-1]) if '.' in format_ else 0

    calkowita = f'{abs(int(wartosc)):d}'
    znaki = len(calkowita) + (len(calkowita) - 1) // 3   # separatory tysięcy
    if miejsca:
        znaki += 1 + miejsca                             # przecinek + ułamek
    if wartosc < 0:
        znaki += 1
    return znaki


def _autoszerokosc(ws):
    """
    Szerokość każdej kolumny dobrana do najdłuższej komórki w niej.

    Zamiast stałej szerokości opartej wyłącznie na nagłówku: „Sylwester Rębiś"
    jest dłuższy niż „Pracownik", a „89 817,51" dłuższe niż „Wartość netto".
    """
    scalone = {zakres.min_row for zakres in ws.merged_cells.ranges}
    for kolumna in ws.columns:
        najdluzsza = max(
            (_dlugosc_wyswietlana(k) for k in kolumna if k.row not in scalone),
            default=0)
        litera = kolumna[0].column_letter
        ws.column_dimensions[litera].width = min(
            _MAKS_SZEROKOSC, max(_MIN_SZEROKOSC, najdluzsza + 2))


def _stylizuj_naglowek(ws):
    """Pierwszy wiersz: grafitowe tło, biały tekst, zamrożony przy przewijaniu."""
    for cell in ws[1]:
        cell.font = _FONT_NAGLOWKA
        cell.fill = _TLO_NAGLOWKA
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'


def _formatuj_wiersz(ws, numer_wiersza, formaty):
    """Nadaje format liczbowy kolumnom wskazanym w mapie i wyrównuje je do prawej."""
    for kolumna, format_ in formaty.items():
        komorka = ws.cell(row=numer_wiersza, column=kolumna)
        komorka.number_format = format_
        komorka.alignment = Alignment(horizontal='right')


def _zaokr(wartosc, miejsca):
    """
    Zaokrąglenie zachowujące None.

    round(None, 2) rzuca TypeError, a pusta komórka niesie tu znaczenie:
    „nie dotyczy" to nie to samo co „policzone i wyszło zero".
    """
    return None if wartosc is None else round(wartosc, miejsca)


def _arkusz_dzien(ws, dane):
    """
    Podsumowanie w układzie pionowym etykieta → wartość.

    Pionowo, bo bloki mają różne jednostki: sztuki, m³, złotówki, procenty
    i godziny nie zmieszczą się sensownie w jednej siatce kolumn. Każda pozycja
    niesie własny format liczbowy, bo w jednej kolumnie B sąsiadują sztuki
    i złotówki.
    """
    ws.append(['Raport produkcji', dane['dzien'].strftime('%d.%m.%Y')])
    ws.cell(row=1, column=1).font = _FONT_TYTULU
    ws.cell(row=1, column=2).font = Font(bold=True, size=14, color=_GRAFIT)
    ws.append([])

    wykonanie = dane['wykonanie']
    ludzie = dane['ludzie']
    trakownia = dane['trakownia']

    bloki = (
        ('WYKONANIE', (
            # „netto" wymagało tłumaczenia za każdym razem — działanie w nawiasie
            # mówi to samo bez żargonu.
            ('Sztuki (wykonane − cofnięte)', wykonanie['sztuki'], _FMT_SZTUKI),
            ('m³', round(wykonanie['m3'], 3), _FMT_M3),
            ('Wartość netto (zł)', round(wykonanie['wartosc_netto'], 2), _FMT_ZL),
            ('Pozycji dotkniętych', wykonanie['pozycje'], _FMT_SZTUKI),
            ('Zamówień dotkniętych', wykonanie['zamowienia'], _FMT_SZTUKI),
            ('Cofnięcia (szt.)', wykonanie['cofniecia'], _FMT_SZTUKI),
        )),
        ('LUDZIE', (
            ('Pracowników z pracą', ludzie['osoby'], _FMT_SZTUKI),
            ('Osobogodziny', ludzie['godziny'], _FMT_DZIESIETNY),
            ('Pokrycie atrybucją (%)', ludzie['pokrycie_proc'], _FMT_DZIESIETNY),
        )),
        ('TRAKOWNIA', (
            ('Kłody', trakownia['klody'], _FMT_SZTUKI),
            ('m³', round(trakownia['m3'], 3), _FMT_M3),
        )),
        ('ZOSTAŁO', tuple(
            [('W kolejce (szt.)',
              sum(s['kolejka_szt'] or 0 for s in dane['stanowiska']), _FMT_SZTUKI),
             ('W kolejce (m³)',
              round(sum(s['kolejka_m3'] or 0 for s in dane['stanowiska']), 3), _FMT_M3)]
            + [(etykieta, dane['terminy'][klucz], _FMT_SZTUKI)
               for klucz, etykieta in _ETYKIETY_TERMINOW]
        )),
    )

    for tytul, pozycje in bloki:
        ws.append([tytul, None])
        for kolumna in (1, 2):
            komorka = ws.cell(row=ws.max_row, column=kolumna)
            komorka.font = _FONT_BLOKU
            komorka.fill = _TLO_BLOKU
        for etykieta, wartosc, format_ in pozycje:
            ws.append([etykieta, wartosc])
            _formatuj_wiersz(ws, ws.max_row, {2: format_})
        ws.append([])

    _autoszerokosc(ws)


def _arkusz_stanowiska(ws, dane):
    """
    Wiersz na stanowisko produkcyjne, potem trakownia, na końcu SUMA.

    Trakownia jest tu wierszem arkusza, ale NIE jest stanowiskiem w modelu:
    agregat zwraca ją osobnym blokiem, bo nie ma statusów kolejki ani wartości
    sprzedaży. Wiersz składamy tutaj, na poziomie prezentacji — dzięki temu
    agregat nie musi udawać, że hala ma osiem stanowisk. Osobne tło podkreśla,
    że to inny pion, a nie ósme stanowisko hali.
    """
    ws.append(list(NAGLOWKI_STANOWISKA))
    _stylizuj_naglowek(ws)

    for wiersz in dane['stanowiska']:
        ws.append([
            _zaokr(wiersz.get(klucz), _ZAOKRAGLENIA_STANOWISKA[klucz])
            if klucz in _ZAOKRAGLENIA_STANOWISKA else wiersz.get(klucz)
            for klucz in _KLUCZE_STANOWISKA
        ])
        _formatuj_wiersz(ws, ws.max_row, _FORMATY_STANOWISKA)

    # Kłody w kolumnie „Sztuki", m³ w swojej. Kolumna zamówień, wartość netto
    # i obie kolumny kolejki zostają PUSTE — trakownia ich nie ma i zero byłoby
    # kłamstwem. Trakownia ma własne zlecenia (prod_sawmill_orders), ale to
    # inny byt niż zamówienie klienta z BaseLinkera; wstawienie ich w tę samą
    # kolumnę sugerowałoby, że da się je zsumować.
    trakownia = dane['trakownia']
    ws.append([STATION_LABELS['sawmill'], None, trakownia['klody'],
               round(trakownia['m3'], 3), None, None, None, None])
    _formatuj_wiersz(ws, ws.max_row, _FORMATY_STANOWISKA)
    for cell in ws[ws.max_row]:
        cell.fill = _TLO_TRAKOWNI

    # Zamówienia biorą się z globalnego licznika, a NIE z sumy kolumny wyżej:
    # jedno zamówienie zwykle przechodzi tego dnia przez kilka stanowisk, więc
    # zsumowanie wierszy policzyłoby je wielokrotnie. Suma kolumny „Zamówienia"
    # jest z definicji większa lub równa tej liczbie i to jest poprawne.
    ws.append([
        'SUMA',
        dane['wykonanie']['zamowienia'],
        sum(w['sztuki'] for w in dane['stanowiska']),
        round(sum(w['m3'] for w in dane['stanowiska']), 3),
        round(sum(w['wartosc_netto'] or 0 for w in dane['stanowiska']), 2),
        sum(w['cofniecia'] for w in dane['stanowiska']),
        sum(w['kolejka_szt'] or 0 for w in dane['stanowiska']),
        round(sum(w['kolejka_m3'] or 0 for w in dane['stanowiska']), 3),
    ])
    _formatuj_wiersz(ws, ws.max_row, _FORMATY_STANOWISKA)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color=_GRAFIT)
        cell.fill = _TLO_SUMY
        cell.border = _OBRAMOWANIE_SUMY

    # Przypis pod tabelą. Scalony przez szerokość arkusza i z zawijaniem, żeby
    # nie rozpychał kolumny A — autoszerokość świadomie go pomija (patrz
    # _autoszerokosc), inaczej pierwsza kolumna miałaby szerokość zdania.
    ws.append([])
    ws.append([_PRZYPIS_ZAMOWIENIA])
    wiersz_przypisu = ws.max_row
    ws.merge_cells(start_row=wiersz_przypisu, start_column=1,
                   end_row=wiersz_przypisu, end_column=len(NAGLOWKI_STANOWISKA))
    komorka = ws.cell(row=wiersz_przypisu, column=1)
    komorka.font = Font(italic=True, size=9, color=_SZARY_TEKST)
    komorka.alignment = Alignment(horizontal='left', vertical='top',
                                  wrap_text=True)
    ws.row_dimensions[wiersz_przypisu].height = 28

    _autoszerokosc(ws)


def _arkusz_ludzie(ws, dane):
    ws.append(list(NAGLOWKI_LUDZIE))
    _stylizuj_naglowek(ws)

    for wiersz in dane['ludzie']['wiersze']:
        ws.append([wiersz.get(klucz) for klucz in _KLUCZE_LUDZIE])
        _formatuj_wiersz(ws, ws.max_row, _FORMATY_LUDZIE)

    # Wiersz obowiązkowy, także gdy wynosi zero: bez niego suma tego arkusza
    # nie zgadza się z sumą arkusza „Stanowiska" i raport wygląda na zepsuty.
    nieprzypisane = dane['ludzie']['nieprzypisane']
    ws.append(['Nieprzypisane', '', nieprzypisane['sztuki'], nieprzypisane['m3'],
               None, None, None])
    _formatuj_wiersz(ws, ws.max_row, _FORMATY_LUDZIE)
    for cell in ws[ws.max_row]:
        cell.font = _FONT_NIEPRZYPISANE

    _autoszerokosc(ws)


def build_daily_xlsx(dane):
    """
    Buduje trzyarkuszowy raport z dictu zbierz_dane(). Zwraca bajty pliku.

    Args:
        dane: dict o kontrakcie daily_report_service.zbierz_dane()

    Returns:
        bytes: zawartość pliku .xlsx
    """
    wb = Workbook()

    ws_dzien = wb.active
    ws_dzien.title = 'Dzień'
    _arkusz_dzien(ws_dzien, dane)

    _arkusz_stanowiska(wb.create_sheet('Stanowiska'), dane)
    _arkusz_ludzie(wb.create_sheet('Ludzie'), dane)

    strumien = io.BytesIO()
    wb.save(strumien)
    return strumien.getvalue()
